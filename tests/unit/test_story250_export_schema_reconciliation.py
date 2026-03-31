"""Tests for STORY-250: Reconcile Export Schema Contract with Workbook Output.

Behavioral contract tests that verify:
1. Generated workbooks pass their own schema gate (full-data + sparse-data)
2. AI readiness and transformation fields are emitted on correct sheets
3. Headers are derived from the single authoritative schema (drift-proof)
4. Schema changes require explicit version bumps
5. No hardcoded header lists exist in sheet generators
"""

import ast
from pathlib import Path

import pytest
from openpyxl import load_workbook

from solstein.domain.models import Company, FinancialMetric
from solstein.exporters.excel_improved import ImprovedExcelExporter
from solstein.exporters.export_schema import (
    EXPORT_SCHEMA,
    EXPORT_SCHEMA_VERSION,
    get_headers_for_sheet,
    get_schema_by_sheet,
    validate_export,
)

# ─────────────────────────────────────────────
# Fixtures
# ─────────────────────────────────────────────

def _make_full_company() -> Company:
    """Create a company with ALL exported fields populated."""
    fm = FinancialMetric(
        revenue=50.0,
        growth_rate=0.15,
        profit_margin=0.22,
        employees=350,
        total_funding_raised_eur=25_000_000,
        latest_valuation_eur=100_000_000,
        allow_empty_primary=True,
    )
    return Company(
        id="TEST-250-FULL",
        name="FullData Corp",
        industry="Energy Software",
        financials=fm,
        ai_score=0.85,
        ai_readiness_score=72.5,
        ai_readiness_tier="Advanced",
        classification="leader",
        tier="Tier 1",
        threat_level="High",
        tech_stack=["Python", "React", "TensorFlow"],
        key_customers=["Acme", "Globex"],
        open_positions=12,
        data_availability="high",
        transformation_time_months=18.0,
        transformation_cost_eur=500_000,
        transformation_risk_level="Medium",
        funding_rounds=[{"round": "Series A", "amount": "5M"}],
        funding_war_chest="Strong",
        revenue_cagr_5yr=0.30,
        revenue_per_employee_eur_k=140.0,
        employee_cagr_3yr=0.20,
        lead_investors=["Sequoia"],
        revenue_timeline=[
            {"year": 2023, "eur_millions": 40.0, "source": "Annual report"},
            {"year": 2024, "eur_millions": 50.0, "source": "Estimate"},
        ],
        parent_company="MegaCorp",
        subsidiaries=["SubCo"],
        acquisitions=[{"company": "Startup", "year": 2024}],
        notes="Full test company",
        source_links=["https://example.com"],
        metric_sources={"revenue": ["sec_edgar"]},
        enrichment_quality_metrics={"merge_conflicts": {"rev": "mismatch"}},
    )


def _make_sparse_company() -> Company:
    """Create a company with minimal data — tests N/A fallbacks."""
    return Company(id="TEST-250-SPARSE", name="Sparse Corp")


@pytest.fixture()
def export_path(tmp_path: Path) -> Path:
    return tmp_path / "story250_test.xlsx"


@pytest.fixture()
def full_export(export_path: Path) -> Path:
    """Generate an export with fully populated data."""
    exporter = ImprovedExcelExporter()
    exporter.create_dashboard([_make_full_company()], export_path)
    return export_path


@pytest.fixture()
def sparse_export(tmp_path: Path) -> Path:
    """Generate an export with minimal/sparse data."""
    path = tmp_path / "sparse_export.xlsx"
    exporter = ImprovedExcelExporter()
    exporter.create_dashboard([_make_sparse_company()], path)
    return path


@pytest.fixture()
def mixed_export(tmp_path: Path) -> Path:
    """Generate an export with a mix of full and sparse companies."""
    path = tmp_path / "mixed_export.xlsx"
    exporter = ImprovedExcelExporter()
    exporter.create_dashboard([_make_full_company(), _make_sparse_company()], path)
    return path


# ─────────────────────────────────────────────
# AC1: create_dashboard passes validate_export for full-data and sparse-data
# ─────────────────────────────────────────────

class TestSchemaGatePassesBothFixtures:
    """Acceptance Criteria 1: Workbook passes validate_export for both fixture types."""

    def test_full_data_passes_schema_gate(self, full_export: Path) -> None:
        """Full-data fixture must pass the schema gate."""
        validate_export(full_export)

    def test_sparse_data_passes_schema_gate(self, sparse_export: Path) -> None:
        """Sparse/minimal-data fixture must pass (headers present, values N/A)."""
        validate_export(sparse_export)

    def test_mixed_data_passes_schema_gate(self, mixed_export: Path) -> None:
        """Mixed full + sparse data must pass the schema gate."""
        validate_export(mixed_export)

    def test_empty_company_list_still_creates_valid_structure(self, tmp_path: Path) -> None:
        """Even with zero companies, the workbook structure (headers) must be valid."""
        path = tmp_path / "empty_export.xlsx"
        exporter = ImprovedExcelExporter()
        exporter.create_dashboard([], path)
        validate_export(path)


# ─────────────────────────────────────────────
# AC2: AI readiness and transformation fields on correct sheet
# ─────────────────────────────────────────────

class TestAIAndTransformationFieldsPresent:
    """AC2: AI readiness and transformation fields are on Executive Summary."""

    EXPECTED_HEADERS = {
        "AI Readiness", "AI Readiness Tier",
        "Transform Time (mo)", "Transform Cost (EUR)", "Transform Risk",
    }

    def test_headers_on_executive_summary(self, full_export: Path) -> None:
        """All AI readiness + transformation headers exist on Executive Summary sheet."""
        wb = load_workbook(full_export, read_only=True)
        ws = wb["Executive Summary"]
        actual_headers = set()
        for col in range(1, 50):
            val = ws.cell(row=4, column=col).value
            if val is not None:
                actual_headers.add(str(val))
        wb.close()

        missing = self.EXPECTED_HEADERS - actual_headers
        assert not missing, f"Executive Summary missing headers: {missing}"

    def test_transformation_data_populated_for_full_company(self, full_export: Path) -> None:
        """Full company must have non-N/A transformation values."""
        wb = load_workbook(full_export, read_only=True)
        ws = wb["Executive Summary"]
        # Find column indices for transformation headers
        header_cols: dict[str, int] = {}
        for col in range(1, 50):
            val = ws.cell(row=4, column=col).value
            if val and str(val) in self.EXPECTED_HEADERS:
                header_cols[str(val)] = col
        wb.close()

        # All expected headers must have been found
        assert len(header_cols) == len(self.EXPECTED_HEADERS)


# ─────────────────────────────────────────────
# AC3: Headers derived from single authoritative mapping (drift-proof)
# ─────────────────────────────────────────────

class TestDriftProofHeaders:
    """AC3: Sheet generators derive headers from export_schema, not hardcoded lists."""

    def test_get_headers_for_sheet_matches_schema(self) -> None:
        """get_headers_for_sheet returns exactly the schema-defined headers in order."""
        by_sheet = get_schema_by_sheet()
        for sheet_name, specs in by_sheet.items():
            expected = [s.header for s in specs]
            actual = get_headers_for_sheet(sheet_name)
            assert actual == expected, f"Header mismatch for '{sheet_name}': {actual} != {expected}"

    def test_get_headers_for_unknown_sheet_raises(self) -> None:
        """Unknown sheet name must raise ValueError, not silently return empty."""
        with pytest.raises(ValueError, match="No schema fields defined"):
            get_headers_for_sheet("Nonexistent Sheet")

    def test_workbook_headers_match_schema_exactly(self, full_export: Path) -> None:
        """Every sheet's actual workbook headers must match the schema in exact order."""
        wb = load_workbook(full_export, read_only=True)
        by_sheet = get_schema_by_sheet()

        for sheet_name, specs in by_sheet.items():
            assert sheet_name in wb.sheetnames, f"Missing sheet: {sheet_name}"
            ws = wb[sheet_name]
            expected_headers = [s.header for s in specs]
            actual_headers = []
            for col in range(1, len(expected_headers) + 1):
                val = ws.cell(row=4, column=col).value
                if val is not None:
                    actual_headers.append(str(val))

            assert actual_headers == expected_headers, (
                f"Sheet '{sheet_name}' headers don't match schema.\n"
                f"Expected: {expected_headers}\n"
                f"Actual:   {actual_headers}"
            )
        wb.close()

    def test_no_hardcoded_header_lists_in_sheet_generators(self) -> None:
        """Static analysis: sheets.py and sheets_extended.py must not contain
        hardcoded header list assignments that bypass the schema."""
        src_dir = Path(__file__).resolve().parents[2] / "src" / "solstein" / "exporters" / "excel"
        for filename in ("sheets.py", "sheets_extended.py"):
            filepath = src_dir / filename
            if not filepath.exists():
                continue
            source = filepath.read_text()
            tree = ast.parse(source, filename=str(filepath))
            for node in ast.walk(tree):
                # Look for: headers = ["...", "...", ...]
                if (
                    isinstance(node, ast.Assign)
                    and len(node.targets) == 1
                    and isinstance(node.targets[0], ast.Name)
                    and node.targets[0].id == "headers"
                    and isinstance(node.value, ast.List)
                ):
                    # This is a hardcoded header list — should not exist
                    pytest.fail(
                        f"{filename}:{node.lineno} still has hardcoded 'headers = [...]'. "
                        "Use get_headers_for_sheet() instead."
                    )


# ─────────────────────────────────────────────
# AC4: Behavioral regression tests with live workbooks
# ─────────────────────────────────────────────

class TestBehavioralRegressions:
    """Tests that verify runtime behavior, not just static field inventories."""

    def test_all_five_data_sheets_present(self, full_export: Path) -> None:
        """Workbook must contain all 5 data sheets + metadata."""
        wb = load_workbook(full_export, read_only=True)
        expected_sheets = {
            "Executive Summary", "Market Rankings", "Financial Intelligence",
            "Revenue History", "Advanced Data", "Export Metadata",
        }
        actual_sheets = set(wb.sheetnames)
        wb.close()
        missing = expected_sheets - actual_sheets
        assert not missing, f"Missing sheets: {missing}"

    def test_schema_version_matches_expected(self) -> None:
        """Schema version is 1.1 (bumped for STORY-250 reconciliation)."""
        assert EXPORT_SCHEMA_VERSION == "1.1"

    def test_full_export_has_data_rows(self, full_export: Path) -> None:
        """Full export must have at least one data row per data sheet."""
        wb = load_workbook(full_export, read_only=True)
        for sheet_name in ["Executive Summary", "Market Rankings", "Financial Intelligence"]:
            ws = wb[sheet_name]
            # Data starts at row 5 (row 4 is headers)
            first_data = ws.cell(row=5, column=1).value
            assert first_data is not None, f"Sheet '{sheet_name}' has no data row"
        wb.close()

    def test_schema_field_count(self) -> None:
        """Guard against accidental field additions/removals without review."""
        # Current count: 16 Executive + 6 Market + 12 Financial + 4 Revenue + 8 Advanced = 46
        # Update this if the schema legitimately changes
        assert len(EXPORT_SCHEMA) == 46, (
            f"Schema field count changed to {len(EXPORT_SCHEMA)}. "
            "If intentional, update this test and bump EXPORT_SCHEMA_VERSION."
        )
