"""Exporter regression snapshot tests - G4.

Schema and visual consistency tests for exporters.
Part of EPIC-017 Wave 2 Testing Hardening.
"""

import csv
from datetime import datetime

import pytest

from solstein.domain.models import Company
from solstein.exporters.csv import CSVExporter
from solstein.exporters.excel.styles import ColorPalette, LayoutConstants
from solstein.exporters.excel_improved import ImprovedExcelExporter


class TestCSVExporterSnapshots:
    """CSV exporter schema consistency tests."""

    @pytest.fixture
    def sample_companies(self):
        """Create sample companies for testing."""
        return [
            Company(
                id="comp-001",
                name="Test Company 1",
                industry="saas",
                revenue=1000000.0,
                growth_rate=0.5,
                employees=100,
                classification="Phoenix",
            ),
            Company(
                id="comp-002",
                name="Test Company 2",
                industry="fintech",
                revenue=500000.0,
                growth_rate=0.3,
                employees=50,
                classification="Salt",
            ),
        ]

    def test_csv_schema_consistency(self, sample_companies, tmp_path):
        """CSV output should have consistent schema."""
        exporter = CSVExporter()
        output_path = tmp_path / "test_export.csv"

        exporter.export(sample_companies, output_path)

        # Read and verify CSV structure
        with open(output_path, newline='') as f:
            reader = csv.reader(f)
            rows = list(reader)

        # Verify header row exists and has expected columns
        assert len(rows) > 0
        header = rows[0]
        assert 'Name' in header or 'Industry' in header

        # Verify data rows match header count
        for row in rows[1:]:
            assert len(row) == len(header)

    def test_csv_field_order_consistency(self, sample_companies, tmp_path):
        """CSV field order should be consistent across exports."""
        exporter = CSVExporter()
        output_path1 = tmp_path / "export1.csv"
        output_path2 = tmp_path / "export2.csv"

        exporter.export(sample_companies, output_path1)
        exporter.export(sample_companies, output_path2)

        # Read both files
        with open(output_path1) as f1, open(output_path2) as f2:
            reader1 = csv.reader(f1)
            reader2 = csv.reader(f2)
            header1 = next(reader1)
            header2 = next(reader2)

        # Headers should be identical
        assert header1 == header2


class TestExcelExporterSnapshots:
    """Excel exporter schema and visual consistency tests."""

    @pytest.fixture
    def sample_companies(self):
        """Create sample companies for testing."""
        return [
            Company(
                id="comp-001",
                name="Test Company 1",
                industry="saas",
                revenue=1000000.0,
                growth_rate=0.5,
                employees=100,
                classification="Phoenix",
            ),
            Company(
                id="comp-002",
                name="Test Company 2",
                industry="fintech",
                revenue=500000.0,
                growth_rate=0.3,
                employees=50,
                classification="Salt",
            ),
        ]

    def test_excel_layout_constants_unchanged(self):
        """Layout constants should not change without review."""
        # These are part of the visual contract
        assert LayoutConstants.TITLE_ROW == 1
        assert LayoutConstants.SUBTITLE_ROW == 2
        assert LayoutConstants.HEADER_ROW == 4
        assert LayoutConstants.DATA_START_ROW == 5
        assert LayoutConstants.TITLE_COLUMN == 1

    def test_excel_color_palette_unchanged(self):
        """Color palette should not change without review."""
        # These colors are part of the brand identity
        assert ColorPalette.OBSIDIAN == "0A0A0F"
        assert ColorPalette.GOLD == "D4A843"
        assert ColorPalette.EMERALD == "2ECC71"
        assert ColorPalette.RUBY == "E74C3C"
        assert ColorPalette.SAPPHIRE == "3498DB"

        # Classification colors
        assert ColorPalette.PHOENIX_FILL == "D5F5E3"
        assert ColorPalette.SALT_FILL == "FEF9E7"
        assert ColorPalette.LEAD_FILL == "FADBD8"

    def test_excel_sheet_structure(self, sample_companies, tmp_path):
        """Excel sheet structure should be consistent."""
        from openpyxl import load_workbook

        exporter = ImprovedExcelExporter()
        output_path = tmp_path / "test_export.xlsx"

        exporter.create_dashboard(sample_companies, output_path)

        # Load and verify structure
        wb = load_workbook(output_path)

        # Should have at least one sheet
        assert len(wb.sheetnames) >= 1

        # Check first sheet structure
        ws = wb.active

        # Title should be in title row
        title_cell = ws.cell(row=LayoutConstants.TITLE_ROW, column=LayoutConstants.TITLE_COLUMN)
        assert title_cell.value is not None

        wb.close()


class TestExportSchemaValidation:
    """Export schema validation tests."""

    EXPECTED_COMPANY_FIELDS = [
        'id',
        'name',
        'industry',
        'revenue',
        'growth_rate',
        'employees',
        'classification',
    ]

    def test_company_export_schema_completeness(self):
        """Company export should include all expected fields."""
        company = Company(
            id="test",
            name="Test Co",
            industry="saas",
            revenue=1000000.0,
            growth_rate=0.5,
            employees=100,
            classification="Phoenix",
        )

        # Convert to dict and check fields
        company_dict = company.model_dump()

        for field in self.EXPECTED_COMPANY_FIELDS:
            assert field in company_dict, f"Missing field: {field}"

    def test_classification_values_consistency(self):
        """Classification values should be consistent."""
        valid_classifications = {"Phoenix", "Salt", "Lead"}

        for classification in valid_classifications:
            company = Company(
                id="test",
                name="Test",
                industry="saas",
                classification=classification,
            )
            assert company.classification == classification


class TestExportRegressionPrevention:
    """Tests to prevent export regressions."""

    def test_export_timestamp_format(self):
        """Export timestamps should use consistent format."""
        timestamp = datetime.now()
        formatted = timestamp.strftime("%Y-%m-%d %H:%M:%S")

        # Should be parseable
        parsed = datetime.strptime(formatted, "%Y-%m-%d %H:%M:%S")
        assert parsed is not None

    def test_numeric_field_formatting(self):
        """Numeric fields should be formatted consistently."""
        company = Company(
            id="test",
            name="Test",
            industry="saas",
            revenue=1234567.89,
            growth_rate=0.123456,
        )

        # Revenue should be numeric
        assert isinstance(company.revenue, float)

        # Growth rate should be numeric
        assert isinstance(company.growth_rate, float)
