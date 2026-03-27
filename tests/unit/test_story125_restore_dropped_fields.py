"""Tests for STORY-125: Restore 20 dropped fields to Excel export.

Validates that all fields identified in the EPIC-033 forensic audit
appear in the correct sheets of the exported Excel workbook.
"""

from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from solstein.domain.models import Company, FinancialMetric
from solstein.exporters.excel_improved import ImprovedExcelExporter


def _make_company(**overrides: Any) -> Company:
    """Create a test company with all 20 dropped fields populated."""
    fm = FinancialMetric(
        revenue=50.0,
        growth_rate=0.15,
        profit_margin=0.22,
        employees=350,
        total_funding_raised_eur=25_000_000,
        latest_valuation_eur=100_000_000,
        allow_empty_primary=True,
    )
    defaults: dict[str, Any] = {
        "id": "TEST-EXPORT-001",
        "name": "TestExport Corp",
        "industry": "Energy Software",
        "financials": fm,
        # Company sheet fields
        "tech_stack": ["Python", "React", "PostgreSQL"],
        "key_customers": ["Acme Inc", "Globex Corp"],
        "open_positions": 12,
        "data_availability": "high",
        # Financial sheet fields
        "funding_rounds": [
            {"round": "Series A", "amount": "5M"},
            {"round": "Series B", "amount": "20M"},
        ],
        "funding_war_chest": "Strong - 18 months runway",
        "revenue_cagr_5yr": 0.35,
        "revenue_per_employee_eur_k": 142.8,
        "employee_cagr_3yr": 0.25,
        "lead_investors": ["Sequoia", "a16z"],
        # Revenue History fields
        "revenue_timeline": [
            {"year": 2023, "eur_millions": 30.0, "source": "SEC filing"},
            {"year": 2024, "eur_millions": 40.0, "source": "Annual report"},
            {"year": 2025, "eur_millions": 50.0, "source": "Estimate"},
        ],
        # Advanced Data fields
        "parent_company": "MegaCorp Holdings",
        "subsidiaries": ["SubCo Alpha", "SubCo Beta"],
        "acquisitions": [{"company": "StartupX", "year": 2024, "amount": "10M"}],
        "notes": "Strong growth in EU market. Watch for regulatory changes.",
        "source_links": [
            "https://example.com/report",
            "https://crunchbase.com/org/testexport",
        ],
        "metric_sources": {
            "revenue": ["sec_edgar", "annual_report"],
            "employees": ["linkedin", "company_website"],
        },
        "enrichment_quality_metrics": {
            "merge_conflicts": {"revenue": "SEC vs Crunchbase mismatch"},
        },
    }
    defaults.update(overrides)
    return Company(**defaults)


@pytest.fixture()
def export_path(tmp_path: Path) -> Path:
    """Return a temporary path for the exported Excel file."""
    return tmp_path / "test_export.xlsx"


@pytest.fixture()
def exported_workbook(export_path: Path) -> Any:
    """Export a workbook with a fully-populated test company and return it."""
    company = _make_company()
    exporter = ImprovedExcelExporter()
    exporter.create_dashboard([company], export_path)
    return load_workbook(export_path)


class TestSheetPresence:
    """Verify all required sheets exist."""

    def test_revenue_history_sheet_exists(self, exported_workbook: Any) -> None:
        assert "Revenue History" in exported_workbook.sheetnames

    def test_advanced_data_sheet_exists(self, exported_workbook: Any) -> None:
        assert "Advanced Data" in exported_workbook.sheetnames

    def test_existing_sheets_preserved(self, exported_workbook: Any) -> None:
        assert "Executive Summary" in exported_workbook.sheetnames
        assert "Market Rankings" in exported_workbook.sheetnames
        assert "Financial Intelligence" in exported_workbook.sheetnames


class TestCompanySheetFields:
    """Verify the 4 new fields on the Executive Summary (Company) sheet."""

    def test_tech_stack_column_exists(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Executive Summary"]
        headers = [ws.cell(row=4, column=c).value for c in range(1, 20)]
        assert "Tech Stack" in headers

    def test_tech_stack_value(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Executive Summary"]
        # Tech Stack is column 8 (index from 1)
        cell_value = ws.cell(row=5, column=8).value
        assert "Python" in cell_value
        assert "React" in cell_value

    def test_key_customers_column_exists(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Executive Summary"]
        headers = [ws.cell(row=4, column=c).value for c in range(1, 20)]
        assert "Key Customers" in headers

    def test_open_positions_column_exists(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Executive Summary"]
        headers = [ws.cell(row=4, column=c).value for c in range(1, 20)]
        assert "Open Positions" in headers

    def test_data_availability_column_exists(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Executive Summary"]
        headers = [ws.cell(row=4, column=c).value for c in range(1, 20)]
        assert "Data Availability" in headers


class TestFinancialSheetFields:
    """Verify the 5 new fields on the Financial Intelligence sheet."""

    def test_funding_rounds_column_exists(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Financial Intelligence"]
        headers = [ws.cell(row=4, column=c).value for c in range(1, 20)]
        assert "Funding Rounds" in headers

    def test_funding_rounds_value(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Financial Intelligence"]
        cell_value = ws.cell(row=5, column=8).value
        assert "Series A" in cell_value

    def test_funding_war_chest_column_exists(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Financial Intelligence"]
        headers = [ws.cell(row=4, column=c).value for c in range(1, 20)]
        assert "Funding War Chest" in headers

    def test_revenue_cagr_5yr_column_exists(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Financial Intelligence"]
        headers = [ws.cell(row=4, column=c).value for c in range(1, 20)]
        assert "Revenue CAGR 5yr" in headers

    def test_revenue_per_employee_column_exists(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Financial Intelligence"]
        headers = [ws.cell(row=4, column=c).value for c in range(1, 20)]
        assert "Revenue/Employee (€K)" in headers

    def test_employee_cagr_3yr_column_exists(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Financial Intelligence"]
        headers = [ws.cell(row=4, column=c).value for c in range(1, 20)]
        assert "Employee CAGR 3yr" in headers


class TestRevenueHistorySheet:
    """Verify the Revenue History sheet."""

    def test_headers_correct(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Revenue History"]
        headers = [ws.cell(row=4, column=c).value for c in range(1, 5)]
        assert headers == ["Company", "Year", "Revenue (EUR M)", "Source"]

    def test_timeline_rows(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Revenue History"]
        # 3 timeline entries should produce 3 data rows
        years = [ws.cell(row=5 + i, column=2).value for i in range(3)]
        assert 2023 in years
        assert 2024 in years
        assert 2025 in years

    def test_company_name_on_each_row(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Revenue History"]
        for i in range(3):
            assert ws.cell(row=5 + i, column=1).value == "TestExport Corp"


class TestAdvancedDataSheet:
    """Verify the Advanced Data sheet."""

    def test_headers_correct(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Advanced Data"]
        headers = [ws.cell(row=4, column=c).value for c in range(1, 9)]
        expected = [
            "Company", "Parent Company", "Subsidiaries", "Acquisitions",
            "Notes", "Source Links", "Data Sources Per Field", "Merge Conflicts",
        ]
        assert headers == expected

    def test_parent_company_value(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Advanced Data"]
        assert ws.cell(row=5, column=2).value == "MegaCorp Holdings"

    def test_subsidiaries_value(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Advanced Data"]
        val = ws.cell(row=5, column=3).value
        assert "SubCo Alpha" in val
        assert "SubCo Beta" in val

    def test_acquisitions_value(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Advanced Data"]
        val = ws.cell(row=5, column=4).value
        assert "StartupX" in val

    def test_notes_value(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Advanced Data"]
        assert "Strong growth" in ws.cell(row=5, column=5).value

    def test_source_links_value(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Advanced Data"]
        val = ws.cell(row=5, column=6).value
        assert "example.com" in val

    def test_source_links_has_hyperlink(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Advanced Data"]
        cell = ws.cell(row=5, column=6)
        assert cell.hyperlink is not None

    def test_data_sources_per_field_value(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Advanced Data"]
        val = ws.cell(row=5, column=7).value
        assert "revenue" in val
        assert "sec_edgar" in val

    def test_merge_conflicts_value(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Advanced Data"]
        val = ws.cell(row=5, column=8).value
        assert "mismatch" in val.lower() or "revenue" in val.lower()


class TestNullFieldHandling:
    """Export must not fail when any of the 20 fields is null."""

    def test_export_with_all_null_fields(self, export_path: Path) -> None:
        """A bare-minimum company (no optional fields) exports without error."""
        company = Company(id="TEST-NULL", name="Null Corp")
        exporter = ImprovedExcelExporter()
        exporter.create_dashboard([company], export_path)
        wb = load_workbook(export_path)
        assert "Executive Summary" in wb.sheetnames
        assert "Revenue History" in wb.sheetnames
        assert "Advanced Data" in wb.sheetnames

    def test_null_revenue_timeline_gets_na_row(self, export_path: Path) -> None:
        """Company with no revenue_timeline gets an N/A row."""
        company = Company(id="TEST-EMPTY-TIMELINE", name="Empty Timeline Corp")
        exporter = ImprovedExcelExporter()
        exporter.create_dashboard([company], export_path)
        wb = load_workbook(export_path)
        ws = wb["Revenue History"]
        assert ws.cell(row=5, column=1).value == "Empty Timeline Corp"
        assert ws.cell(row=5, column=2).value == "N/A"


class TestBackwardCompatibility:
    """Existing columns must not be removed or reordered."""

    def test_executive_summary_original_columns_preserved(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Executive Summary"]
        original_headers = ["Company", "Industry", "Revenue (€M)", "Growth", "AI Score", "Tier", "Threat Level"]
        actual = [ws.cell(row=4, column=c).value for c in range(1, 8)]
        assert actual == original_headers

    def test_financial_original_columns_preserved(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Financial Intelligence"]
        original_headers = [
            "Company", "Revenue (€M)", "Growth Rate", "Profit Margin",
            "Total Funding", "Latest Valuation", "Investors",
        ]
        actual = [ws.cell(row=4, column=c).value for c in range(1, 8)]
        assert actual == original_headers

    def test_market_rankings_unchanged(self, exported_workbook: Any) -> None:
        ws = exported_workbook["Market Rankings"]
        original_headers = ["Rank", "Company", "Market Share", "Competitive Score", "Growth Rate", "Employees"]
        actual = [ws.cell(row=4, column=c).value for c in range(1, 7)]
        assert actual == original_headers


class TestAllTwentyFieldsPresent:
    """Regression test: all 20 field names must appear somewhere in the export."""

    FIELD_TO_HEADER = {
        "tech_stack": "Tech Stack",
        "key_customers": "Key Customers",
        "open_positions": "Open Positions",
        "data_availability": "Data Availability",
        "funding_rounds": "Funding Rounds",
        "funding_war_chest": "Funding War Chest",
        "lead_investors": "Investors",
        "revenue_cagr_5yr": "Revenue CAGR 5yr",
        "revenue_per_employee_eur_k": "Revenue/Employee (€K)",
        "employee_cagr_3yr": "Employee CAGR 3yr",
        "revenue_timeline": "Revenue (EUR M)",  # On Revenue History sheet
        "parent_company": "Parent Company",
        "subsidiaries": "Subsidiaries",
        "acquisitions": "Acquisitions",
        "notes": "Notes",
        "source_links": "Source Links",
        "data_source_per_field": "Data Sources Per Field",
        "merge_conflicts": "Merge Conflicts",
        "profit_margin": "Profit Margin",  # Already existed
        "employee_count": "Employees",  # Already existed on Market Rankings
    }

    def test_all_field_headers_present(self, exported_workbook: Any) -> None:
        """Every one of the 20 fields maps to a header present in the workbook."""
        all_headers: set[str] = set()
        for sheet_name in exported_workbook.sheetnames:
            ws = exported_workbook[sheet_name]
            for col in range(1, 30):
                val = ws.cell(row=4, column=col).value
                if val:
                    all_headers.add(val)

        missing = []
        for field, header in self.FIELD_TO_HEADER.items():
            if header not in all_headers:
                missing.append(f"{field} -> {header}")
        assert not missing, f"Missing headers: {missing}"
