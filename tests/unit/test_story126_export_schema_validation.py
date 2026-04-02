"""Tests for STORY-126: Export Schema Validation.

Validates that:
- ExportSchema defines all required fields
- Validation catches missing fields
- Validation catches wrong-sheet placement
- Schema version is embedded in export metadata
- Auto-generated documentation covers all fields
"""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from solstein.domain.models import Company, FinancialMetric
from solstein.exporters.excel_improved import ImprovedExcelExporter
from solstein.exporters.export_schema import (
    EXPORT_SCHEMA,
    EXPORT_SCHEMA_VERSION,
    ExportValidationError,
    FieldSpec,
    generate_schema_docs,
    get_schema_by_sheet,
    validate_export,
)


def _make_full_company() -> Company:
    """Create a company with all exported fields populated."""
    fm = FinancialMetric(
        revenue=50.0,
        growth_rate=0.15,
        profit_margin=0.22,
        employees=350,
        total_funding_raised_eur=25_000_000,
        latest_valuation_eur=100_000_000,
        allow_empty_primary=True,
    )
    return Company(
        id="TEST-SCHEMA-001",
        name="SchemaTest Corp",
        industry="Energy Software",
        financials=fm,
        tech_stack=["Python", "React"],
        key_customers=["Acme"],
        open_positions=5,
        data_availability="high",
        funding_rounds=[{"round": "Series A", "amount": "5M"}],
        funding_war_chest="Strong",
        revenue_cagr_5yr=0.30,
        revenue_per_employee_eur_k=140.0,
        employee_cagr_3yr=0.20,
        lead_investors=["Sequoia"],
        revenue_timeline=[{"year": 2024, "eur_millions": 40.0, "source": "Annual report"}],
        parent_company="MegaCorp",
        subsidiaries=["SubCo"],
        acquisitions=[{"company": "Startup", "year": 2024}],
        notes="Test notes",
        source_links=["https://example.com"],
        metric_sources={"revenue": ["sec_edgar"]},
        enrichment_quality_metrics={"merge_conflicts": {"rev": "mismatch"}},
    )


@pytest.fixture()
def export_path(tmp_path: Path) -> Path:
    return tmp_path / "schema_test.xlsx"


@pytest.fixture()
def valid_export(export_path: Path) -> Path:
    """Create a valid export file and return its path."""
    company = _make_full_company()
    exporter = ImprovedExcelExporter()
    exporter.create_dashboard([company], export_path)
    return export_path


class TestExportSchemaDefinition:
    """Schema definition completeness."""

    def test_schema_has_version(self) -> None:
        assert EXPORT_SCHEMA_VERSION == "1.0"

    def test_schema_covers_all_five_sheets(self) -> None:
        by_sheet = get_schema_by_sheet()
        assert "Executive Summary" in by_sheet
        assert "Market Rankings" in by_sheet
        assert "Financial Intelligence" in by_sheet
        assert "Revenue History" in by_sheet
        assert "Advanced Data" in by_sheet

    def test_schema_covers_story125_fields(self) -> None:
        """All 20 STORY-125 restored fields are represented in the schema."""
        story125_headers = {
            "Tech Stack",
            "Key Customers",
            "Open Positions",
            "Data Availability",
            "Funding Rounds",
            "Funding War Chest",
            "Investors",
            "Revenue CAGR 5yr",
            "Revenue/Employee (€K)",
            "Employee CAGR 3yr",
            "Company",
            "Year",
            "Revenue (EUR M)",
            "Source",
            "Parent Company",
            "Subsidiaries",
            "Acquisitions",
            "Notes",
            "Source Links",
            "Data Sources Per Field",
            "Merge Conflicts",
            "Profit Margin",
            "Employees",
        }
        schema_headers = {spec.header for spec in EXPORT_SCHEMA}
        missing = story125_headers - schema_headers
        assert not missing, f"Schema missing headers for STORY-125 fields: {missing}"

    def test_all_fields_have_required_attributes(self) -> None:
        for spec in EXPORT_SCHEMA:
            assert spec.name, "Field spec missing name"
            assert spec.header, f"Field {spec.name} missing header"
            assert spec.sheet, f"Field {spec.name} missing sheet"
            assert spec.data_type in {"string", "number", "percentage", "integer", "list", "structured"}, (
                f"Field {spec.name} has invalid data_type: {spec.data_type}"
            )


class TestValidationPasses:
    """Positive tests: valid exports pass validation."""

    def test_valid_export_passes(self, valid_export: Path) -> None:
        """Complete export passes schema validation without error."""
        validate_export(valid_export)

    def test_export_with_null_fields_passes(self, export_path: Path) -> None:
        """Company with null optional data still passes (columns present, values N/A)."""
        company = Company(id="TEST-NULL", name="Null Corp")
        exporter = ImprovedExcelExporter()
        exporter.create_dashboard([company], export_path)
        # Validation should pass because headers are present even with N/A values
        validate_export(export_path)


class TestValidationFails:
    """Negative tests: invalid exports trigger ExportValidationError."""

    def test_missing_field_raises_error(self, export_path: Path) -> None:
        """Deliberately removing a required header causes validation to fail."""
        company = _make_full_company()
        exporter = ImprovedExcelExporter()

        # Patch add_executive_summary to skip tech_stack column
        # Create the export normally first, then tamper with the file
        exporter.create_dashboard([company], export_path)

        # Tamper: remove Tech Stack header from the exported file
        wb = load_workbook(export_path)
        ws = wb["Executive Summary"]
        # Find and clear the Tech Stack header
        for col in range(1, 20):
            if ws.cell(row=4, column=col).value == "Tech Stack":
                ws.cell(row=4, column=col).value = None
                break
        wb.save(export_path)

        with pytest.raises(ExportValidationError) as exc_info:
            validate_export(export_path)
        assert "tech_stack" in str(exc_info.value)

    def test_error_includes_field_name(self, export_path: Path) -> None:
        """ExportValidationError names the specific missing field."""
        # Validate against a custom schema with a field that doesn't exist
        fake_schema = [
            FieldSpec("nonexistent_field", "This Header Does Not Exist", "Executive Summary", "string"),
        ]
        company = _make_full_company()
        exporter = ImprovedExcelExporter()
        exporter.create_dashboard([company], export_path)

        with pytest.raises(ExportValidationError) as exc_info:
            validate_export(export_path, schema=fake_schema)
        assert "nonexistent_field" in str(exc_info.value)
        assert exc_info.value.missing_fields  # Has at least one missing field


class TestSchemaVersionInMetadata:
    """Schema version is embedded in the exported file."""

    def test_metadata_sheet_has_schema_version(self, valid_export: Path) -> None:
        wb = load_workbook(valid_export)
        assert "Export Metadata" in wb.sheetnames
        ws = wb["Export Metadata"]
        # Find the schema version row
        found = False
        for row in range(1, 20):
            key = ws.cell(row=row, column=1).value
            if key == "export_schema_version":
                assert ws.cell(row=row, column=2).value == EXPORT_SCHEMA_VERSION
                found = True
                break
        assert found, "export_schema_version not found in Export Metadata sheet"


class TestDocumentationGeneration:
    """Schema documentation is auto-generated."""

    def test_generate_docs_includes_version(self) -> None:
        docs = generate_schema_docs()
        assert EXPORT_SCHEMA_VERSION in docs

    def test_generate_docs_includes_all_sheets(self) -> None:
        docs = generate_schema_docs()
        assert "## Executive Summary" in docs
        assert "## Market Rankings" in docs
        assert "## Financial Intelligence" in docs
        assert "## Revenue History" in docs
        assert "## Advanced Data" in docs

    def test_generate_docs_includes_changelog(self) -> None:
        docs = generate_schema_docs()
        assert "Schema Changelog" in docs
        assert "1.0" in docs

    def test_generate_docs_includes_all_field_headers(self) -> None:
        docs = generate_schema_docs()
        for spec in EXPORT_SCHEMA:
            assert spec.header in docs, f"Header '{spec.header}' not in generated docs"
