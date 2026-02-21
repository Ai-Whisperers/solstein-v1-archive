"""
Unit tests for SolStein exporters.
"""


import pytest
from openpyxl import load_workbook

from solstein.domain.models import (
    AIMaturity,
    Company,
    CompanyTier,
    FinancialMetric,
    ThreatLevel,
)
from solstein.exporters.excel import ExcelExporter, TemplateExporter


@pytest.fixture
def sample_profiles():
    return [
        Company(
            id="c1",
            name="Alpha Corp",
            tier=CompanyTier.TIER_1,
            threat_level=ThreatLevel.HIGH,
            ai_maturity=AIMaturity.STRONG,
            growth_score=8.5,
            saas_maturity=5,
            geographic_presence=["US", "EU"],
            key_customers=["A", "B"],
            financials=FinancialMetric(
                revenue=50_000_000.0,
                growth_rate=25.0,
                profit_margin=15.0,
                funding_raised=20_000_000.0,
                valuation=250_000_000.0,
                employees=200,
            ),
        ),
        Company(
            id="c2",
            name="Beta Ltd",
            tier=CompanyTier.TIER_3,
            threat_level=ThreatLevel.LOW,
            ai_maturity=AIMaturity.NONE,
            growth_score=3.2,
            saas_maturity=1,
            geographic_presence=["UK"],
            key_customers=["C"],
            financials=FinancialMetric(
                revenue=5_000_000.0,
                growth_rate=-5.0,
                profit_margin=-10.0,
                funding_raised=2_000_000.0,
                valuation=15_000_000.0,
                employees=30,
            ),
        ),
    ]


class TestExcelExporter:
    def test_create_styles(self):
        exporter = ExcelExporter()
        styles = exporter._create_styles()
        assert "title_font" in styles
        assert "header_font" in styles
        assert "data_fill_even" in styles
        assert "phoenix_fill" in styles

    def test_create_dashboard(self, sample_profiles, tmp_path):
        output_file = tmp_path / "glyph_report.xlsx"
        exporter = ExcelExporter()

        exporter.create_dashboard(sample_profiles, output_file)

        assert output_file.exists()

        # Verify content
        wb = load_workbook(output_file)
        expected_sheets = ["Executive Summary", "Market Rankings", "Financial Intelligence", "Tech & AI Maturity"]
        for sheet in expected_sheets:
            assert sheet in wb.sheetnames

        ws_summary = wb["Executive Summary"]
        # Banner check
        assert "SOLSTEIN | EXECUTIVE SUMMARY" in ws_summary["A1"].value
        
        # KPI check
        assert "Total Portfolio Size" in str(ws_summary["A5"].value)
        assert str(len(sample_profiles)) in str(ws_summary["B5"].value)

        # Rankings check
        ws_rankings = wb["Market Rankings"]
        assert ws_rankings["B4"].value == "Company"
        assert ws_rankings["B5"].value == "Alpha Corp"

        wb.close()

    def test_create_dashboard_empty(self, tmp_path):
        output_file = tmp_path / "empty_dashboard.xlsx"
        exporter = ExcelExporter()

        exporter.create_dashboard([], output_file)

        assert output_file.exists()


class TestTemplateExporter:
    def test_create_dashboard_not_implemented(self, sample_profiles, tmp_path):
        output_file = tmp_path / "template.xlsx"
        exporter = TemplateExporter()

        with pytest.raises(NotImplementedError):
            exporter.create_dashboard(sample_profiles, output_file)
