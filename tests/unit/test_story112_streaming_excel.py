"""Tests for STORY-112: Streaming Excel Export for Large Datasets.

Validates:
- REQ-1: OpenPyXL write_only mode is used for O(1) memory usage
- REQ-2: Four sheets generated (Summary, Companies, Signals, Financials)
- REQ-3: Progress callback updates export_jobs.progress_pct
- REQ-4: Integration with Celery export task via _generate_file
- REQ-5: Limitations documented (no auto-width, no merging, no cell styling)
"""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook

from solstein.infrastructure.models.export import ExportJobRecord

# Load export_tasks module directly to avoid __init__.py chain issues
_EXPORT_TASKS_PATH = str(Path(__file__).parent.parent.parent / "src" / "solstein" / "worker" / "export_tasks.py")
_spec = importlib.util.spec_from_file_location(
    "solstein.worker.export_tasks",
    _EXPORT_TASKS_PATH,
    submodule_search_locations=[],
)
_export_mod = importlib.util.module_from_spec(_spec)  # type: ignore[arg-type]
sys.modules["solstein.worker.export_tasks"] = _export_mod
_spec.loader.exec_module(_export_mod)  # type: ignore[union-attr]

# Import the streaming exporter. The PYTHONPATH includes src/ so we
# can import directly. We avoid loading via importlib here because
# the relative import chain (..domain.models) requires the full
# package tree to be resolved.
from solstein.exporters.excel_streaming import StreamingExcelExporter

_stream_mod = sys.modules["solstein.exporters.excel_streaming"]

_STREAMING_PATH = str(Path(__file__).parent.parent.parent / "src" / "solstein" / "exporters" / "excel_streaming.py")

# Source text for structural inspection
_STREAMING_SOURCE = Path(_STREAMING_PATH).read_text()
_EXPORT_TASKS_SOURCE = Path(_EXPORT_TASKS_PATH).read_text()
_EXPORT_MODEL_SOURCE = Path(
    Path(__file__).parent.parent.parent / "src" / "solstein" / "infrastructure" / "models" / "export.py"
).read_text()


# ---------------------------------------------------------------------------
# Helper: Fake Company domain object
# ---------------------------------------------------------------------------
class _FakeCompany:
    """Minimal Company stand-in for exporter tests."""

    def __init__(self, name: str = "Acme Corp", industry: str = "Technology", classification: str = "Growth") -> None:
        self.name, self.industry, self.classification = name, industry, classification
        self.overall_score, self.growth_score = 0.85, 0.72
        self.ai_maturity, self.description = "Advanced", "A technology company"
        self.website, self.country = "https://acme.example.com", "US"
        self.employee_count, self.founded_year = 500, 2015
        self.revenue, self.revenue_growth = 10_000_000, 0.25
        self.total_funding, self.last_funding_round = 50_000_000, "Series B"
        self.gross_margin, self.ebitda = 0.65, 2_000_000
        self.net_income, self.burn_rate, self.runway_months = 1_500_000, 500_000, 24
        self.signals = [{"type": "hiring", "value": "50 open roles", "confidence": 0.9, "source": "LinkedIn"}]


# ---------------------------------------------------------------------------
# REQ-1: OpenPyXL write_only mode for O(1) memory
# ---------------------------------------------------------------------------
class TestWriteOnlyMode:
    """StreamingExcelExporter must use write_only=True."""

    def test_source_uses_write_only(self) -> None:
        """write_only=True must appear in the exporter source."""
        assert "write_only=True" in _STREAMING_SOURCE

    def test_workbook_created_with_write_only(
        self,
        tmp_path: Path,
    ) -> None:
        """Workbook constructor is called with write_only=True."""
        exporter = StreamingExcelExporter()
        companies = [_FakeCompany()]
        output = tmp_path / "test.xlsx"

        with patch(
            "solstein.exporters.excel_streaming.Workbook",
        ) as mock_wb_cls:
            mock_wb = MagicMock()
            mock_wb.create_sheet.return_value = MagicMock()
            mock_wb_cls.return_value = mock_wb

            exporter.create_dashboard(companies, output)
            mock_wb_cls.assert_called_once_with(write_only=True)

    def test_source_documents_o1_memory(self) -> None:
        """Docstring must mention O(1) memory usage."""
        assert "O(1)" in _STREAMING_SOURCE

    def test_no_cell_level_access(self) -> None:
        """Source must not use ws.cell() or ws.merge_cells() calls."""
        assert "ws.cell(" not in _STREAMING_SOURCE
        assert ".merge_cells(" not in _STREAMING_SOURCE


# ---------------------------------------------------------------------------
# REQ-2: Four sheets generated
# ---------------------------------------------------------------------------
class TestFourSheets:
    """Exporter must create Summary, Companies, Signals, Financials."""

    def test_creates_four_sheets(self, tmp_path: Path) -> None:
        """Output Excel file must contain exactly 4 sheets."""

        exporter = StreamingExcelExporter()
        companies = [_FakeCompany()]
        output = tmp_path / "test_sheets.xlsx"

        exporter.create_dashboard(companies, output)

        wb = load_workbook(str(output), read_only=True)
        assert len(wb.sheetnames) == 4
        wb.close()

    def test_sheet_names_correct(self, tmp_path: Path) -> None:
        """Sheets must be named Summary, Companies, Signals, Financials."""

        exporter = StreamingExcelExporter()
        companies = [_FakeCompany()]
        output = tmp_path / "test_names.xlsx"

        exporter.create_dashboard(companies, output)

        wb = load_workbook(str(output), read_only=True)
        expected = ["Summary", "Companies", "Signals", "Financials"]
        assert wb.sheetnames == expected
        wb.close()

    def test_summary_sheet_has_header(self, tmp_path: Path) -> None:
        """Summary sheet must have Export Summary header."""

        exporter = StreamingExcelExporter()
        output = tmp_path / "test_summary.xlsx"
        exporter.create_dashboard([_FakeCompany()], output)

        wb = load_workbook(str(output), read_only=True)
        ws = wb["Summary"]
        rows = list(ws.rows)
        first_row_values = [c.value for c in rows[0]]
        assert "Export Summary" in first_row_values
        wb.close()

    def test_companies_sheet_has_headers(self, tmp_path: Path) -> None:
        """Companies sheet must have column headers."""

        exporter = StreamingExcelExporter()
        output = tmp_path / "test_companies.xlsx"
        exporter.create_dashboard([_FakeCompany()], output)

        wb = load_workbook(str(output), read_only=True)
        ws = wb["Companies"]
        rows = list(ws.rows)
        header_values = [c.value for c in rows[0]]
        assert "Name" in header_values
        assert "Industry" in header_values
        wb.close()

    def test_signals_sheet_has_data(self, tmp_path: Path) -> None:
        """Signals sheet must contain signal data rows."""

        exporter = StreamingExcelExporter()
        company = _FakeCompany()
        output = tmp_path / "test_signals.xlsx"
        exporter.create_dashboard([company], output)

        wb = load_workbook(str(output), read_only=True)
        ws = wb["Signals"]
        rows = list(ws.rows)
        # Header + 1 signal row
        assert len(rows) >= 2
        wb.close()

    def test_financials_sheet_has_data(self, tmp_path: Path) -> None:
        """Financials sheet must contain financial data."""

        exporter = StreamingExcelExporter()
        output = tmp_path / "test_financials.xlsx"
        exporter.create_dashboard([_FakeCompany()], output)

        wb = load_workbook(str(output), read_only=True)
        ws = wb["Financials"]
        rows = list(ws.rows)
        # Header + 1 data row
        assert len(rows) >= 2
        header_values = [c.value for c in rows[0]]
        assert "Revenue" in header_values
        wb.close()

    def test_source_defines_total_sheets_four(self) -> None:
        """_total_sheets must be 4 in the exporter source."""
        assert "_total_sheets = 4" in _STREAMING_SOURCE


# ---------------------------------------------------------------------------
# REQ-3: Progress callback updates progress_pct
# ---------------------------------------------------------------------------
class TestProgressCallback:
    """Progress callback must be invoked and update DB."""

    def test_callback_called_four_times(self, tmp_path: Path) -> None:
        """Callback is called once per sheet (4 times)."""
        exporter = StreamingExcelExporter()
        output = tmp_path / "test_progress.xlsx"
        callback = MagicMock()

        exporter.create_dashboard([_FakeCompany()], output, callback)

        assert callback.call_count == 4

    def test_callback_values_increase(self, tmp_path: Path) -> None:
        """Callback values must be monotonically increasing to 100."""
        exporter = StreamingExcelExporter()
        output = tmp_path / "test_progress_vals.xlsx"
        values: list[int] = []

        def track(pct: int) -> None:
            values.append(pct)

        exporter.create_dashboard([_FakeCompany()], output, track)

        assert len(values) == 4
        assert values == sorted(values)
        assert values[-1] == 100

    def test_progress_pct_capped_at_99_in_task(self) -> None:
        """_update_progress must cap at 99 (100 only on completion)."""
        assert "min(pct, 99)" in _EXPORT_TASKS_SOURCE

    def test_export_model_has_progress_pct(self) -> None:
        """ExportJobRecord must have progress_pct field."""
        assert "progress_pct" in _EXPORT_MODEL_SOURCE
        assert hasattr(ExportJobRecord, "progress_pct")

    def test_progress_callback_wired_in_task(self) -> None:
        """_execute_export must define _update_progress and _sync_progress."""
        assert "_update_progress" in _EXPORT_TASKS_SOURCE
        assert "_sync_progress" in _EXPORT_TASKS_SOURCE
        # _sync_progress must be passed to _generate_file (on its own line, after filters)
        assert "filters,\n        _sync_progress," in _EXPORT_TASKS_SOURCE

    def test_callback_error_does_not_crash(
        self,
        tmp_path: Path,
    ) -> None:
        """A failing callback must not crash the export."""
        exporter = StreamingExcelExporter()
        output = tmp_path / "test_callback_err.xlsx"

        def bad_callback(pct: int) -> None:
            raise RuntimeError("callback broke")

        # Should not raise
        exporter.create_dashboard(
            [_FakeCompany()],
            output,
            bad_callback,
        )
        assert output.exists()


# ---------------------------------------------------------------------------
# REQ-4: Integration with Celery export task
# ---------------------------------------------------------------------------
class TestCeleryIntegration:
    """Streaming exporter must integrate with the export task."""

    def test_generate_file_accepts_progress_callback(self) -> None:
        """_generate_file must accept progress_callback parameter."""
        assert "progress_callback" in _EXPORT_TASKS_SOURCE
        # Check the function signature
        assert "progress_callback: Any | None = None" in _EXPORT_TASKS_SOURCE

    def test_generate_excel_uses_streaming(self) -> None:
        """_generate_excel must use StreamingExcelExporter."""
        assert "StreamingExcelExporter" in _EXPORT_TASKS_SOURCE

    def test_generate_excel_passes_callback(self) -> None:
        """_generate_excel must pass progress_callback to exporter."""
        assert "exporter.create_dashboard" in _EXPORT_TASKS_SOURCE
        assert "progress_callback" in _EXPORT_TASKS_SOURCE

    def test_generate_file_passes_callback_to_excel(self) -> None:
        """_generate_file must pass callback to _generate_excel."""
        # The excel branch must pass progress_callback
        assert "await _generate_excel(output_path, filters, progress_callback)" in _EXPORT_TASKS_SOURCE

    def test_streaming_exporter_in_init(self) -> None:
        """StreamingExcelExporter must be exported from __init__.py."""
        init_path = Path(__file__).parent.parent.parent / "src" / "solstein" / "exporters" / "__init__.py"
        init_source = init_path.read_text()
        assert "StreamingExcelExporter" in init_source


# ---------------------------------------------------------------------------
# REQ-5: Limitations documented
# ---------------------------------------------------------------------------
class TestDocumentation:
    """Streaming exporter limitations must be documented."""

    def test_limitations_documented(self) -> None:
        """Source must document write_only limitations."""
        src = _STREAMING_SOURCE.lower()
        assert "write_only" in src
        assert "column width" in src
        assert "merge" in src
        assert "styling" in src

    def test_story_references(self) -> None:
        """STORY-112 must be referenced in all modified files."""
        assert "STORY-112" in _STREAMING_SOURCE
        assert "STORY-112" in _EXPORT_TASKS_SOURCE
        assert "STORY-112" in _EXPORT_MODEL_SOURCE


# ---------------------------------------------------------------------------
# Edge cases and data handling
# ---------------------------------------------------------------------------
class TestEdgeCases:
    """Edge cases for the streaming exporter."""

    def test_empty_company_list(self, tmp_path: Path) -> None:
        """Exporter must handle empty company list."""
        exporter = StreamingExcelExporter()
        output = tmp_path / "test_empty.xlsx"
        exporter.create_dashboard([], output)
        assert output.exists()

    def test_company_without_signals(self, tmp_path: Path) -> None:
        """Company with no signals gets placeholder row."""

        company = _FakeCompany()
        company.signals = []
        exporter = StreamingExcelExporter()
        output = tmp_path / "test_no_signals.xlsx"
        exporter.create_dashboard([company], output)

        wb = load_workbook(str(output), read_only=True)
        ws = wb["Signals"]
        rows = list(ws.rows)
        # Header + placeholder row
        assert len(rows) >= 2
        data_row = [c.value for c in rows[1]]
        assert "N/A" in data_row
        wb.close()

    def test_company_with_none_attributes(
        self,
        tmp_path: Path,
    ) -> None:
        """Company with None attributes must not crash."""
        company = _FakeCompany()
        company.revenue = None
        company.description = None
        company.signals = None
        exporter = StreamingExcelExporter()
        output = tmp_path / "test_none_attrs.xlsx"
        exporter.create_dashboard([company], output)
        assert output.exists()

    def test_multiple_companies(self, tmp_path: Path) -> None:
        """Exporter must handle multiple companies."""

        companies = [
            _FakeCompany(name="Alpha Corp", industry="Tech"),
            _FakeCompany(name="Beta Inc", industry="Finance"),
            _FakeCompany(name="Gamma LLC", industry="Health"),
        ]
        exporter = StreamingExcelExporter()
        output = tmp_path / "test_multi.xlsx"
        exporter.create_dashboard(companies, output)

        wb = load_workbook(str(output), read_only=True)
        ws = wb["Companies"]
        rows = list(ws.rows)
        # Header + 3 data rows
        assert len(rows) == 4
        wb.close()

    def test_signal_as_object_not_dict(self, tmp_path: Path) -> None:
        """Signals that are objects (not dicts) must be handled."""

        company = _FakeCompany()

        class _SignalObj:
            type = "tech_adoption"
            value = "high"
            confidence = 0.95
            source = "Internal"

        company.signals = [_SignalObj()]
        exporter = StreamingExcelExporter()
        output = tmp_path / "test_signal_obj.xlsx"
        exporter.create_dashboard([company], output)

        wb = load_workbook(str(output), read_only=True)
        ws = wb["Signals"]
        rows = list(ws.rows)
        data_row = [c.value for c in rows[1]]
        assert "tech_adoption" in data_row
        wb.close()


# ---------------------------------------------------------------------------
# Helper function tests
# ---------------------------------------------------------------------------
class TestHelperFunctions:
    """Test _safe_str and _safe_num helper functions."""

    def test_safe_str_with_none(self) -> None:
        """_safe_str(None) returns None."""
        assert _stream_mod._safe_str(None) is None

    def test_safe_str_with_string(self) -> None:
        """_safe_str('hello') returns 'hello'."""
        assert _stream_mod._safe_str("hello") == "hello"

    def test_safe_str_with_number(self) -> None:
        """_safe_str(42) returns '42'."""
        assert _stream_mod._safe_str(42) == "42"

    def test_safe_str_with_empty(self) -> None:
        """_safe_str('') returns None."""
        assert _stream_mod._safe_str("") is None

    def test_safe_num_with_none(self) -> None:
        """_safe_num(None) returns None."""
        assert _stream_mod._safe_num(None) is None

    def test_safe_num_with_int(self) -> None:
        """_safe_num(42) returns 42."""
        assert _stream_mod._safe_num(42) == 42

    def test_safe_num_with_float(self) -> None:
        """_safe_num(3.14) returns 3.14."""
        assert _stream_mod._safe_num(3.14) == 3.14

    def test_safe_num_with_numeric_string(self) -> None:
        """_safe_num('123') returns 123.0."""
        assert _stream_mod._safe_num("123") == 123.0

    def test_safe_num_with_non_numeric(self) -> None:
        """_safe_num('abc') returns None."""
        assert _stream_mod._safe_num("abc") is None
