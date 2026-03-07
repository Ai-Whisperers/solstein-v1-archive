"""Snapshot testing utilities using syrupy.

EPIC-029 Story 3: Snapshot testing for API responses and exports.

Usage:
    from tests.snapshots import CompanySnapshot

    def test_company_response(snapshot):
        company = CompanyFactory()
        response = serialize_company(company)
        assert response == snapshot
"""

import json
from pathlib import Path
from typing import Any

import pytest


class SnapshotMixin:
    """Mixin for snapshot testing utilities."""

    def exclude_fields(self, data: dict[str, Any], fields: list[str]) -> dict[str, Any]:
        """Exclude volatile fields from snapshot comparison.

        Args:
            data: Data dictionary.
            fields: Fields to exclude.

        Returns:
            Filtered data.
        """
        return {k: v for k, v in data.items() if k not in fields}

    def normalize_dates(self, data: dict[str, Any]) -> dict[str, Any]:
        """Normalize date fields for consistent snapshots.

        Args:
            data: Data dictionary.

        Returns:
            Data with normalized dates.
        """
        result = {}
        for k, v in data.items():
            if isinstance(v, str) and self._is_iso_date(v):
                result[k] = "<ISO_DATE>"
            elif isinstance(v, dict):
                result[k] = self.normalize_dates(v)
            elif isinstance(v, list):
                result[k] = [self.normalize_dates(item) if isinstance(item, dict) else item for item in v]
            else:
                result[k] = v
        return result

    def _is_iso_date(self, value: str) -> bool:
        """Check if string is ISO date format.

        Args:
            value: String to check.

        Returns:
            True if ISO date.
        """
        import re

        pattern = r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}"
        return bool(re.match(pattern, value))


class CompanySnapshot(SnapshotMixin):
    """Snapshot helper for company data."""

    EXCLUDED_FIELDS = ["id", "created_at", "updated_at", "analysis_date"]

    def prepare(self, company: Any) -> dict[str, Any]:
        """Prepare company for snapshot.

        Args:
            company: Company object.

        Returns:
            Serialized company data.
        """
        data = company.to_dict() if hasattr(company, "to_dict") else dict(company)
        data = self.exclude_fields(data, self.EXCLUDED_FIELDS)
        data = self.normalize_dates(data)
        return data


class APISnapshot(SnapshotMixin):
    """Snapshot helper for API responses."""

    EXCLUDED_FIELDS = ["id", "timestamp", "request_id", "trace_id"]

    def prepare(self, response: dict[str, Any]) -> dict[str, Any]:
        """Prepare API response for snapshot.

        Args:
            response: API response dict.

        Returns:
            Serialized response.
        """
        data = self.exclude_fields(response, self.EXCLUDED_FIELDS)
        data = self.normalize_dates(data)
        return data


class ExcelSnapshot(SnapshotMixin):
    """Snapshot helper for Excel exports."""

    def prepare(self, excel_data: bytes) -> dict[str, Any]:
        """Prepare Excel data for snapshot.

        Args:
            excel_data: Excel file bytes.

        Returns:
            Structure snapshot (not binary data).
        """
        try:
            import openpyxl
            from io import BytesIO

            wb = openpyxl.load_workbook(BytesIO(excel_data))
            structure = {"sheets": list(wb.sheetnames), "sheet_data": {}}

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                # Get headers from first row
                headers = [cell.value for cell in sheet[1]]
                structure["sheet_data"][sheet_name] = {
                    "headers": headers,
                    "row_count": sheet.max_row - 1,  # Excluding header
                }

            return structure
        except Exception:
            return {"error": "Could not parse Excel"}


class PDFSnapshot(SnapshotMixin):
    """Snapshot helper for PDF exports."""

    def prepare(self, pdf_data: bytes) -> dict[str, Any]:
        """Prepare PDF for snapshot.

        Args:
            pdf_data: PDF file bytes.

        Returns:
            PDF metadata snapshot.
        """
        try:
            from pypdf import PdfReader
            from io import BytesIO

            reader = PdfReader(BytesIO(pdf_data))
            return {
                "num_pages": len(reader.pages),
                "title": reader.metadata.title if reader.metadata else None,
                "pages": [page.extract_text()[:100] + "..." for page in reader.pages[:3]],
            }
        except Exception:
            return {"size_bytes": len(pdf_data)}


# Pytest fixtures
@pytest.fixture
def company_snapshot():
    """Fixture providing company snapshot helper.

    Returns:
        CompanySnapshot instance.
    """
    return CompanySnapshot()


@pytest.fixture
def api_snapshot():
    """Fixture providing API snapshot helper.

    Returns:
        APISnapshot instance.
    """
    return APISnapshot()


@pytest.fixture
def excel_snapshot():
    """Fixture providing Excel snapshot helper.

    Returns:
        ExcelSnapshot instance.
    """
    return ExcelSnapshot()


@pytest.fixture
def pdf_snapshot():
    """Fixture providing PDF snapshot helper.

    Returns:
        PDFSnapshot instance.
    """
    return PDFSnapshot()


def pytest_addoption(parser):
    """Add pytest options for snapshot testing."""
    parser.addoption("--update-snapshots", action="store_true", default=False, help="Update test snapshots")


# Configure syrupy
@pytest.fixture(scope="session")
def snapshot_config():
    """Configure snapshot testing."""
    return {
        "include": {
            "*.json",
            "*.txt",
            "*.yaml",
        },
        "exclude": {
            "*.pyc",
            "__pycache__",
        },
    }
