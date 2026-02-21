#!/usr/bin/env python3
"""Generate a formatted Excel workbook from extracted competitor data.

Produces a multi-sheet workbook with conditional formatting, auto-filters,
frozen headers, and charts for strategic analysis.

Usage:
    python generate_excel_report.py --input competitor_data.json --output financial-dashboard.xlsx
    python generate_excel_report.py --source tickets/COMPETITION/ --output financial-dashboard.xlsx
    python generate_excel_report.py --input data.json --output out.xlsx --profile

Requirements:
    Python 3.10+
    openpyxl >= 3.1.0
    rich >= 13.0 (optional, for progress bars)

Performance (29 competitors, 12 sheets):
    Total pipeline: ~0.45s
    Sheet writing: ~0.17s (37% of total, linear with competitor count)
    Workbook save: ~0.27s (60% of total, dominated by openpyxl XML serialization)
"""

import argparse
import json
import logging
import os
import re
import statistics
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

from competitor_utils import (
    CLASSIFICATION_ORDER,
    calc_hiring_efficiency,
    calc_growth_roi,
    calc_rev_per_employee_eur_k,
    calc_rev_per_eur_m_raised,
    get_acquihire_score,
    get_ai_capabilities,
    get_ai_in_production,
    get_ai_leadership_count,
    get_ai_score,
    get_ai_signal_level,
    get_ai_staff_pct,
    get_ai_talent_pct_total,
    get_ai_talent_team_size,
    get_classification,
    get_cloud_revenue_pct,
    get_composite,
    get_concentration_risk,
    get_countries_count,
    get_deployment_model,
    get_ebitda_margin,
    get_international_revenue_pct,
    get_key_hires_count,
    get_revenue_per_employee,
    get_score,
    get_talent_flow,
    has_ai_talent_data,
    is_eneve,
    parse_total_raised_eur_m,
    timed_phase,
)

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, BubbleChart, DoughnutChart, Reference, ScatterChart, Series
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print(
        "openpyxl is required. Install with: pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)

# Native sparklines disabled: openpyxl generates malformed sparkline XML that
# triggers Excel "Removed Records: Formula" repair warnings on open.  Text
# sparklines (Unicode block chars) are written to cells instead.
_HAS_NATIVE_SPARKLINES = False

try:
    from rich.console import Console
    from rich.progress import BarColumn, Progress, SpinnerColumn, TextColumn
    _HAS_RICH = True
except ImportError:
    _HAS_RICH = False

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
ENEVE_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
ROCKET_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
DINOSAUR_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
LEAD_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TRAIL_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
DATA_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
LINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")
LINK_BOLD_FONT = Font(name="Calibri", size=10, bold=True, color="0563C1", underline="single")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

REVENUE_FORMAT = '#,##0.0'
PERCENT_FORMAT = '0.0"%"'
SCORE_FORMAT = '0.0'

EXEC_TITLE_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
EXEC_TITLE_FONT = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
EXEC_SECTION_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
EXEC_SECTION_FONT = Font(name="Calibri", size=14, bold=True, color="1B2A4A")
EXEC_KPI_LABEL_FONT = Font(name="Calibri", size=10, color="808080")
EXEC_KPI_VALUE_FONT = Font(name="Calibri", size=24, bold=True, color="1B2A4A")
EXEC_KPI_SUBTITLE_FONT = Font(name="Calibri", size=9, color="A0A0A0")
EXEC_INSIGHT_FONT = Font(name="Calibri", size=12, bold=True, color="1B2A4A")


SPARK_CHARS = " _.-~^"


def iter_with_progress(items, description="Processing"):
    """Iterate with a progress bar (Rich) or simple stderr counter (fallback)."""
    if _HAS_RICH:
        with Progress(
            SpinnerColumn(),
            TextColumn("{task.description}"),
            BarColumn(),
            TextColumn("{task.completed}/{task.total}"),
            console=Console(stderr=True),
        ) as progress:
            task = progress.add_task(description, total=len(items))
            for item in items:
                yield item
                progress.advance(task)
    else:
        for i, item in enumerate(items, 1):
            print(f"\r{description}: {i}/{len(items)}", end="", file=sys.stderr)
            yield item
        print(file=sys.stderr)


def text_sparkline(values: list[Optional[float]]) -> str:
    """Create a text-based mini sparkline using Unicode block characters."""
    clean = [v for v in values if v is not None]
    if len(clean) < 2:
        return ""
    vmin, vmax = min(clean), max(clean)
    if vmax == vmin:
        return "―" * len(clean)
    chars = []
    for v in values:
        if v is None:
            chars.append(" ")
        else:
            idx = int((v - vmin) / (vmax - vmin) * (len(SPARK_CHARS) - 1))
            chars.append(SPARK_CHARS[idx])
    return "".join(chars)


def _get_max_timeline_length(competitors: list[dict]) -> int:
    """Return the maximum number of timeline data points across all competitors."""
    max_len = 0
    for comp in competitors:
        rev_tl = comp.get("revenue", {}).get("timeline", [])
        emp_tl = comp.get("employees", {}).get("timeline", [])
        max_len = max(max_len, len(rev_tl), len(emp_tl))
    return max_len


def _add_sparkline_columns(
    ws,
    competitors: list[dict],
    num_headers: int,
    max_years: int,
) -> None:
    """Write timeline data to hidden helper columns and attempt native sparkline groups.

    Assumes the last two visible headers are 'Revenue Trend' and 'Employee Trend'.
    Text sparkline cell values should already be written by the caller.
    """
    if max_years == 0:
        return

    helper_start = num_headers + 1
    rev_helper = helper_start
    emp_helper = helper_start + max_years
    rev_col = num_headers - 1
    emp_col = num_headers

    for row_idx, comp in enumerate(competitors, 2):
        for i, entry in enumerate(comp.get("revenue", {}).get("timeline", [])):
            ws.cell(row=row_idx, column=rev_helper + i, value=entry.get("eur_millions"))
        for i, entry in enumerate(comp.get("employees", {}).get("timeline", [])):
            ws.cell(row=row_idx, column=emp_helper + i, value=entry.get("headcount"))

    for col in range(rev_helper, emp_helper + max_years):
        ws.column_dimensions[get_column_letter(col)].hidden = True


def style_header_row(ws, num_cols: int) -> None:
    """Apply header styling to the first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(cell, is_eneve: bool = False, classification: Optional[str] = None, row_idx: int = 0) -> None:
    """Apply standard data cell styling with alternating row shading."""
    cell.font = BOLD_FONT if is_eneve else DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    if is_eneve:
        cell.fill = ENEVE_FILL
    elif classification == "Rocket":
        cell.fill = ROCKET_FILL
    elif classification == "Dinosaur":
        cell.fill = DINOSAUR_FILL
    elif row_idx % 2 == 0:
        cell.fill = ALT_ROW_FILL


def auto_fit_columns(ws, min_width: int = 10, max_width: int = 45) -> None:
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value or "")
            max_len = max(max_len, len(val))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def add_autofilter(ws, num_cols: int) -> None:
    """Add auto-filter to header row."""
    last_col = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"


def add_company_link(cell, comp: dict, link_base: Optional[Path]) -> None:
    """Add a hyperlink from a company cell to its financial-growth.md source file."""
    if link_base is None:
        return
    folder = comp.get("folder", "")
    if not folder:
        return
    target = str(link_base / folder / "financial-growth.md").replace("\\", "/")
    cell.hyperlink = target
    cell.font = LINK_BOLD_FONT if cell.font.bold else LINK_FONT


def format_value(val):
    """Return display value, using '--' for missing data."""
    if val is None:
        return "--"
    return val


def apply_number_format(cell, col_header: str) -> None:
    """Apply appropriate number format based on column header."""
    if not isinstance(cell.value, (int, float)):
        return
    header_lower = col_header.lower()
    if "revenue" in header_lower and ("eur" in header_lower or "employee" in header_lower):
        cell.number_format = REVENUE_FORMAT
    elif "%" in col_header or "cagr" in header_lower or "growth" in header_lower:
        cell.number_format = PERCENT_FORMAT
    elif "score" in header_lower or "composite" in header_lower:
        cell.number_format = SCORE_FORMAT


def setup_print_layout(ws, num_cols: int) -> None:
    """Configure sheet for professional printing."""
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.oddFooter.left.text = "Solstein by AI-Whisperers | Confidential"
    ws.oddFooter.center.text = "Generated &D"
    ws.oddFooter.right.text = "Page &P of &N"
    ws.print_title_rows = "1:1"
    last_col = get_column_letter(num_cols)
    ws.print_area = f"A1:{last_col}{ws.max_row}"


def add_trend_icons(ws, col_letter: str, last_row: int) -> None:
    """Add conditional formatting trend arrows on a growth column."""
    ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{last_row}",
        IconSetRule(
            icon_style="3Arrows",
            type="num",
            values=[0, 0, 0],
            showValue=True,
        ),
    )


def add_data_bars(ws, col_letter: str, last_row: int, max_val: int = 10) -> None:
    """Add conditional formatting data bars on a score column."""
    ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{last_row}",
        DataBarRule(
            start_type="num", start_value=0,
            end_type="num", end_value=max_val,
            color="B4C6E7",
        ),
    )


def write_executive_summary(wb: Workbook, competitors: list[dict], link_base: Optional[Path] = None) -> None:
    """Executive summary sheet with headline KPIs, top threats, and insight callouts.

    Produces a presentation-quality overview designed for CTO/Board viewing:
    - Headline KPI tiles (large fonts, color-coded)
    - Top 5 competitive threats table
    - Dynamic insight callouts computed from competitor data
    """
    ws = wb.create_sheet("Executive Summary")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # --- KPI Computation ---
    eneve = next((c for c in competitors if is_eneve(c)), None)
    non_eneve = [c for c in competitors if not is_eneve(c)]

    total_count = len(competitors)
    rockets = [c for c in competitors if get_classification(c) == "Rocket"]
    rocket_count = len(rockets)

    market_cagrs = [
        c.get("revenue", {}).get("cagr_3yr_pct")
        for c in non_eneve
        if c.get("revenue", {}).get("cagr_3yr_pct") is not None
    ]
    avg_market_cagr = sum(market_cagrs) / len(market_cagrs) if market_cagrs else None
    eneve_cagr = eneve.get("revenue", {}).get("cagr_3yr_pct") if eneve else None

    composites = [get_composite(c) for c in non_eneve if get_composite(c) is not None]
    avg_composite = sum(composites) / len(composites) if composites else None
    eneve_composite = get_composite(eneve) if eneve else None
    eneve_classification = get_classification(eneve) if eneve else None

    faster_than_eneve = len([
        c for c in non_eneve
        if (c.get("revenue", {}).get("cagr_3yr_pct") or 0) > (eneve_cagr or 0)
    ])
    funded_competitors = [
        c for c in non_eneve
        if (c.get("funding", {}).get("total_raised_text") or "").strip()
    ]
    funded_count = len(funded_competitors)

    # --- Column widths for 12-column layout ---
    for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
        ws.column_dimensions[col_letter].width = 14

    # --- Row 1-2: Title bar ---
    ws.merge_cells("A1:L2")
    title_cell = ws["A1"]
    title_cell.value = "SOLSTEIN - COMPETITIVE INTELLIGENCE DASHBOARD"
    title_cell.font = EXEC_TITLE_FONT
    title_cell.fill = EXEC_TITLE_FILL
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28

    # --- Row 3: Subtitle with branding and generation timestamp ---
    ws.merge_cells("A3:L3")
    subtitle_cell = ws["A3"]
    generated_ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    subtitle_cell.value = f"Powered by AI-Whisperers  |  Generated: {generated_ts}"
    subtitle_cell.font = Font(name="Calibri", size=11, italic=True, color="808080")
    subtitle_cell.alignment = center

    # --- Row 5-7: KPI tiles (5 tiles in merged column pairs) ---
    tiles = [
        ("A", "B", "COMPETITORS TRACKED", str(total_count), "Total in analysis", None),
        ("C", "D", "ROCKETS IDENTIFIED", str(rocket_count),
         f"of {total_count} competitors", "2E7D32"),
        ("E", "F", "REVENUE CAGR",
         f"{eneve_cagr:.1f}%" if eneve_cagr is not None else "N/A",
         f"Eneve vs Market avg {avg_market_cagr:.1f}%" if avg_market_cagr is not None else "Market avg N/A",
         None),
        ("G", "H", "COMPOSITE SCORE",
         f"{eneve_composite:.1f}" if eneve_composite is not None else "N/A",
         f"Eneve vs Market avg {avg_composite:.1f}" if avg_composite is not None else "Market avg N/A",
         None),
        ("I", "J", "ENEVE CLASSIFICATION", eneve_classification or "N/A",
         "Current standing", None),
    ]

    for c1, c2, label, value, subtitle, value_color in tiles:
        ws.merge_cells(f"{c1}4:{c2}4")
        label_cell = ws[f"{c1}4"]
        label_cell.value = label
        label_cell.font = EXEC_KPI_LABEL_FONT
        label_cell.alignment = center

        ws.merge_cells(f"{c1}5:{c2}5")
        value_cell = ws[f"{c1}5"]
        value_cell.value = value
        value_cell.font = (
            Font(name="Calibri", size=24, bold=True, color=value_color)
            if value_color
            else EXEC_KPI_VALUE_FONT
        )
        value_cell.alignment = center

        ws.merge_cells(f"{c1}6:{c2}6")
        sub_cell = ws[f"{c1}6"]
        sub_cell.value = subtitle
        sub_cell.font = EXEC_KPI_SUBTITLE_FONT
        sub_cell.alignment = center

    # Green highlight on Rockets tile
    ws["C5"].fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    # Color badge on Eneve classification tile
    cls_fills = {
        "Rocket": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Riser": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        "Steady": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "Dinosaur": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }
    if eneve_classification in cls_fills:
        ws["I5"].fill = cls_fills[eneve_classification]

    ws.row_dimensions[5].height = 36

    # --- Row 8: Section header ---
    ws.merge_cells("A8:L8")
    section_cell = ws["A8"]
    section_cell.value = "TOP 5 COMPETITIVE THREATS"
    section_cell.font = EXEC_SECTION_FONT
    section_cell.fill = EXEC_SECTION_FILL
    section_cell.alignment = center

    # --- Row 9: Threat table headers ---
    threat_headers = [
        "Rank", "Company", "Classification",
        "Composite Score", "Revenue CAGR %", "Latest Revenue (EUR M)",
    ]
    for col, h in enumerate(threat_headers, 1):
        cell = ws.cell(row=9, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = center
        cell.border = THIN_BORDER

    # --- Rows 10-14: Top 5 threats (highest composite, excluding Eneve) ---
    top_threats = sorted(
        [c for c in non_eneve if get_composite(c) is not None],
        key=lambda c: get_composite(c) or 0,
        reverse=True,
    )[:5]

    for idx, comp in enumerate(top_threats):
        row = 10 + idx
        cls = get_classification(comp)
        rev = comp.get("revenue", {})
        values = [
            idx + 1,
            comp.get("company_name", ""),
            cls or "",
            get_composite(comp),
            rev.get("cagr_3yr_pct"),
            rev.get("latest_revenue_eur_m"),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = left_align
            if cls == "Rocket":
                cell.fill = ROCKET_FILL

    # --- Rows 16-18: Insight callouts ---
    callouts = [
        f"{faster_than_eneve} competitors grew revenue faster than Eneve",
        f"{funded_count} competitors have raised external funding",
        f"{rocket_count} competitors are classified as Rockets",
    ]
    for i, text in enumerate(callouts):
        row = 16 + i
        ws.merge_cells(f"A{row}:L{row}")
        cell = ws[f"A{row}"]
        cell.value = text
        cell.font = EXEC_INSIGHT_FONT
        cell.alignment = center

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.oddFooter.left.text = "Solstein by AI-Whisperers | Confidential"
    ws.oddFooter.center.text = "Generated &D"
    ws.oddFooter.right.text = "Page &P of &N"
    ws.print_area = "A1:L18"


def write_summary_sheet(wb: Workbook, competitors: list[dict], link_base: Optional[Path] = None) -> None:
    """Summary sheet with all competitors and scorecard dimensions."""
    ws = wb.active
    ws.title = "Summary"

    headers = [
        "Company", "Tier", "Classification",
        "Revenue Growth", "Funding Momentum", "Employee Growth",
        "Geographic Expansion", "M&A Activity", "SaaS Maturity",
        "Composite Score", "Latest Revenue (EUR M)", "Latest Headcount",
        "Revenue Trend", "Employee Trend",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"

    max_years = _get_max_timeline_length(competitors)

    for row_idx, comp in enumerate(competitors, 2):
        cls = get_classification(comp)
        eneve = is_eneve(comp)
        rev_tl = [e.get("eur_millions") for e in comp.get("revenue", {}).get("timeline", [])]
        emp_tl = [e.get("headcount") for e in comp.get("employees", {}).get("timeline", [])]
        values = [
            comp.get("company_name", ""),
            comp.get("tier", ""),
            cls or "",
            get_score(comp, "Revenue Growth"),
            get_score(comp, "Funding Momentum"),
            get_score(comp, "Employee Growth"),
            get_score(comp, "Geographic Expansion"),
            get_score(comp, "M&A Activity"),
            get_score(comp, "SaaS Maturity"),
            get_composite(comp),
            comp.get("revenue", {}).get("latest_revenue_eur_m"),
            comp.get("employees", {}).get("latest_headcount"),
            text_sparkline(rev_tl),
            text_sparkline(emp_tl),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=format_value(val))
            style_data_cell(cell, is_eneve=eneve, classification=cls, row_idx=row_idx)
            apply_number_format(cell, headers[col - 1])
        add_company_link(ws.cell(row=row_idx, column=1), comp, link_base)

    if max_years > 0:
        _add_sparkline_columns(ws, competitors, len(headers), max_years)

    # Color scale on score columns (D through I) and composite (J)
    last_row = len(competitors) + 1
    for col_letter in ["D", "E", "F", "G", "H", "I", "J"]:
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{last_row}",
            ColorScaleRule(
                start_type="num", start_value=1, start_color="F8696B",
                mid_type="num", mid_value=5, mid_color="FFEB84",
                end_type="num", end_value=10, end_color="63BE7B",
            ),
        )
        add_data_bars(ws, col_letter, last_row)

    auto_fit_columns(ws)
    add_autofilter(ws, len(headers))
    setup_print_layout(ws, len(headers))

    if len(competitors) >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Composite Growth Score - All Competitors"
        chart.y_axis.title = "Score (1-10)"
        chart.x_axis.title = "Company"
        chart.style = 10

        data_ref = Reference(ws, min_col=10, min_row=1, max_row=len(competitors) + 1)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(competitors) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        chart.width = 25
        chart.height = 14

        chart_col = get_column_letter(len(headers) + 2 * max_years + 2)
        ws.add_chart(chart, f"{chart_col}2")


def write_revenue_leaderboard(wb: Workbook, competitors: list[dict], link_base: Optional[Path] = None) -> None:
    """Revenue leaderboard sorted by CAGR."""
    ws = wb.create_sheet("Revenue Leaderboard")

    headers = ["Rank", "Company", "Tier", "Latest Revenue (EUR M)", "Revenue CAGR 3yr (%)", "Classification"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"

    sorted_comps = sorted(
        competitors,
        key=lambda c: c.get("revenue", {}).get("cagr_3yr_pct") or 0,
        reverse=True,
    )

    for row_idx, comp in enumerate(sorted_comps, 2):
        rank = row_idx - 1
        cls = get_classification(comp)
        eneve = is_eneve(comp)
        rev = comp.get("revenue", {})
        values = [
            rank,
            comp.get("company_name", ""),
            comp.get("tier", ""),
            rev.get("latest_revenue_eur_m"),
            rev.get("cagr_3yr_pct"),
            cls or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=format_value(val))
            style_data_cell(cell, is_eneve=eneve, classification=cls, row_idx=row_idx)
            apply_number_format(cell, headers[col - 1])
        add_company_link(ws.cell(row=row_idx, column=2), comp, link_base)

    last_row = len(sorted_comps) + 1
    add_trend_icons(ws, "E", last_row)

    auto_fit_columns(ws)
    add_autofilter(ws, len(headers))
    setup_print_layout(ws, len(headers))

    # Bar chart for Revenue CAGR
    if len(sorted_comps) >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Revenue CAGR 3yr (%) - All Competitors"
        chart.y_axis.title = "CAGR %"
        chart.x_axis.title = "Company"
        chart.style = 10

        data_ref = Reference(ws, min_col=5, min_row=1, max_row=len(sorted_comps) + 1)
        cats_ref = Reference(ws, min_col=2, min_row=2, max_row=len(sorted_comps) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        chart.width = 25
        chart.height = 14

        ws.add_chart(chart, "H2")


def write_funding_leaderboard(wb: Workbook, competitors: list[dict], link_base: Optional[Path] = None) -> None:
    """Funding leaderboard sorted by funding score."""
    ws = wb.create_sheet("Funding Leaderboard")

    headers = [
        "Rank", "Company", "Tier", "Funding Score",
        "Total Raised", "Latest Valuation", "Classification",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"

    sorted_comps = sorted(
        competitors,
        key=lambda c: get_score(c, "Funding Momentum") or 0,
        reverse=True,
    )

    for row_idx, comp in enumerate(sorted_comps, 2):
        rank = row_idx - 1
        cls = get_classification(comp)
        eneve = is_eneve(comp)
        funding = comp.get("funding", {})
        values = [
            rank,
            comp.get("company_name", ""),
            comp.get("tier", ""),
            get_score(comp, "Funding Momentum"),
            funding.get("total_raised_text", ""),
            funding.get("latest_valuation_text", ""),
            cls or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=format_value(val))
            style_data_cell(cell, is_eneve=eneve, classification=cls, row_idx=row_idx)
            apply_number_format(cell, headers[col - 1])
        add_company_link(ws.cell(row=row_idx, column=2), comp, link_base)

    last_row = len(sorted_comps) + 1
    add_data_bars(ws, "D", last_row)

    auto_fit_columns(ws)
    add_autofilter(ws, len(headers))
    setup_print_layout(ws, len(headers))

    if len(sorted_comps) >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Funding Momentum Score - All Competitors"
        chart.y_axis.title = "Score (1-10)"
        chart.x_axis.title = "Company"
        chart.style = 10

        data_ref = Reference(ws, min_col=4, min_row=1, max_row=len(sorted_comps) + 1)
        cats_ref = Reference(ws, min_col=2, min_row=2, max_row=len(sorted_comps) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        chart.width = 25
        chart.height = 14

        ws.add_chart(chart, "I2")


def write_employee_growth(wb: Workbook, competitors: list[dict], link_base: Optional[Path] = None) -> None:
    """Employee growth leaderboard."""
    ws = wb.create_sheet("Employee Growth")

    headers = [
        "Rank", "Company", "Tier", "Latest Headcount",
        "Employee CAGR (%)", "Open Positions", "Classification",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"

    sorted_comps = sorted(
        competitors,
        key=lambda c: c.get("employees", {}).get("employee_cagr_pct") or 0,
        reverse=True,
    )

    for row_idx, comp in enumerate(sorted_comps, 2):
        rank = row_idx - 1
        cls = get_classification(comp)
        eneve = is_eneve(comp)
        emp = comp.get("employees", {})
        values = [
            rank,
            comp.get("company_name", ""),
            comp.get("tier", ""),
            emp.get("latest_headcount"),
            emp.get("employee_cagr_pct"),
            emp.get("open_positions"),
            cls or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=format_value(val))
            style_data_cell(cell, is_eneve=eneve, classification=cls, row_idx=row_idx)
            apply_number_format(cell, headers[col - 1])
        add_company_link(ws.cell(row=row_idx, column=2), comp, link_base)

    last_row = len(sorted_comps) + 1
    add_trend_icons(ws, "E", last_row)

    auto_fit_columns(ws)
    add_autofilter(ws, len(headers))
    setup_print_layout(ws, len(headers))

    if len(sorted_comps) >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Employee CAGR (%) - All Competitors"
        chart.y_axis.title = "CAGR %"
        chart.x_axis.title = "Company"
        chart.style = 10

        data_ref = Reference(ws, min_col=5, min_row=1, max_row=len(sorted_comps) + 1)
        cats_ref = Reference(ws, min_col=2, min_row=2, max_row=len(sorted_comps) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        chart.width = 25
        chart.height = 14

        ws.add_chart(chart, "H2")


def write_saas_maturity(wb: Workbook, competitors: list[dict], link_base: Optional[Path] = None) -> None:
    """SaaS maturity ranking."""
    ws = wb.create_sheet("SaaS Maturity")

    headers = [
        "Rank", "Company", "Tier", "SaaS Score",
        "Recurring Revenue %", "Classification",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"

    sorted_comps = sorted(
        competitors,
        key=lambda c: get_score(c, "SaaS Maturity") or 0,
        reverse=True,
    )

    for row_idx, comp in enumerate(sorted_comps, 2):
        rank = row_idx - 1
        cls = get_classification(comp)
        eneve = is_eneve(comp)
        prof = comp.get("profitability", {})
        values = [
            rank,
            comp.get("company_name", ""),
            comp.get("tier", ""),
            get_score(comp, "SaaS Maturity"),
            prof.get("recurring_revenue_pct"),
            cls or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=format_value(val))
            style_data_cell(cell, is_eneve=eneve, classification=cls, row_idx=row_idx)
            apply_number_format(cell, headers[col - 1])
        add_company_link(ws.cell(row=row_idx, column=2), comp, link_base)

    last_row = len(sorted_comps) + 1
    add_data_bars(ws, "D", last_row)

    auto_fit_columns(ws)
    add_autofilter(ws, len(headers))
    setup_print_layout(ws, len(headers))

    if len(sorted_comps) >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.title = "SaaS Maturity Score - All Competitors"
        chart.y_axis.title = "Score (1-10)"
        chart.x_axis.title = "Company"
        chart.style = 10

        data_ref = Reference(ws, min_col=4, min_row=1, max_row=len(sorted_comps) + 1)
        cats_ref = Reference(ws, min_col=2, min_row=2, max_row=len(sorted_comps) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        chart.width = 25
        chart.height = 14

        ws.add_chart(chart, "H2")


def write_classification_matrix(wb: Workbook, competitors: list[dict], link_base: Optional[Path] = None) -> None:
    """Classification matrix grouped by Rocket/Riser/Steady/Dinosaur."""
    ws = wb.create_sheet("Classification Matrix")

    headers = [
        "Classification", "Company", "Tier", "Composite Score",
        "Revenue Growth", "Funding", "Employees", "Geography", "M&A", "SaaS",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"

    grouped = sorted(
        competitors,
        key=lambda c: (
            CLASSIFICATION_ORDER.index(get_classification(c))
            if get_classification(c) in CLASSIFICATION_ORDER
            else 99,
            -(get_composite(c) or 0),
        ),
    )

    for row_idx, comp in enumerate(grouped, 2):
        cls = get_classification(comp)
        eneve = is_eneve(comp)
        values = [
            cls or "Unclassified",
            comp.get("company_name", ""),
            comp.get("tier", ""),
            get_composite(comp),
            get_score(comp, "Revenue Growth"),
            get_score(comp, "Funding Momentum"),
            get_score(comp, "Employee Growth"),
            get_score(comp, "Geographic Expansion"),
            get_score(comp, "M&A Activity"),
            get_score(comp, "SaaS Maturity"),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=format_value(val))
            style_data_cell(cell, is_eneve=eneve, classification=cls, row_idx=row_idx)
            apply_number_format(cell, headers[col - 1])
        add_company_link(ws.cell(row=row_idx, column=2), comp, link_base)

    last_row = len(grouped) + 1
    for col_letter in ["D", "E", "F", "G", "H", "I", "J"]:
        add_data_bars(ws, col_letter, last_row)

    auto_fit_columns(ws)
    add_autofilter(ws, len(headers))
    setup_print_layout(ws, len(headers))

    # Doughnut chart: write classification counts to helper columns (M-N)
    classification_counts: dict[str, int] = {}
    for comp in grouped:
        cls = get_classification(comp) or "Unclassified"
        classification_counts[cls] = classification_counts.get(cls, 0) + 1

    if classification_counts:
        helper_label_col = 13  # Column M
        helper_value_col = 14  # Column N
        ws.cell(row=1, column=helper_label_col, value="Classification")
        ws.cell(row=1, column=helper_value_col, value="Count")
        for i, (cls_name, count) in enumerate(classification_counts.items(), 2):
            ws.cell(row=i, column=helper_label_col, value=cls_name)
            ws.cell(row=i, column=helper_value_col, value=count)

        num_categories = len(classification_counts)
        chart = DoughnutChart()
        chart.title = "Competitor Classification Distribution"
        chart.style = 10
        chart.width = 20
        chart.height = 14

        data_ref = Reference(ws, min_col=helper_value_col, min_row=1, max_row=num_categories + 1)
        cats_ref = Reference(ws, min_col=helper_label_col, min_row=2, max_row=num_categories + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        ws.add_chart(chart, "L2")


def write_efficiency_sheet(wb: Workbook, competitors: list[dict], link_base: Optional[Path] = None) -> None:
    """Efficiency & Profitability leaderboard sorted by Revenue per Employee."""
    ws = wb.create_sheet("Efficiency & Profitability")

    headers = [
        "Rank", "Company", "Tier", "Revenue/Employee (EUR K)",
        "EBITDA Margin (%)", "Recurring Revenue (%)", "SaaS Score", "Classification",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"

    sorted_comps = sorted(
        competitors,
        key=lambda c: get_revenue_per_employee(c) or 0,
        reverse=True,
    )

    for row_idx, comp in enumerate(sorted_comps, 2):
        rank = row_idx - 1
        cls = get_classification(comp)
        eneve = is_eneve(comp)
        values = [
            rank,
            comp.get("company_name", ""),
            comp.get("tier", ""),
            get_revenue_per_employee(comp),
            get_ebitda_margin(comp),
            comp.get("profitability", {}).get("recurring_revenue_pct"),
            get_score(comp, "SaaS Maturity"),
            cls or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=format_value(val))
            style_data_cell(cell, is_eneve=eneve, classification=cls, row_idx=row_idx)
            apply_number_format(cell, headers[col - 1])
        add_company_link(ws.cell(row=row_idx, column=2), comp, link_base)

    last_row = len(sorted_comps) + 1
    add_data_bars(ws, "G", last_row)

    auto_fit_columns(ws)
    add_autofilter(ws, len(headers))
    setup_print_layout(ws, len(headers))

    if len(sorted_comps) >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Revenue per Employee (EUR K) - All Competitors"
        chart.y_axis.title = "EUR K"
        chart.x_axis.title = "Company"
        chart.style = 10

        data_ref = Reference(ws, min_col=4, min_row=1, max_row=len(sorted_comps) + 1)
        cats_ref = Reference(ws, min_col=2, min_row=2, max_row=len(sorted_comps) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        chart.width = 25
        chart.height = 14

        ws.add_chart(chart, "J2")


# Quartile fill styles for investment efficiency conditional highlighting
Q_TOP_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
Q_BOTTOM_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _compute_quartile_bounds(values: list[float]) -> tuple[float, float]:
    """Return (Q1, Q3) thresholds from a list of non-None numeric values."""
    if len(values) < 4:
        return (float("-inf"), float("inf"))
    s = sorted(values)
    return s[len(s) // 4], s[3 * len(s) // 4]


def write_investment_efficiency_sheet(wb: Workbook, competitors: list[dict], link_base: Optional[Path] = None) -> None:
    """Investment Efficiency Ratios sheet with capital efficiency metrics and chart."""
    ws = wb.create_sheet("Investment Efficiency")

    headers = [
        "Rank", "Company", "Tier", "Revenue (EUR M)", "Headcount",
        "Total Raised", "Rev/Emp (EUR K)", "Rev/EUR M Raised",
        "Hiring Eff.", "Growth ROI", "Classification",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"

    enriched = []
    for c in competitors:
        enriched.append({
            "comp": c,
            "rev_per_emp": calc_rev_per_employee_eur_k(c),
            "rev_per_raised": calc_rev_per_eur_m_raised(c),
            "hiring_eff": calc_hiring_efficiency(c),
            "growth_roi": calc_growth_roi(c),
        })

    sorted_data = sorted(enriched, key=lambda x: x["rev_per_emp"] or 0, reverse=True)

    rpe_vals = [e["rev_per_emp"] for e in enriched if e["rev_per_emp"] is not None]
    rpr_vals = [e["rev_per_raised"] for e in enriched if e["rev_per_raised"] is not None]
    he_vals = [e["hiring_eff"] for e in enriched if e["hiring_eff"] is not None]

    rpe_q1, rpe_q3 = _compute_quartile_bounds(rpe_vals)
    rpr_q1, rpr_q3 = _compute_quartile_bounds(rpr_vals)
    he_q1, he_q3 = _compute_quartile_bounds(he_vals)

    for row_idx, entry in enumerate(sorted_data, 2):
        rank = row_idx - 1
        c = entry["comp"]
        cls = get_classification(c)
        eneve = is_eneve(c)
        rev_m = c.get("revenue", {}).get("latest_revenue_eur_m")
        hc = c.get("employees", {}).get("latest_headcount")
        raised_text = c.get("funding", {}).get("total_raised_text") or ""

        values = [
            rank,
            c.get("company_name", ""),
            c.get("tier", ""),
            rev_m,
            hc,
            raised_text,
            entry["rev_per_emp"],
            entry["rev_per_raised"],
            entry["hiring_eff"],
            entry["growth_roi"],
            cls or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=format_value(val))
            style_data_cell(cell, is_eneve=eneve, classification=cls, row_idx=row_idx)
            apply_number_format(cell, headers[col - 1])
        add_company_link(ws.cell(row=row_idx, column=2), c, link_base)

        # Quartile highlights on ratio columns (override base styling for non-Eneve)
        if not eneve:
            rpe = entry["rev_per_emp"]
            if rpe is not None:
                if rpe >= rpe_q3:
                    ws.cell(row=row_idx, column=7).fill = Q_TOP_FILL
                elif rpe <= rpe_q1:
                    ws.cell(row=row_idx, column=7).fill = Q_BOTTOM_FILL

            rpr = entry["rev_per_raised"]
            if rpr is not None:
                if rpr >= rpr_q3:
                    ws.cell(row=row_idx, column=8).fill = Q_TOP_FILL
                elif rpr <= rpr_q1:
                    ws.cell(row=row_idx, column=8).fill = Q_BOTTOM_FILL

            # Hiring efficiency: lower is better
            he = entry["hiring_eff"]
            if he is not None:
                if he <= he_q1:
                    ws.cell(row=row_idx, column=9).fill = Q_TOP_FILL
                elif he >= he_q3:
                    ws.cell(row=row_idx, column=9).fill = Q_BOTTOM_FILL

    last_row = len(sorted_data) + 1

    # Rev/Emp number format
    for row in range(2, last_row + 1):
        cell = ws.cell(row=row, column=7)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0'
        cell = ws.cell(row=row, column=8)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '0.00'
        cell = ws.cell(row=row, column=9)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '0.00'
        cell = ws.cell(row=row, column=10)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '0.00'

    auto_fit_columns(ws)
    add_autofilter(ws, len(headers))
    setup_print_layout(ws, len(headers))

    # Bar chart: Revenue per Employee for all competitors
    if len(sorted_data) >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Revenue per Employee (EUR K) - Sorted by Efficiency"
        chart.y_axis.title = "EUR K / Employee"
        chart.x_axis.title = "Company"
        chart.style = 10

        data_ref = Reference(ws, min_col=7, min_row=1, max_row=last_row)
        cats_ref = Reference(ws, min_col=2, min_row=2, max_row=last_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        chart.width = 28
        chart.height = 16

        ws.add_chart(chart, "M2")


def write_market_reach_sheet(wb: Workbook, competitors: list[dict], link_base: Optional[Path] = None) -> None:
    """Market Reach & Geographic Expansion leaderboard sorted by International Revenue %."""
    ws = wb.create_sheet("Market Reach")

    headers = [
        "Rank", "Company", "Tier", "International Revenue (%)",
        "Countries Active", "Deployment Model", "Cloud Revenue (%)",
        "Geographic Score", "Classification",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"

    sorted_comps = sorted(
        competitors,
        key=lambda c: get_international_revenue_pct(c) or 0,
        reverse=True,
    )

    for row_idx, comp in enumerate(sorted_comps, 2):
        rank = row_idx - 1
        cls = get_classification(comp)
        eneve = is_eneve(comp)
        values = [
            rank,
            comp.get("company_name", ""),
            comp.get("tier", ""),
            get_international_revenue_pct(comp),
            get_countries_count(comp),
            get_deployment_model(comp),
            get_cloud_revenue_pct(comp),
            get_score(comp, "Geographic Expansion"),
            cls or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=format_value(val))
            style_data_cell(cell, is_eneve=eneve, classification=cls, row_idx=row_idx)
            apply_number_format(cell, headers[col - 1])
        add_company_link(ws.cell(row=row_idx, column=2), comp, link_base)

    last_row = len(sorted_comps) + 1
    add_data_bars(ws, "H", last_row)

    auto_fit_columns(ws)
    add_autofilter(ws, len(headers))
    setup_print_layout(ws, len(headers))

    if len(sorted_comps) >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.title = "International Revenue (%) - All Competitors"
        chart.y_axis.title = "International Revenue %"
        chart.x_axis.title = "Company"
        chart.style = 10

        data_ref = Reference(ws, min_col=4, min_row=1, max_row=len(sorted_comps) + 1)
        cats_ref = Reference(ws, min_col=2, min_row=2, max_row=len(sorted_comps) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        chart.width = 25
        chart.height = 14

        ws.add_chart(chart, "K2")


AI_SIGNAL_LEVELS = {
    "None": 0,
    "Low": 1,
    "Moderate": 2,
    "Strong": 3,
    "Very Strong": 4,
}


def write_ai_maturity_sheet(wb: Workbook, competitors: list[dict], link_base: Optional[Path] = None) -> None:
    """AI Maturity Matrix ranking all competitors by AI adoption score.

    Includes heatmap conditional formatting, Eneve highlight, and a horizontal
    bar chart sorted by AI score descending.
    """
    ws = wb.create_sheet("AI Maturity")

    headers = [
        "Rank", "Company", "Tier", "AI Score", "AI Signal Level",
        "Key AI Capabilities", "AI Staff %", "AI in Production",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"

    sorted_comps = sorted(
        competitors,
        key=lambda c: (get_ai_score(c) or 0, AI_SIGNAL_LEVELS.get(get_ai_signal_level(c) or "None", 0)),
        reverse=True,
    )

    for row_idx, comp in enumerate(sorted_comps, 2):
        rank = row_idx - 1
        cls = get_classification(comp)
        eneve = is_eneve(comp)
        ai_score = get_ai_score(comp)
        signal = get_ai_signal_level(comp)
        staff_pct = get_ai_staff_pct(comp)
        in_prod = get_ai_in_production(comp)
        caps = get_ai_capabilities(comp)
        no_data = comp.get("ai", {}).get("data_source") == "No Data" and ai_score == 0

        values = [
            rank,
            comp.get("company_name", ""),
            comp.get("tier", ""),
            ai_score if ai_score is not None else 0,
            f"{signal} (No Data)" if no_data and signal == "None" else (signal or "None"),
            caps or ("No Data" if no_data else "--"),
            staff_pct,
            "Y" if in_prod else "N",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=format_value(val))
            style_data_cell(cell, is_eneve=eneve, classification=cls, row_idx=row_idx)
            apply_number_format(cell, headers[col - 1])
        add_company_link(ws.cell(row=row_idx, column=2), comp, link_base)

    last_row = len(sorted_comps) + 1

    # Heatmap: red-yellow-green gradient on AI Score column (D)
    ws.conditional_formatting.add(
        f"D2:D{last_row}",
        ColorScaleRule(
            start_type="num", start_value=0, start_color="F8696B",
            mid_type="num", mid_value=5, mid_color="FFEB84",
            end_type="num", end_value=10, end_color="63BE7B",
        ),
    )

    # Horizontal bar chart sorted by AI score descending
    if len(sorted_comps) >= 2:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "AI Maturity Score - All Competitors"
        chart.x_axis.title = "AI Score (0-10)"
        chart.y_axis.title = "Company"
        chart.style = 10

        data_ref = Reference(ws, min_col=4, min_row=1, max_row=last_row)
        cats_ref = Reference(ws, min_col=2, min_row=2, max_row=last_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        chart.width = 28
        chart.height = 20

        ws.add_chart(chart, "J2")

    auto_fit_columns(ws)
    add_autofilter(ws, len(headers))
    setup_print_layout(ws, len(headers))


# --- M&A Landscape classification constants ---

MNA_ACQUIRER_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
MNA_TARGET_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
MNA_SUSTAINING_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

CONFIDENTIAL_FILL = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
CONFIDENTIAL_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
HIGH_RISK_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ACQUIHIRE_TARGET_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

MNA_ROLE_ORDER = ["Likely Acquirer", "Likely Target", "Neutral", "Self-Sustaining"]

KNOWN_MNA_NARRATIVES: dict[str, str] = {
    "hansen": (
        "Could acquire small energy software using powercloud playbook "
        "(EUR 30M template); A$59M annual cash flow"
    ),
    "octopus": (
        "15+ acquisitions globally; acquired Jedlix (Rotterdam) — "
        "Dutch tech hub with $1B fresh capital (Dec 2025)"
    ),
    "volue": (
        "Aggressive M&A post-delisting (PowerBot + Quorum); "
        "EUR 1.5B valuation enables further bolt-on acquisitions"
    ),
    "asseco": (
        "10+ acquisitions pattern; acquires market access in new regions "
        "through bolt-on acquisitions"
    ),
    "eg": (
        "EUR 896M total capital deployed; acquired Bright Energy (Dec 2025) "
        "for AI-driven energy management"
    ),
    "cgi": (
        "Systematic acquirer with $2.2B+ annual cash flow; "
        "global consulting firm absorbing energy software targets"
    ),
}


def _classify_mna(comp: dict) -> str:
    """Classify competitor M&A role using scoring criteria from plan requirements.

    Acquirer: M&A score >= 7 (proven acquisition track record)
    Target:   revenue < EUR 50M AND unfunded AND composite < 5 AND SaaS <= 5
    Self-Sustaining: composite >= 5 AND unfunded AND revenue >= EUR 40M
    Neutral:  everything else
    """
    if is_eneve(comp):
        return "Neutral"

    mna = get_score(comp, "M&A Activity") or 0
    composite = get_composite(comp) or 0
    saas = get_score(comp, "SaaS Maturity") or 0
    rev_m = comp.get("revenue", {}).get("latest_revenue_eur_m")
    raised = parse_total_raised_eur_m(
        comp.get("funding", {}).get("total_raised_text"),
    )
    is_unfunded = raised is None or raised == 0

    if mna >= 7:
        return "Likely Acquirer"

    if (
        rev_m is not None
        and rev_m < 50
        and is_unfunded
        and composite < 5
        and saas <= 5
    ):
        return "Likely Target"

    if composite >= 5 and is_unfunded and (rev_m is None or rev_m >= 40):
        return "Self-Sustaining"

    return "Neutral"


def _mna_rationale(comp: dict, role: str) -> str:
    """Generate a brief rationale explaining why the competitor got this M&A role."""
    folder = comp.get("folder", "").lower()
    segments = folder.split("-")
    for key, narrative in KNOWN_MNA_NARRATIVES.items():
        if key in segments:
            return narrative

    mna = get_score(comp, "M&A Activity") or 0
    composite = get_composite(comp) or 0
    saas = get_score(comp, "SaaS Maturity") or 0
    rev_m = comp.get("revenue", {}).get("latest_revenue_eur_m")
    raised = parse_total_raised_eur_m(
        comp.get("funding", {}).get("total_raised_text"),
    )

    if role == "Likely Acquirer":
        return f"M&A score {mna}/10, active acquisition history"

    if role == "Likely Target":
        parts = []
        if rev_m is not None:
            parts.append(f"EUR {rev_m:.0f}M rev")
        if raised is None or raised == 0:
            parts.append("no external funding")
        else:
            parts.append(f"EUR {raised:.0f}M raised")
        parts.append(f"composite {composite:.1f}, SaaS {saas}/10")
        return "Vulnerable: " + ", ".join(parts)

    if role == "Self-Sustaining":
        parts = [f"composite {composite:.1f}"]
        if rev_m is not None:
            parts.append(f"EUR {rev_m:.0f}M revenue")
        parts.append("self-funded, stable operations")
        return "Independent: " + ", ".join(parts)

    return f"Composite {composite:.1f}, no strong M&A signals"


def _cash_flow_indicator(comp: dict) -> str:
    """Derive a qualitative cash flow indicator from available metrics."""
    funding = get_score(comp, "Funding Momentum") or 0
    rev_m = comp.get("revenue", {}).get("latest_revenue_eur_m")
    if funding >= 8:
        return "Strong (funded)"
    if rev_m is not None and rev_m >= 100:
        return "Strong (large co)"
    if rev_m is not None and rev_m >= 30:
        return "Moderate"
    return "Limited"


def write_mna_landscape_sheet(
    wb: Workbook,
    competitors: list[dict],
    link_base: Optional[Path] = None,
) -> None:
    """M&A Landscape sheet classifying all competitors by acquisition role.

    Produces a table with classification, rationale, cash flow indicators,
    and a scatter chart plotting Funding Score vs Revenue color-coded
    by M&A role (Acquirer=blue, Target=red, Neutral=grey, Self-Sustaining=green).
    """
    ws = wb.create_sheet("M&A Landscape")

    headers = [
        "Company", "Tier", "M&A Role", "Revenue (EUR M)",
        "Funding Score", "M&A Score", "Cash Flow Indicator",
        "Acquirer/Target Rationale",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"

    enriched = []
    for c in competitors:
        role = _classify_mna(c)
        enriched.append({
            "comp": c,
            "role": role,
            "rationale": _mna_rationale(c, role),
        })

    enriched.sort(
        key=lambda x: (
            MNA_ROLE_ORDER.index(x["role"])
            if x["role"] in MNA_ROLE_ORDER
            else 99,
            -(get_composite(x["comp"]) or 0),
        ),
    )

    role_fills = {
        "Likely Acquirer": MNA_ACQUIRER_FILL,
        "Likely Target": MNA_TARGET_FILL,
        "Self-Sustaining": MNA_SUSTAINING_FILL,
    }

    for row_idx, entry in enumerate(enriched, 2):
        c = entry["comp"]
        role = entry["role"]
        eneve = is_eneve(c)
        values = [
            c.get("company_name", ""),
            c.get("tier", ""),
            role,
            c.get("revenue", {}).get("latest_revenue_eur_m"),
            get_score(c, "Funding Momentum"),
            get_score(c, "M&A Activity"),
            _cash_flow_indicator(c),
            entry["rationale"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=format_value(val))
            cell.font = BOLD_FONT if eneve else DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if eneve:
                cell.fill = ENEVE_FILL
            elif role in role_fills:
                cell.fill = role_fills[role]
            elif row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL
            apply_number_format(cell, headers[col - 1])
        add_company_link(ws.cell(row=row_idx, column=1), c, link_base)

    last_row = len(enriched) + 1
    add_data_bars(ws, "F", last_row)

    auto_fit_columns(ws)
    ws.column_dimensions["H"].width = 55
    add_autofilter(ws, len(headers))
    setup_print_layout(ws, len(headers))

    # --- Scatter chart: Funding Score (x) vs Revenue (y) by M&A role ---
    if len(enriched) >= 2:
        grouped: dict[str, list[tuple[float, float]]] = {}
        for entry in enriched:
            c = entry["comp"]
            fs = get_score(c, "Funding Momentum")
            rv = c.get("revenue", {}).get("latest_revenue_eur_m")
            if fs is not None and rv is not None:
                grouped.setdefault(entry["role"], []).append((fs, rv))

        chart = ScatterChart()
        chart.title = "M&A Landscape: Funding vs Revenue"
        chart.x_axis.title = "Funding Score (1-10)"
        chart.y_axis.title = "Revenue (EUR M)"
        chart.style = 13
        chart.width = 22
        chart.height = 16

        helper_col = 20
        for role_name in MNA_ROLE_ORDER:
            points = grouped.get(role_name, [])
            if not points:
                continue

            x_col = helper_col
            y_col = helper_col + 1

            ws.cell(row=1, column=x_col, value=f"{role_name} Funding")
            ws.cell(row=1, column=y_col, value=f"{role_name} Revenue")
            for i, (fs, rv) in enumerate(points, 2):
                ws.cell(row=i, column=x_col, value=fs)
                ws.cell(row=i, column=y_col, value=rv)

            data_end = len(points) + 1
            x_ref = Reference(ws, min_col=x_col, min_row=1, max_row=data_end)
            y_ref = Reference(ws, min_col=y_col, min_row=1, max_row=data_end)
            series = Series(y_ref, x_ref, title=role_name)
            series.graphicalProperties.line.noFill = True
            chart.series.append(series)

            ws.column_dimensions[get_column_letter(x_col)].hidden = True
            ws.column_dimensions[get_column_letter(y_col)].hidden = True
            helper_col += 2

        chart_anchor = f"{get_column_letter(len(headers) + 2)}2"
        ws.add_chart(chart, chart_anchor)


def write_raw_data(wb: Workbook, competitors: list[dict], link_base: Optional[Path] = None) -> None:
    """Flat raw data sheet for pivot tables and ad-hoc analysis."""
    ws = wb.create_sheet("Raw Data")

    headers = [
        "Company", "Folder", "Tier", "Data Availability",
        "Revenue Growth Score", "Funding Momentum Score", "Employee Growth Score",
        "Geographic Expansion Score", "M&A Activity Score", "SaaS Maturity Score",
        "Composite Score", "Classification",
        "Latest Revenue EUR M", "Revenue CAGR 3yr %", "Revenue CAGR 5yr %",
        "Recurring Revenue %",
        "Latest Headcount", "Employee CAGR %", "Open Positions",
        "Total Raised", "Latest Valuation",
        "Revenue Trend", "Employee Trend",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"

    max_years = _get_max_timeline_length(competitors)

    for row_idx, comp in enumerate(competitors, 2):
        eneve = is_eneve(comp)
        cls = get_classification(comp)
        rev = comp.get("revenue", {})
        emp = comp.get("employees", {})
        prof = comp.get("profitability", {})
        funding = comp.get("funding", {})
        rev_tl = [e.get("eur_millions") for e in rev.get("timeline", [])]
        emp_tl = [e.get("headcount") for e in emp.get("timeline", [])]
        values = [
            comp.get("company_name", ""),
            comp.get("folder", ""),
            comp.get("tier", ""),
            comp.get("data_availability", ""),
            get_score(comp, "Revenue Growth"),
            get_score(comp, "Funding Momentum"),
            get_score(comp, "Employee Growth"),
            get_score(comp, "Geographic Expansion"),
            get_score(comp, "M&A Activity"),
            get_score(comp, "SaaS Maturity"),
            get_composite(comp),
            cls or "",
            rev.get("latest_revenue_eur_m"),
            rev.get("cagr_3yr_pct"),
            rev.get("cagr_5yr_pct"),
            prof.get("recurring_revenue_pct"),
            emp.get("latest_headcount"),
            emp.get("employee_cagr_pct"),
            emp.get("open_positions"),
            funding.get("total_raised_text", ""),
            funding.get("latest_valuation_text", ""),
            text_sparkline(rev_tl),
            text_sparkline(emp_tl),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=format_value(val))
            style_data_cell(cell, is_eneve=eneve, classification=cls, row_idx=row_idx)
            apply_number_format(cell, headers[col - 1])
        add_company_link(ws.cell(row=row_idx, column=1), comp, link_base)

    if max_years > 0:
        _add_sparkline_columns(ws, competitors, len(headers), max_years)

    auto_fit_columns(ws)
    add_autofilter(ws, len(headers))
    setup_print_layout(ws, len(headers))


def write_data_explorer(
    wb: Workbook, competitors: list[dict], link_base: Optional[Path] = None
) -> None:
    """Data Explorer sheet with an Excel Table for interactive filtering.

    Creates a styled table sorted by Composite Score descending, with columns
    for Company, Tier, Classification, and the six key metrics.  The Excel
    Table format enables autofilter dropdowns on every column (immediate
    slicer-like filtering) and lets analysts create native pivot tables or
    slicers in two clicks via Insert > PivotTable / Insert > Slicer.

    openpyxl cannot create pivot tables or slicers programmatically -- this is
    a known library limitation.  The Excel Table provides equivalent interactive
    filtering and is the recommended foundation for analyst-created pivot views.
    """
    ws = wb.create_sheet("Data Explorer")

    # --- Instruction area (rows 1-3) ---
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "DATA EXPLORER - Interactive Competitor Analysis"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:I2")
    subtitle = ws["A2"]
    subtitle.value = (
        "Use the dropdown arrows on each column header to filter by any value.  "
        "To create a pivot table: select any cell in the table > Insert > PivotTable.  "
        "To add slicers: click the table > Table Design > Insert Slicer."
    )
    subtitle.font = Font(name="Calibri", size=10, italic=True, color="1B2A4A")
    subtitle.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 36

    ws.merge_cells("A3:I3")
    reset_note = ws["A3"]
    reset_note.value = (
        "Reset Filters: click the dropdown arrow on a filtered column > "
        "Clear Filter, or right-click any slicer > Clear Filter.  "
        "Keyboard: Alt+Down on a header cell to open the filter menu."
    )
    reset_note.font = Font(name="Calibri", size=10, color="808080")
    reset_note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 28

    # --- Data table starts at row 5 (row 4 is blank spacer) ---
    header_row = 5
    headers = [
        "Company",
        "Tier",
        "Classification",
        "Revenue (EUR M)",
        "Revenue CAGR 3yr %",
        "Composite Score",
        "AI Score",
        "SaaS Score",
        "Employees",
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col, value=h)
    style_header_row(ws, len(headers))
    # Re-apply header styling on the correct row (style_header_row targets row 1)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER

    # Sort competitors by composite score descending (default view)
    sorted_comps = sorted(
        competitors,
        key=lambda c: get_composite(c) or 0,
        reverse=True,
    )

    for row_idx, comp in enumerate(sorted_comps, header_row + 1):
        cls = get_classification(comp)
        eneve = is_eneve(comp)
        rev = comp.get("revenue", {})
        emp = comp.get("employees", {})
        values = [
            comp.get("company_name", ""),
            comp.get("tier", ""),
            cls or "",
            rev.get("latest_revenue_eur_m"),
            rev.get("cagr_3yr_pct"),
            get_composite(comp),
            get_ai_score(comp),
            get_score(comp, "SaaS Maturity"),
            emp.get("latest_headcount"),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=format_value(val))
            style_data_cell(
                cell, is_eneve=eneve, classification=cls, row_idx=row_idx
            )
            apply_number_format(cell, headers[col - 1])
        add_company_link(ws.cell(row=row_idx, column=1), comp, link_base)

    last_data_row = header_row + len(sorted_comps)

    # --- Create an Excel Table (ListObject) over the data range ---
    table_ref = f"A{header_row}:{get_column_letter(len(headers))}{last_data_row}"
    table = Table(displayName="CompetitorExplorer", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    # --- Conditional formatting on score columns ---
    score_cols = {
        "F": ("Composite Score", 1, 10),
        "G": ("AI Score", 0, 10),
        "H": ("SaaS Score", 1, 10),
    }
    for col_letter, (_label, lo, hi) in score_cols.items():
        cell_range = f"{col_letter}{header_row + 1}:{col_letter}{last_data_row}"
        ws.conditional_formatting.add(
            cell_range,
            ColorScaleRule(
                start_type="num", start_value=lo, start_color="F8696B",
                mid_type="num", mid_value=(lo + hi) / 2, mid_color="FFEB84",
                end_type="num", end_value=hi, end_color="63BE7B",
            ),
        )

    # --- Column widths ---
    col_widths = [30, 12, 16, 18, 20, 16, 12, 12, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = f"A{header_row + 1}"
    ws.sheet_view.showGridLines = False
    setup_print_layout(ws, len(headers))


def compute_market_stats(
    values: list[Optional[float]], companies: list[str]
) -> dict:
    """Compute average, median, best value and best company from competitor values."""
    pairs = [(v, c) for v, c in zip(values, companies) if v is not None]
    if not pairs:
        return {"avg": None, "median": None, "best": None, "best_company": None}
    clean = [p[0] for p in pairs]
    best_idx = clean.index(max(clean))
    return {
        "avg": statistics.mean(clean),
        "median": statistics.median(clean),
        "best": max(clean),
        "best_company": pairs[best_idx][1],
    }


def write_eneve_positioning(wb: Workbook, competitors: list[dict], link_base: Optional[Path] = None) -> None:
    """Eneve vs Market comparison sheet with conditional formatting and grouped bar chart."""
    ws = wb.create_sheet("Eneve vs Market")

    eneve = next((c for c in competitors if is_eneve(c)), None)
    non_eneve = [c for c in competitors if not is_eneve(c)]

    metrics = [
        ("Revenue CAGR 3yr (%)", lambda c: c.get("revenue", {}).get("cagr_3yr_pct")),
        ("Employee CAGR (%)", lambda c: c.get("employees", {}).get("employee_cagr_pct")),
        ("Composite Score", lambda c: get_composite(c)),
        ("SaaS Maturity Score", lambda c: get_score(c, "SaaS Maturity")),
        ("Recurring Revenue (%)", lambda c: c.get("profitability", {}).get("recurring_revenue_pct")),
        ("Latest Revenue (EUR M)", lambda c: c.get("revenue", {}).get("latest_revenue_eur_m")),
        ("Latest Headcount", lambda c: c.get("employees", {}).get("latest_headcount")),
    ]

    headers = ["Metric", "Eneve", "Market Average", "Market Median", "Best-in-Class", "Best Company"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"

    company_names = [c.get("company_name", "") for c in non_eneve]

    for row_idx, (metric_name, extractor) in enumerate(metrics, 2):
        eneve_val = extractor(eneve) if eneve else None
        market_vals = [extractor(c) for c in non_eneve]
        stats = compute_market_stats(market_vals, company_names)

        metric_cell = ws.cell(row=row_idx, column=1, value=metric_name)
        metric_cell.font = BOLD_FONT
        metric_cell.border = THIN_BORDER

        eneve_cell = ws.cell(
            row=row_idx, column=2,
            value=eneve_val if eneve_val is not None else "N/A",
        )
        eneve_cell.font = BOLD_FONT
        eneve_cell.border = THIN_BORDER

        if eneve_val is not None and stats["avg"] is not None:
            eneve_cell.fill = LEAD_FILL if eneve_val >= stats["avg"] else TRAIL_FILL

        for col, val in [(3, stats["avg"]), (4, stats["median"]), (5, stats["best"])]:
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

        company_cell = ws.cell(row=row_idx, column=6, value=stats["best_company"] or "")
        company_cell.font = DATA_FONT
        company_cell.border = THIN_BORDER

    auto_fit_columns(ws)
    setup_print_layout(ws, len(headers))

    last_data_row = len(metrics) + 1
    if last_data_row > 2:
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.title = "Eneve vs Market Average"
        chart.y_axis.title = "Value"
        chart.style = 10

        data_ref = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=last_data_row)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=last_data_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        chart.width = 25
        chart.height = 14

        ws.add_chart(chart, f"A{last_data_row + 2}")


METHODOLOGY_SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1B2A4A")
METHODOLOGY_TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="FFFFFF")


def write_methodology_sheet(wb: Workbook, data: dict) -> None:
    """Methodology and data quality notes — static text sheet for board credibility."""
    ws = wb.create_sheet("Methodology")

    sections = [
        ("DATA SOURCES", [
            "Financial data is sourced from a combination of:",
            "  - Public annual reports and financial filings",
            "  - Market research databases (Gartner, IDC, Forrester)",
            "  - Company press releases and investor presentations",
            "  - Industry analyst reports and news articles",
            "  - LinkedIn data for employee headcount trends",
            "  - Crunchbase/PitchBook for funding data",
        ]),
        ("CONFIDENCE LEVELS", [
            "Each data point is assessed for confidence:",
            "  - High: Direct from official source (annual report, regulatory filing)",
            "  - Medium: From reputable secondary source or cross-referenced",
            "  - Low: Single unverified source or analyst estimate",
            "  - Estimated: Calculated/interpolated from available data points",
        ]),
        ("SCORING METHODOLOGY", [
            "Each competitor is scored across 6 dimensions on a 1-10 scale:",
            "  1. Revenue Growth: YoY and CAGR trajectory",
            "  2. Funding Momentum: Capital raised, valuation trajectory",
            "  3. Employee Growth: Headcount CAGR, hiring signals",
            "  4. Geographic Expansion: International presence, new markets",
            "  5. M&A Activity: Acquisition cadence, strategic fit",
            "  6. SaaS Maturity: Cloud revenue %, recurring revenue %",
            "",
            "Composite Score = Average of all 6 dimension scores",
        ]),
        ("CLASSIFICATION THRESHOLDS", [
            "Based on composite score:",
            "  - Rocket: 7.0 - 10.0 (high-growth, aggressive competitor)",
            "  - Riser: 5.0 - 6.9 (growing steadily, gaining momentum)",
            "  - Steady: 3.0 - 4.9 (stable, moderate growth)",
            "  - Dinosaur: 1.0 - 2.9 (declining or stagnant)",
        ]),
        ("CURRENCY CONVERSION", [
            "All revenue figures converted to EUR using annual average exchange rates.",
            "Source currencies include USD, NOK, SEK, PLN, GBP.",
            "Conversion rates sourced from European Central Bank annual averages.",
        ]),
        ("DATA FRESHNESS", [
            f"Total competitors tracked: {data.get('metadata', {}).get('total_folders', 'N/A')}",
            f"Competitors with financial data: {data.get('metadata', {}).get('with_financial_data', 'N/A')}",
            f"Competitors without financial data: {data.get('metadata', {}).get('without_financial_data', 'N/A')}",
        ]),
        ("CAVEATS & LIMITATIONS", [
            "- Not all competitors disclose financial data publicly",
            "- Revenue figures for private companies are estimates unless noted",
            "- Employee counts from LinkedIn may include contractors",
            "- Funding data may be incomplete for bootstrapped companies",
            "- Geographic data reflects known market presence, not exhaustive coverage",
            "- Scores reflect relative positioning within this competitor set",
        ]),
    ]

    ws.column_dimensions["A"].width = 100
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells("A1:A1")
    title_cell = ws["A1"]
    title_cell.value = "SOLSTEIN - METHODOLOGY & DATA QUALITY"
    title_cell.font = METHODOLOGY_TITLE_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 36

    row = 3
    for section_title, lines in sections:
        ws.cell(row=row, column=1, value=section_title).font = METHODOLOGY_SECTION_FONT
        row += 1
        for line in lines:
            cell = ws.cell(row=row, column=1, value=line)
            cell.font = DATA_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        row += 1  # blank row between sections


THREAT_TIMELINE_YEARS = list(range(2024, 2030))
THREAT_TIMELINE_QUARTERS = ["Q1", "Q2", "Q3", "Q4"]

THREAT_TIMELINE_DATA: list[dict] = [
    {
        "competitor": "Dexter Energy",
        "threat_level": "High",
        "nl_status": "Active (Amsterdam HQ)",
        "active_start": (2024, 1),
        "events": {
            (2024, 2): "Computable Award; AI in Real Life feature",
            (2025, 1): "Energy Risk 'One to Watch'",
            (2025, 3): "Series C EUR 23M; 13 EU countries",
            (2025, 4): "BESS cross-market optimization",
            (2026, 1): "Solar Nowcasting product",
        },
    },
    {
        "competitor": "Kraken (Octopus Energy)",
        "threat_level": "High",
        "nl_status": "Rotterdam hub (Oct 2024)",
        "active_start": (2024, 4),
        "events": {
            (2024, 1): "$370M follow-on; $9B valuation",
            (2024, 3): "Jedlix acquisition (Rotterdam)",
            (2025, 2): "National Grid US 6.5M accounts",
            (2025, 3): "Kraken spin-off announced",
            (2025, 4): "$1B funding at $8.65B",
            (2026, 2): "Kraken demerger completion",
            (2027, 1): "Kraken IPO (potential $15B)",
        },
    },
    {
        "competitor": "Hansen Technologies",
        "threat_level": "High",
        "nl_status": "EDSN embedded (since 2001)",
        "active_start": (2024, 1),
        "events": {
            (2024, 1): "powercloud EUR 30M (DE market)",
            (2024, 2): "Vattenfall EDM go-live",
            (2025, 1): "Dial AI investment (AI CX)",
            (2025, 3): "powercloud rebranded Hansen DE",
            (2025, 4): "Digitalk GBP 33M acquisition",
            (2026, 1): "E-world 2026 exhibition",
        },
    },
    {
        "competitor": "Volue ASA",
        "threat_level": "Very High",
        "nl_status": "Adjacent (DACH/Nordics)",
        "active_start": (2024, 3),
        "events": {
            (2024, 3): "Take-private NOK 6.1B",
            (2024, 4): "PowerBot acquired; delisted",
            (2025, 1): "Reorg: 3 BUs (Energy focus)",
            (2025, 2): "Scanmatic divested (HW exit)",
            (2025, 3): "Infrastructure sold (focus)",
            (2025, 4): "smartPulse acquired (trading)",
            (2026, 1): "TA Associates EUR 1.5B; Optimeering acq.",
            (2027, 1): "Expected PE exit window opens",
            (2028, 1): "Potential strategic buyer/IPO",
        },
    },
    {
        "competitor": "tem energy",
        "threat_level": "Medium",
        "nl_status": "Not yet (UK focus)",
        "active_start": (2024, 4),
        "events": {
            (2024, 4): "RED platform launch (UK)",
            (2025, 2): "1,500 UK customers",
            (2026, 1): "Series B $75M; $300M+ valuation",
            (2026, 3): "Texas + Australia market entry",
        },
    },
    {
        "competitor": "Engrate AB",
        "threat_level": "Medium",
        "nl_status": "NL entry H1 2025 (TenneT)",
        "active_start": (2025, 1),
        "events": {
            (2024, 1): "Founded; Norrsken Accelerator",
            (2024, 3): "Pre-seed EUR 500K+",
            (2025, 1): "Schedule API: DE + NL (TenneT)",
            (2025, 2): "Seed EUR 2.5M (Maniv-led)",
            (2025, 4): "DSO Connect; MCP Server",
            (2026, 1): "4 new products in pipeline",
        },
    },
    {
        "competitor": "EG A/S (Utility Div.)",
        "threat_level": "Medium",
        "nl_status": "Nordic only (DK/SE/NO/FI)",
        "active_start": (2024, 1),
        "events": {
            (2023, 4): "EUR 400M growth equity; >EUR 1B val.",
            (2024, 1): "Mestro AB acquired (SE, AI energy)",
            (2024, 2): "Timma Oy acquired (FI entry)",
            (2024, 4): "India 850+ employees",
            (2025, 4): "Bright Energy AI acquisition",
        },
    },
    {
        "competitor": "EU Harmonization",
        "threat_level": "Structural",
        "nl_status": "Regulatory framework",
        "active_start": (2024, 1),
        "events": {
            (2022, 2): "PICASSO go-live (aFRR)",
            (2022, 4): "MARI go-live (mFRR)",
            (2025, 4): "MARI: TenneT NL joins",
            (2025, 4): "TERRE: operations cease",
            (2026, 1): "TERRE: decommissioned",
            (2026, 2): "Full cross-border balancing active",
        },
    },
]

THREAT_LEVEL_FILLS = {
    "Very High": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
    "High": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    "Medium": PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"),
    "Structural": PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid"),
}

THREAT_LEVEL_FONTS = {
    "Very High": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
    "High": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
    "Medium": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
    "Structural": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
}

# Quarterly proximity fills: the closer a quarter is to "now", the redder it gets
# Index 0 = far past, higher index = closer to now and beyond
_PROXIMITY_FILLS = [
    PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),  # far past: light blue-gray
    PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # past: light green
    PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # recent past: pale green
    PatternFill(start_color="FFEB84", end_color="FFEB84", fill_type="solid"),  # approaching: yellow
    PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),  # near: orange
    PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid"),  # imminent: dark orange
    PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # now: red
    PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid"),  # future: dark red
]


def _quarter_index(year: int, quarter: int) -> int:
    """Convert (year, quarter) to linear index starting from (2024, 1) = 0."""
    return (year - 2024) * 4 + (quarter - 1)


def _proximity_fill(year: int, quarter: int, now_year: int = 2026, now_quarter: int = 1) -> PatternFill:
    """Get fill color based on how close a quarter is to the current date."""
    now_idx = _quarter_index(now_year, now_quarter)
    cell_idx = _quarter_index(year, quarter)
    diff = cell_idx - now_idx  # negative = past, positive = future

    if diff <= -6:
        return _PROXIMITY_FILLS[0]
    elif diff <= -3:
        return _PROXIMITY_FILLS[1]
    elif diff <= -1:
        return _PROXIMITY_FILLS[2]
    elif diff == 0:
        return _PROXIMITY_FILLS[6]
    elif diff <= 2:
        return _PROXIMITY_FILLS[5]
    elif diff <= 4:
        return _PROXIMITY_FILLS[4]
    elif diff <= 8:
        return _PROXIMITY_FILLS[3]
    else:
        return _PROXIMITY_FILLS[2]


def write_threat_timeline_sheet(wb: Workbook, competitors: list[dict], link_base: Optional[Path] = None) -> None:
    """Threat Convergence Timeline: Gantt-style view of when competitors threaten the NL market.

    Rows: one per competitor/threat actor.
    Columns: Competitor | Threat Level | NL Status | then one column per quarter (2024-Q1 to 2029-Q4).
    Cells with events show brief text; active periods are shaded.
    Proximity-based conditional coloring: cells get progressively redder as threats approach current date.
    Summary row at bottom counts active threats per year.
    """
    ws = wb.create_sheet("Threat Timeline")

    info_cols = 3  # Competitor, Threat Level, NL Status
    num_quarters = len(THREAT_TIMELINE_YEARS) * 4  # 24 quarters

    # --- Row 1: Year headers (merged across 4 quarter columns each) ---
    ws.cell(row=1, column=1, value="Competitor").font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=1).border = THIN_BORDER
    ws.cell(row=1, column=2, value="Threat Level").font = HEADER_FONT
    ws.cell(row=1, column=2).fill = HEADER_FILL
    ws.cell(row=1, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=2).border = THIN_BORDER
    ws.cell(row=1, column=3, value="NL Status").font = HEADER_FONT
    ws.cell(row=1, column=3).fill = HEADER_FILL
    ws.cell(row=1, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=3).border = THIN_BORDER

    for yi, year in enumerate(THREAT_TIMELINE_YEARS):
        start_col = info_cols + 1 + yi * 4
        end_col = start_col + 3
        ws.merge_cells(
            start_row=1, start_column=start_col,
            end_row=1, end_column=end_col,
        )
        year_cell = ws.cell(row=1, column=start_col)
        year_cell.value = str(year)
        year_cell.font = HEADER_FONT
        year_cell.fill = HEADER_FILL
        year_cell.alignment = Alignment(horizontal="center", vertical="center")
        year_cell.border = THIN_BORDER
        for qc in range(start_col + 1, end_col + 1):
            ws.cell(row=1, column=qc).border = THIN_BORDER

    # --- Row 2: Quarter sub-headers ---
    ws.cell(row=2, column=1).fill = HEADER_FILL
    ws.cell(row=2, column=1).border = THIN_BORDER
    ws.cell(row=2, column=2).fill = HEADER_FILL
    ws.cell(row=2, column=2).border = THIN_BORDER
    ws.cell(row=2, column=3).fill = HEADER_FILL
    ws.cell(row=2, column=3).border = THIN_BORDER

    for yi, year in enumerate(THREAT_TIMELINE_YEARS):
        for qi, qtr in enumerate(THREAT_TIMELINE_QUARTERS):
            col = info_cols + 1 + yi * 4 + qi
            cell = ws.cell(row=2, column=col, value=qtr)
            cell.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

    # --- Data rows: one per threat actor ---
    event_font = Font(name="Calibri", size=8)
    event_font_white = Font(name="Calibri", size=8, color="FFFFFF")
    active_fill_light = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")

    for row_offset, threat in enumerate(THREAT_TIMELINE_DATA):
        data_row = 3 + row_offset
        threat_level = threat["threat_level"]
        active_start_idx = _quarter_index(*threat["active_start"])

        # Competitor name
        name_cell = ws.cell(row=data_row, column=1, value=threat["competitor"])
        name_cell.font = BOLD_FONT
        name_cell.border = THIN_BORDER
        name_cell.alignment = Alignment(vertical="center")

        # Threat level with color badge
        level_cell = ws.cell(row=data_row, column=2, value=threat_level)
        level_cell.fill = THREAT_LEVEL_FILLS.get(threat_level, ALT_ROW_FILL)
        level_cell.font = THREAT_LEVEL_FONTS.get(threat_level, DATA_FONT)
        level_cell.border = THIN_BORDER
        level_cell.alignment = Alignment(horizontal="center", vertical="center")

        # NL Status
        status_cell = ws.cell(row=data_row, column=3, value=threat["nl_status"])
        status_cell.font = DATA_FONT
        status_cell.border = THIN_BORDER
        status_cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Quarter cells
        for yi, year in enumerate(THREAT_TIMELINE_YEARS):
            for qi in range(4):
                quarter = qi + 1
                col = info_cols + 1 + yi * 4 + qi
                cell_idx = _quarter_index(year, quarter)
                cell = ws.cell(row=data_row, column=col)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                event_text = threat["events"].get((year, quarter), "")

                if event_text:
                    cell.value = event_text
                    prox_fill = _proximity_fill(year, quarter)
                    cell.fill = prox_fill
                    # Use white font on darker fills
                    dark_colors = {"FF0000", "CC0000", "FF6600", "C00000", "7030A0"}
                    if prox_fill.start_color and prox_fill.start_color.rgb and str(prox_fill.start_color.rgb).replace("00", "", 1) in dark_colors:
                        cell.font = event_font_white
                    else:
                        cell.font = event_font
                elif cell_idx >= active_start_idx:
                    # Active period with no specific event: light shading
                    cell.fill = active_fill_light
                    cell.font = event_font

    # --- Summary row: count active threats per year ---
    summary_row = 3 + len(THREAT_TIMELINE_DATA) + 1  # blank row then summary
    ws.cell(row=summary_row, column=1, value="Active Threats in NL Market")
    ws.cell(row=summary_row, column=1).font = Font(name="Calibri", size=11, bold=True, color="1B2A4A")
    ws.cell(row=summary_row, column=1).border = THIN_BORDER

    for yi, year in enumerate(THREAT_TIMELINE_YEARS):
        count = 0
        for threat in THREAT_TIMELINE_DATA:
            active_start_idx = _quarter_index(*threat["active_start"])
            year_mid_idx = _quarter_index(year, 2)
            if year_mid_idx >= active_start_idx:
                count += 1

        # Merge 4 quarter columns for yearly count
        start_col = info_cols + 1 + yi * 4
        end_col = start_col + 3
        ws.merge_cells(
            start_row=summary_row, start_column=start_col,
            end_row=summary_row, end_column=end_col,
        )
        count_cell = ws.cell(row=summary_row, column=start_col)
        count_cell.value = count
        count_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        count_cell.alignment = Alignment(horizontal="center", vertical="center")
        count_cell.border = THIN_BORDER

        # Color-code by count (more threats = redder)
        if count <= 3:
            count_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        elif count <= 5:
            count_cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
        else:
            count_cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")

    # --- Column widths ---
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 24
    for qi in range(num_quarters):
        col_letter = get_column_letter(info_cols + 1 + qi)
        ws.column_dimensions[col_letter].width = 18

    # --- Row heights for data rows ---
    for r in range(3, 3 + len(THREAT_TIMELINE_DATA)):
        ws.row_dimensions[r].height = 38

    ws.freeze_panes = "D3"
    ws.sheet_view.showGridLines = False

    total_cols = info_cols + num_quarters
    setup_print_layout(ws, total_cols)


# ---------------------------------------------------------------------------
# Data Confidence helpers
# ---------------------------------------------------------------------------

_CONFIDENCE_RE = re.compile(
    r"\|\s*(Confirmed(?:\s*\([^)]*\))?|Estimated(?:\s*\([^)]*\))?|Unknown(?:\s*\([^)]*\))?)\s*\|",
    re.IGNORECASE,
)

HIGH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MEDIUM_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
LOW_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

HIGH_FONT = Font(name="Calibri", size=10, bold=True, color="006100")
MEDIUM_FONT = Font(name="Calibri", size=10, bold=True, color="9C6500")
LOW_FONT = Font(name="Calibri", size=10, bold=True, color="9C0006")


def _classify_confidence(tag: str) -> str:
    """Map a raw confidence string to one of three canonical buckets."""
    lower = tag.strip().lower()
    if lower.startswith("confirmed"):
        return "Confirmed"
    if lower.startswith("estimated"):
        return "Estimated"
    return "Unknown"


def _count_confidence_from_markdown(md_path: Path) -> tuple[int, int, int]:
    """Parse a financial-growth markdown file and count confidence tags."""
    try:
        text = md_path.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return (0, 0, 0)
    confirmed = estimated = unknown = 0
    for match in _CONFIDENCE_RE.finditer(text):
        bucket = _classify_confidence(match.group(1))
        if bucket == "Confirmed":
            confirmed += 1
        elif bucket == "Estimated":
            estimated += 1
        else:
            unknown += 1
    return (confirmed, estimated, unknown)


def _count_confidence_from_json(comp: dict) -> tuple[int, int, int]:
    """Count confidence tags from the already-extracted revenue timeline."""
    confirmed = estimated = unknown = 0
    for entry in comp.get("revenue", {}).get("timeline", []):
        tag = entry.get("confidence", "")
        if not tag:
            continue
        bucket = _classify_confidence(tag)
        if bucket == "Confirmed":
            confirmed += 1
        elif bucket == "Estimated":
            estimated += 1
        else:
            unknown += 1
    return (confirmed, estimated, unknown)


def count_confidence_tags(
    comp: dict, link_base: Optional[Path] = None
) -> tuple[int, int, int]:
    """Return (confirmed, estimated, unknown) counts for a competitor.

    Prefers parsing the full markdown file (covers all table sections) when
    *link_base* is available; falls back to the revenue-timeline data in the
    JSON when it is not.
    """
    if link_base is not None:
        folder = comp.get("folder", "")
        if folder:
            md_path = link_base / folder / "financial-growth.md"
            counts = _count_confidence_from_markdown(md_path)
            if sum(counts) > 0:
                return counts
    return _count_confidence_from_json(comp)


def calc_confidence_ratio(
    confirmed: int, estimated: int, unknown: int
) -> Optional[float]:
    """Confidence % = Confirmed / Total.  Returns None when total is zero."""
    total = confirmed + estimated + unknown
    if total == 0:
        return None
    return round(confirmed / total * 100, 1)


def get_quality_rating(ratio: Optional[float]) -> str:
    """Map a confidence percentage to High / Medium / Low."""
    if ratio is None:
        return "N/A"
    if ratio > 70:
        return "High"
    if ratio >= 40:
        return "Medium"
    return "Low"


def write_confidence_sheet(
    wb: Workbook,
    competitors: list[dict],
    link_base: Optional[Path] = None,
) -> None:
    """Data Confidence sheet with per-competitor research quality metrics."""
    ws = wb.create_sheet("Data Confidence")

    headers = [
        "Company", "Tier", "Confirmed", "Estimated", "Unknown",
        "Confidence %", "Data Quality Rating",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"

    total_confirmed = total_estimated = total_unknown = 0

    sorted_comps = sorted(
        competitors,
        key=lambda c: calc_confidence_ratio(
            *count_confidence_tags(c, link_base)
        ) or 0,
        reverse=True,
    )

    for row_idx, comp in enumerate(sorted_comps, 2):
        conf, est, unk = count_confidence_tags(comp, link_base)
        total_confirmed += conf
        total_estimated += est
        total_unknown += unk

        ratio = calc_confidence_ratio(conf, est, unk)
        rating = get_quality_rating(ratio)

        cls = get_classification(comp)
        eneve = is_eneve(comp)

        values = [
            comp.get("company_name", ""),
            comp.get("tier", ""),
            conf,
            est,
            unk,
            ratio if ratio is not None else "--",
            rating,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=format_value(val))
            style_data_cell(cell, is_eneve=eneve, classification=cls, row_idx=row_idx)

        if isinstance(values[5], (int, float)):
            ws.cell(row=row_idx, column=6).number_format = '0.0"%"'

        add_company_link(ws.cell(row=row_idx, column=1), comp, link_base)

        rating_cell = ws.cell(row=row_idx, column=7)
        if rating == "High":
            rating_cell.fill = HIGH_FILL
            rating_cell.font = HIGH_FONT
        elif rating == "Medium":
            rating_cell.fill = MEDIUM_FILL
            rating_cell.font = MEDIUM_FONT
        elif rating == "Low":
            rating_cell.fill = LOW_FILL
            rating_cell.font = LOW_FONT

    # --- Summary row ---
    summary_row = len(sorted_comps) + 2
    portfolio_ratio = calc_confidence_ratio(
        total_confirmed, total_estimated, total_unknown
    )
    portfolio_rating = get_quality_rating(portfolio_ratio)

    summary_values = [
        "PORTFOLIO TOTAL", "", total_confirmed, total_estimated, total_unknown,
        portfolio_ratio if portfolio_ratio is not None else "--",
        portfolio_rating,
    ]
    for col, val in enumerate(summary_values, 1):
        cell = ws.cell(row=summary_row, column=col, value=format_value(val))
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
        cell.fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")

    if isinstance(summary_values[5], (int, float)):
        ws.cell(row=summary_row, column=6).number_format = '0.0"%"'

    # --- Conditional formatting: color scale on Confidence % column ---
    conf_col = "F"
    data_range = f"{conf_col}2:{conf_col}{len(sorted_comps) + 1}"
    ws.conditional_formatting.add(
        data_range,
        ColorScaleRule(
            start_type="num", start_value=0, start_color="FFC7CE",
            mid_type="num", mid_value=50, mid_color="FFEB9C",
            end_type="num", end_value=100, end_color="C6EFCE",
        ),
    )

    # --- Stacked bar chart ---
    if len(sorted_comps) >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.title = "Data Confidence Breakdown by Competitor"
        chart.y_axis.title = "Data Points"
        chart.x_axis.title = "Company"
        chart.style = 10

        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(sorted_comps) + 1)
        for col_idx, label, color in [
            (3, "Confirmed", "4472C4"),
            (4, "Estimated", "ED7D31"),
            (5, "Unknown", "A5A5A5"),
        ]:
            data_ref = Reference(ws, min_col=col_idx, min_row=1, max_row=len(sorted_comps) + 1)
            series = Series(data_ref, title=label)
            series.graphicalProperties.solidFill = color
            chart.append(series)
        chart.set_categories(cats_ref)
        chart.shape = 4
        chart.width = 30
        chart.height = 16

        ws.add_chart(chart, f"A{summary_row + 2}")

    auto_fit_columns(ws)
    add_autofilter(ws, len(headers))
    setup_print_layout(ws, len(headers))


# ---------------------------------------------------------------------------
# Portfolio Risk helpers
# ---------------------------------------------------------------------------

PROXIMITY_SCORES = {"Immediate": 4, "Near-term": 3, "Long-term": 2, "Negligible": 1}
SEVERITY_SCORES = {"Critical": 4, "High": 3, "Medium": 2, "Low": 1}

RISK_FILLS = {
    "Critical Risk": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "High Risk": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Moderate Risk": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "Low Risk": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}


def _classify_threat_proximity(comp: dict) -> str:
    """Classify how soon a competitor could impact Eneve's market position.

    Uses classification tier + geographic expansion score as proxies for
    market overlap likelihood and expansion trajectory.
    """
    if is_eneve(comp):
        return "Negligible"
    cls = get_classification(comp)
    geo = get_score(comp, "Geographic Expansion") or 0

    if cls == "Rocket" and geo >= 6:
        return "Immediate"
    if cls == "Rocket":
        return "Near-term"
    if cls == "Riser" and geo >= 5:
        return "Near-term"
    if cls in ("Riser", "Steady") and (get_composite(comp) or 0) >= 5:
        return "Long-term"
    return "Negligible"


def _classify_threat_severity(comp: dict) -> str:
    """Classify the damage potential if a competitor enters direct competition.

    Considers revenue scale, funding capacity, and classification tier to
    determine how disruptive a head-to-head encounter would be.
    """
    if is_eneve(comp):
        return "Low"
    cls = get_classification(comp)
    rev_m = comp.get("revenue", {}).get("latest_revenue_eur_m") or 0
    raised = parse_total_raised_eur_m(comp.get("funding", {}).get("total_raised_text")) or 0

    if cls == "Rocket" and (rev_m >= 100 or raised >= 100):
        return "Critical"
    if cls == "Rocket":
        return "High"
    if cls == "Riser" and rev_m >= 50:
        return "High"
    if cls == "Riser":
        return "Medium"
    if (get_composite(comp) or 0) >= 4:
        return "Medium"
    return "Low"


def _risk_category(proximity: str, severity: str) -> str:
    """Combine proximity and severity into a single risk label."""
    score = PROXIMITY_SCORES.get(proximity, 1) * SEVERITY_SCORES.get(severity, 1)
    if score >= 12:
        return "Critical Risk"
    if score >= 6:
        return "High Risk"
    if score >= 3:
        return "Moderate Risk"
    return "Low Risk"


def _primary_risk_factor(comp: dict) -> str:
    """One-line summary of the main competitive risk posed by this company."""
    cls = get_classification(comp)
    rev_m = comp.get("revenue", {}).get("latest_revenue_eur_m")
    raised = parse_total_raised_eur_m(comp.get("funding", {}).get("total_raised_text"))
    ai_prod = get_ai_in_production(comp)
    saas = get_score(comp, "SaaS Maturity") or 0

    parts: list[str] = []
    if cls == "Rocket":
        parts.append("High-growth competitor")
    elif cls == "Riser":
        parts.append("Rising competitor")
    if rev_m and rev_m >= 30:
        parts.append(f"EUR {rev_m:.0f}M revenue")
    if raised and raised >= 50:
        parts.append(f"EUR {raised:.0f}M+ funded")
    if ai_prod:
        parts.append("AI in production")
    if saas >= 8:
        parts.append("Strong SaaS platform")
    return "; ".join(parts[:3]) if parts else "Limited competitive signals"


def _risk_mitigation(comp: dict) -> str:
    """Brief mitigation strategy tailored to the competitor profile."""
    cls = get_classification(comp)
    saas = get_score(comp, "SaaS Maturity") or 0
    ai_prod = get_ai_in_production(comp)

    if cls == "Rocket" and ai_prod:
        return "Accelerate AI roadmap; deepen domain differentiation"
    if cls == "Rocket":
        return "Monitor closely; strengthen customer lock-in"
    if cls == "Riser":
        return "Track growth trajectory; defend key accounts"
    if cls == "Steady" and saas >= 6:
        return "Watch for SaaS acceleration pivot"
    return "Periodic review; low priority"


def _bubble_size(comp: dict) -> float:
    """Revenue or total funding in EUR M for bubble chart sizing.

    Falls back to a minimum size of 1.0 so every competitor is visible.
    """
    rev_m = comp.get("revenue", {}).get("latest_revenue_eur_m")
    if rev_m and rev_m > 0:
        return rev_m
    raised = parse_total_raised_eur_m(comp.get("funding", {}).get("total_raised_text"))
    if raised and raised > 0:
        return raised
    return 1.0


def write_portfolio_risk_sheet(
    wb: Workbook,
    competitors: list[dict],
    link_base: Optional[Path] = None,
) -> None:
    """Portfolio Risk sheet with KPIs, risk matrix, bubble chart, and top-5 summary.

    Aggregates competitive threats at portfolio level to present the view PE
    firms use: across the entire competitive landscape rather than per-company.
    """
    ws = wb.create_sheet("Portfolio Risk")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    eneve_comp = next((c for c in competitors if is_eneve(c)), None)
    non_eneve = [c for c in competitors if not is_eneve(c)]

    # --- KPI computation ---
    rockets = [c for c in non_eneve if get_classification(c) == "Rocket"]
    rocket_count = len(rockets)

    rocket_capital = sum(
        parse_total_raised_eur_m(c.get("funding", {}).get("total_raised_text")) or 0
        for c in rockets
    )

    ai_in_prod_count = sum(1 for c in non_eneve if get_ai_in_production(c))

    # NL proxy: geographic expansion >= 6 in this European energy market
    nl_proxy_count = sum(
        1 for c in non_eneve if (get_score(c, "Geographic Expansion") or 0) >= 6
    )

    saas_scores = [
        get_score(c, "SaaS Maturity")
        for c in non_eneve
        if get_score(c, "SaaS Maturity") is not None
    ]
    avg_saas = statistics.mean(saas_scores) if saas_scores else None
    eneve_saas = get_score(eneve_comp, "SaaS Maturity") if eneve_comp else None

    # --- Layout: column widths ---
    col_widths = {
        "A": 6, "B": 24, "C": 8, "D": 18, "E": 18,
        "F": 16, "G": 34, "H": 38,
    }
    for letter, w in col_widths.items():
        ws.column_dimensions[letter].width = w

    # --- Row 1-2: Title ---
    ws.merge_cells("A1:H2")
    title_cell = ws["A1"]
    title_cell.value = "PORTFOLIO RISK DASHBOARD"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24

    # --- Row 4-6: KPI tiles (5 tiles across column pairs) ---
    kpi_label_font = Font(name="Calibri", size=8, bold=True, color="666666")
    kpi_value_font = Font(name="Calibri", size=18, bold=True, color="1B2A4A")
    kpi_sub_font = Font(name="Calibri", size=8, italic=True, color="999999")

    tiles = [
        ("A", "B", "ROCKETS IN MARKET", str(rocket_count),
         f"of {len(non_eneve)} competitors"),
        ("C", "D", "ROCKET CAPITAL (EUR M)",
         f"{rocket_capital:.0f}" if rocket_capital else "N/A",
         "Total raised by Rockets"),
        ("E", "F", "AI IN PRODUCTION", str(ai_in_prod_count),
         f"of {len(non_eneve)} competitors"),
        ("G", "H", "OPERATING IN NL (est.)", str(nl_proxy_count),
         "Geo expansion >= 6"),
    ]

    for c1, c2, label, value, subtitle in tiles:
        ws.merge_cells(f"{c1}4:{c2}4")
        lbl_cell = ws[f"{c1}4"]
        lbl_cell.value = label
        lbl_cell.font = kpi_label_font
        lbl_cell.alignment = center

        ws.merge_cells(f"{c1}5:{c2}5")
        val_cell = ws[f"{c1}5"]
        val_cell.value = value
        val_cell.font = kpi_value_font
        val_cell.alignment = center

        ws.merge_cells(f"{c1}6:{c2}6")
        sub_cell = ws[f"{c1}6"]
        sub_cell.value = subtitle
        sub_cell.font = kpi_sub_font
        sub_cell.alignment = center

    ws.row_dimensions[5].height = 30

    # 5th KPI below (SaaS Maturity comparison) using a merged row
    ws.merge_cells("A7:B7")
    ws["A7"].value = "AVG SaaS MATURITY"
    ws["A7"].font = kpi_label_font
    ws["A7"].alignment = center
    ws.merge_cells("C7:D7")
    saas_text = f"{avg_saas:.1f}" if avg_saas is not None else "N/A"
    eneve_saas_text = f"{eneve_saas:.1f}" if eneve_saas is not None else "N/A"
    ws["C7"].value = f"Market: {saas_text}  |  Eneve: {eneve_saas_text}"
    ws["C7"].font = Font(name="Calibri", size=10, bold=True, color="1B2A4A")
    ws["C7"].alignment = Alignment(horizontal="left", vertical="center")

    # --- Row 9: Section header ---
    ws.merge_cells("A9:H9")
    sec_cell = ws["A9"]
    sec_cell.value = "COMPETITIVE RISK MATRIX"
    sec_cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    sec_cell.fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    sec_cell.alignment = center

    # --- Row 10: Table headers ---
    headers = [
        "Rank", "Company", "Tier", "Threat Proximity",
        "Threat Severity", "Risk Category", "Primary Risk Factor", "Mitigation",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=10, column=col, value=h)
    style_header_row(ws, len(headers))
    ws.freeze_panes = "A11"

    # --- Classify and sort competitors by combined risk score ---
    risk_data: list[dict] = []
    for comp in competitors:
        prox = _classify_threat_proximity(comp)
        sev = _classify_threat_severity(comp)
        cat = _risk_category(prox, sev)
        combined = PROXIMITY_SCORES.get(prox, 1) * SEVERITY_SCORES.get(sev, 1)
        risk_data.append({
            "comp": comp,
            "proximity": prox,
            "severity": sev,
            "category": cat,
            "combined_score": combined,
            "risk_factor": _primary_risk_factor(comp),
            "mitigation": _risk_mitigation(comp),
            "bubble": _bubble_size(comp),
        })

    risk_data.sort(key=lambda r: r["combined_score"], reverse=True)

    # --- Chart data columns (I=proximity numeric, J=severity numeric, K=bubble size) ---
    chart_col_start = len(headers) + 2  # Column J (skip one spacer)
    ws.cell(row=10, column=chart_col_start, value="ProxNum")
    ws.cell(row=10, column=chart_col_start + 1, value="SevNum")
    ws.cell(row=10, column=chart_col_start + 2, value="BubbleSize")
    for ci in range(3):
        c = ws.cell(row=10, column=chart_col_start + ci)
        c.font = Font(name="Calibri", size=8, color="999999")

    # --- Data rows ---
    for row_idx, rd in enumerate(risk_data, 11):
        rank = row_idx - 10
        comp = rd["comp"]
        eneve = is_eneve(comp)
        cls = get_classification(comp)

        values = [
            rank,
            comp.get("company_name", ""),
            comp.get("tier", ""),
            rd["proximity"],
            rd["severity"],
            rd["category"],
            rd["risk_factor"],
            rd["mitigation"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=format_value(val))
            style_data_cell(cell, is_eneve=eneve, classification=cls, row_idx=row_idx)

        # Colour-code the Risk Category cell
        cat_cell = ws.cell(row=row_idx, column=6)
        if rd["category"] in RISK_FILLS:
            cat_cell.fill = RISK_FILLS[rd["category"]]

        add_company_link(ws.cell(row=row_idx, column=2), comp, link_base)

        # Chart data
        ws.cell(row=row_idx, column=chart_col_start,
                value=PROXIMITY_SCORES.get(rd["proximity"], 1))
        ws.cell(row=row_idx, column=chart_col_start + 1,
                value=SEVERITY_SCORES.get(rd["severity"], 1))
        ws.cell(row=row_idx, column=chart_col_start + 2,
                value=rd["bubble"])

    last_data_row = 10 + len(risk_data)

    # --- Bubble chart: Proximity (X) vs Severity (Y), bubble = revenue/funding ---
    if len(risk_data) >= 2:
        chart = BubbleChart()
        chart.title = "Risk Landscape: Proximity vs Severity"
        chart.x_axis.title = "Threat Proximity (4=Immediate)"
        chart.y_axis.title = "Threat Severity (4=Critical)"
        chart.x_axis.scaling.min = 0
        chart.x_axis.scaling.max = 5
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 5
        chart.style = 18
        chart.width = 22
        chart.height = 15

        xvalues = Reference(ws, min_col=chart_col_start, min_row=11, max_row=last_data_row)
        yvalues = Reference(ws, min_col=chart_col_start + 1, min_row=11, max_row=last_data_row)
        bubbles = Reference(ws, min_col=chart_col_start + 2, min_row=11, max_row=last_data_row)
        series = Series(yvalues, xvalues, bubbles, title="Competitors")
        chart.series.append(series)

        chart_anchor = f"J{last_data_row + 3}"
        ws.add_chart(chart, chart_anchor)

    # --- Top 5 Risks Summary ---
    top5_start = last_data_row + 2
    ws.merge_cells(f"A{top5_start}:H{top5_start}")
    sec2 = ws[f"A{top5_start}"]
    sec2.value = "TOP 5 PORTFOLIO RISKS"
    sec2.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    sec2.fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    sec2.alignment = center

    top5_headers = ["#", "Company", "Risk Category", "Proximity", "Severity", "Risk Summary"]
    hdr_row = top5_start + 1
    for col, h in enumerate(top5_headers, 1):
        cell = ws.cell(row=hdr_row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    top5 = [rd for rd in risk_data if not is_eneve(rd["comp"])][:5]
    for i, rd in enumerate(top5, 1):
        r = hdr_row + i
        comp = rd["comp"]
        ws.cell(row=r, column=1, value=i).font = BOLD_FONT
        name_cell = ws.cell(row=r, column=2, value=comp.get("company_name", ""))
        name_cell.font = BOLD_FONT
        add_company_link(name_cell, comp, link_base)

        cat_cell = ws.cell(row=r, column=3, value=rd["category"])
        if rd["category"] in RISK_FILLS:
            cat_cell.fill = RISK_FILLS[rd["category"]]
        cat_cell.font = DATA_FONT

        ws.cell(row=r, column=4, value=rd["proximity"]).font = DATA_FONT
        ws.cell(row=r, column=5, value=rd["severity"]).font = DATA_FONT
        ws.cell(row=r, column=6, value=rd["risk_factor"]).font = DATA_FONT

    add_autofilter(ws, len(headers))
    setup_print_layout(ws, len(headers))


def write_ai_talent_map_sheet(
    wb: Workbook,
    competitors: list[dict],
    link_base: Optional[Path] = None,
) -> None:
    """AI Talent Map sheet -- CONFIDENTIAL strategic intelligence.

    Maps AI team size, leadership depth, talent concentration risk,
    and acqui-hire attractiveness per competitor. Includes a bubble chart
    plotting AI Team Size vs AI Talent Density with acqui-hire score as
    bubble size.
    """
    ws = wb.create_sheet("AI Talent Map")

    # Sensitivity header spanning full width
    ws.merge_cells("A1:L1")
    banner = ws.cell(row=1, column=1, value="CONFIDENTIAL -- STRATEGIC INTELLIGENCE -- CTO/BOARD ONLY")
    banner.fill = CONFIDENTIAL_FILL
    banner.font = CONFIDENTIAL_FONT
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = [
        "Rank", "Company", "Tier", "Classification",
        "AI Team Size", "AI Density %", "Leadership Count",
        "Concentration Risk", "Acqui-Hire Score",
        "Key Hires (24mo)", "Talent Flow", "Growth Classification",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A3"

    # Sort: Acqui-Hire Score descending, then Concentration Risk descending
    sorted_comps = sorted(
        competitors,
        key=lambda c: (
            get_acquihire_score(c) or 0,
            get_concentration_risk(c) or 0,
        ),
        reverse=True,
    )

    for row_idx, comp in enumerate(sorted_comps, 3):
        rank = row_idx - 2
        cls = get_classification(comp)
        eneve = is_eneve(comp)
        conc_risk = get_concentration_risk(comp)
        acq_score = get_acquihire_score(comp)

        values = [
            rank,
            comp.get("company_name", ""),
            comp.get("tier", ""),
            cls or "",
            get_ai_talent_team_size(comp),
            get_ai_talent_pct_total(comp),
            get_ai_leadership_count(comp),
            conc_risk,
            acq_score,
            get_key_hires_count(comp),
            get_talent_flow(comp) or "Unknown",
            cls or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=format_value(val))
            cell.font = BOLD_FONT if eneve else DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            if eneve:
                cell.fill = ENEVE_FILL
            elif acq_score is not None and acq_score >= 7 and col == 9:
                cell.fill = ACQUIHIRE_TARGET_FILL
            elif conc_risk is not None and conc_risk >= 7 and col == 8:
                cell.fill = HIGH_RISK_FILL
            elif row_idx % 2 == 1:
                cell.fill = ALT_ROW_FILL

            apply_number_format(cell, headers[col - 1])

        add_company_link(ws.cell(row=row_idx, column=2), comp, link_base)

    last_row = len(sorted_comps) + 2

    # Heatmaps on Concentration Risk (H) and Acqui-Hire Score (I)
    if last_row > 3:
        ws.conditional_formatting.add(
            f"H3:H{last_row}",
            ColorScaleRule(
                start_type="num", start_value=1, start_color="63BE7B",
                mid_type="num", mid_value=5, mid_color="FFEB84",
                end_type="num", end_value=10, end_color="F8696B",
            ),
        )
        ws.conditional_formatting.add(
            f"I3:I{last_row}",
            ColorScaleRule(
                start_type="num", start_value=1, start_color="F8696B",
                mid_type="num", mid_value=5, mid_color="FFEB84",
                end_type="num", end_value=10, end_color="63BE7B",
            ),
        )

    # Bubble chart: AI Team Size (x) vs AI Density % (y), bubble = Acqui-Hire Score
    chart_data: list[tuple[float, float, float, str]] = []
    for comp in sorted_comps:
        ts = get_ai_talent_team_size(comp)
        density = get_ai_talent_pct_total(comp)
        acq = get_acquihire_score(comp)
        if ts is not None and density is not None and acq is not None:
            chart_data.append((ts, density, acq, comp.get("company_name", "")))

    if len(chart_data) >= 2:
        chart = BubbleChart()
        chart.title = "AI Talent Landscape: Team Size vs Density (Bubble = Acqui-Hire Score)"
        chart.x_axis.title = "AI Team Size (headcount)"
        chart.y_axis.title = "AI Talent Density (% of total)"
        chart.style = 13
        chart.width = 26
        chart.height = 18

        helper_col = len(headers) + 3
        ws.cell(row=2, column=helper_col, value="Team Size")
        ws.cell(row=2, column=helper_col + 1, value="Density %")
        ws.cell(row=2, column=helper_col + 2, value="Acqui-Hire")

        for i, (ts, density, acq, name) in enumerate(chart_data, 3):
            ws.cell(row=i, column=helper_col, value=ts)
            ws.cell(row=i, column=helper_col + 1, value=density)
            ws.cell(row=i, column=helper_col + 2, value=acq)

        data_end = len(chart_data) + 2
        x_ref = Reference(ws, min_col=helper_col, min_row=2, max_row=data_end)
        y_ref = Reference(ws, min_col=helper_col + 1, min_row=2, max_row=data_end)
        bubble_ref = Reference(ws, min_col=helper_col + 2, min_row=2, max_row=data_end)

        series = Series(y_ref, x_ref, bubble_ref, title="Competitors")
        chart.series.append(series)

        ws.column_dimensions[get_column_letter(helper_col)].hidden = True
        ws.column_dimensions[get_column_letter(helper_col + 1)].hidden = True
        ws.column_dimensions[get_column_letter(helper_col + 2)].hidden = True

        chart_anchor = f"{get_column_letter(len(headers) + 2)}3"
        ws.add_chart(chart, chart_anchor)

    auto_fit_columns(ws)
    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A2:{last_col_letter}{ws.max_row}"
    setup_print_layout(ws, len(headers))


def generate_workbook(data: dict, output_path: Path, *, profile: bool = False) -> None:
    """Build the complete Excel workbook from extracted competitor data."""
    competitors = data.get("competitors", [])
    if not competitors:
        log.error("No competitor data to generate report from.")
        return

    # Compute relative link paths from output location to source files
    source_dir_str = data.get("metadata", {}).get("source_directory")
    link_base = None
    if source_dir_str:
        source_dir = Path(source_dir_str)
        try:
            link_base = Path(os.path.relpath(source_dir, output_path.parent))
        except ValueError:
            link_base = source_dir

    wb = Workbook()

    log.info("Generating workbook with %d competitors, 19 sheets...", len(competitors))

    sheets = [
        ("Summary", write_summary_sheet, [wb, competitors, link_base]),
        ("Revenue Leaderboard", write_revenue_leaderboard, [wb, competitors, link_base]),
        ("Funding Leaderboard", write_funding_leaderboard, [wb, competitors, link_base]),
        ("Employee Growth", write_employee_growth, [wb, competitors, link_base]),
        ("SaaS Maturity", write_saas_maturity, [wb, competitors, link_base]),
        ("Classification Matrix", write_classification_matrix, [wb, competitors, link_base]),
        ("Efficiency & Profitability", write_efficiency_sheet, [wb, competitors, link_base]),
        ("Investment Efficiency", write_investment_efficiency_sheet, [wb, competitors, link_base]),
        ("Market Reach", write_market_reach_sheet, [wb, competitors, link_base]),
        ("AI Maturity", write_ai_maturity_sheet, [wb, competitors, link_base]),
        ("M&A Landscape", write_mna_landscape_sheet, [wb, competitors, link_base]),
        ("AI Talent Map", write_ai_talent_map_sheet, [wb, competitors, link_base]),
        ("Threat Timeline", write_threat_timeline_sheet, [wb, competitors, link_base]),
        ("Portfolio Risk", write_portfolio_risk_sheet, [wb, competitors, link_base]),
        ("Data Confidence", write_confidence_sheet, [wb, competitors, link_base]),
        ("Raw Data", write_raw_data, [wb, competitors, link_base]),
        ("Data Explorer", write_data_explorer, [wb, competitors, link_base]),
        ("Eneve vs Market", write_eneve_positioning, [wb, competitors, link_base]),
        ("Executive Summary", write_executive_summary, [wb, competitors, link_base]),
        ("Methodology", write_methodology_sheet, [wb, data]),
    ]

    for name, writer_fn, writer_args in iter_with_progress(sheets, "Writing sheets"):
        with timed_phase(f"Sheet: {name}", profile=profile):
            log.info("Writing %s...", name)
            writer_fn(*writer_args)

    with timed_phase("Sheet reordering", profile=profile):
        wb.move_sheet("Executive Summary", offset=-wb.sheetnames.index("Executive Summary"))
        wb.move_sheet("Eneve vs Market", offset=-wb.sheetnames.index("Eneve vs Market") + 1)

    with timed_phase("Workbook save", profile=profile):
        wb.save(str(output_path))
    log.info("Workbook saved to %s", output_path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate Excel dashboard from extracted competitor data.",
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        "--input",
        type=Path,
        help="Path to competitor_data.json (pre-extracted)",
    )
    group.add_argument(
        "--source",
        type=Path,
        help="Path to tickets/COMPETITION/ directory (extracts data first)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        required=True,
        help="Output .xlsx file path",
    )
    parser.add_argument(
        "--profile",
        action="store_true",
        help="Log wall-clock timing for each major pipeline phase",
    )
    args = parser.parse_args()

    try:
        with timed_phase("Total Excel pipeline", profile=args.profile):
            with timed_phase("Data loading", profile=args.profile):
                if args.input:
                    if not args.input.exists():
                        log.error("Input file not found: %s", args.input)
                        return 1
                    data = json.loads(args.input.read_text(encoding="utf-8"))
                else:
                    from extract_competitor_data import extract_all_competitors
                    data = extract_all_competitors(args.source, profile=args.profile)

            generate_workbook(data, args.output, profile=args.profile)
        return 0
    except json.JSONDecodeError as exc:
        log.error("Invalid JSON in input file: %s", exc)
        return 1
    except Exception as exc:
        log.error("Unexpected error: %s", exc)
        return 1


if __name__ == "__main__":
    sys.exit(main())
