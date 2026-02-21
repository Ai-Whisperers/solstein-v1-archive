"""Tests for generate_excel_report.py -- KPI logic and workbook structure."""

import io
from pathlib import Path
from unittest.mock import patch

import pytest

from generate_excel_report import (
    compute_market_stats,
    format_value,
    generate_workbook,
    text_sparkline,
    _get_max_timeline_length,
    _classify_mna,
    _mna_rationale,
    _cash_flow_indicator,
)


# ---------------------------------------------------------------------------
# text_sparkline
# ---------------------------------------------------------------------------

class TestTextSparkline:
    def test_ascending_values(self):
        result = text_sparkline([1.0, 3.0, 5.0, 7.0, 9.0])
        assert len(result) == 5
        assert result[0] != result[-1]

    def test_constant_values(self):
        result = text_sparkline([5.0, 5.0, 5.0])
        assert result == "―――"

    def test_single_value(self):
        assert text_sparkline([5.0]) == ""

    def test_empty_list(self):
        assert text_sparkline([]) == ""

    def test_with_nones(self):
        result = text_sparkline([None, 1.0, None, 5.0, 9.0])
        assert len(result) == 5
        assert result[0] == " "
        assert result[2] == " "

    def test_all_nones(self):
        assert text_sparkline([None, None]) == ""

    def test_two_values(self):
        result = text_sparkline([0.0, 10.0])
        assert len(result) == 2

    def test_no_output_starts_with_equals(self):
        """Regression: strings starting with '=' are treated as Excel formulas
        by openpyxl, causing 'Removed Records: Formula' warnings on open."""
        cases = [
            [5.0, 5.0, 5.0],
            [0.0, 0.0],
            [1.0, 3.0, 5.0],
            [None, 2.0, 2.0],
        ]
        for values in cases:
            result = text_sparkline(values)
            assert not result.startswith("="), (
                f"text_sparkline({values}) returned {result!r} which starts "
                f"with '=' and would be misinterpreted as an Excel formula"
            )


# ---------------------------------------------------------------------------
# _get_max_timeline_length
# ---------------------------------------------------------------------------

class TestGetMaxTimelineLength:
    def test_with_timelines(self, sample_competitor, eneve_competitor):
        result = _get_max_timeline_length([sample_competitor, eneve_competitor])
        assert result == 3  # sample_competitor has 3 revenue entries

    def test_empty_list(self):
        assert _get_max_timeline_length([]) == 0

    def test_empty_competitors(self, empty_competitor):
        assert _get_max_timeline_length([empty_competitor]) == 0


# ---------------------------------------------------------------------------
# format_value
# ---------------------------------------------------------------------------

class TestFormatValue:
    def test_none_returns_dashes(self):
        assert format_value(None) == "--"

    def test_number_passes_through(self):
        assert format_value(42) == 42

    def test_string_passes_through(self):
        assert format_value("hello") == "hello"

    def test_zero_passes_through(self):
        assert format_value(0) == 0


# ---------------------------------------------------------------------------
# compute_market_stats
# ---------------------------------------------------------------------------

class TestComputeMarketStats:
    def test_basic_stats(self):
        values = [10.0, 20.0, 30.0]
        companies = ["A", "B", "C"]
        result = compute_market_stats(values, companies)
        assert result["avg"] == pytest.approx(20.0)
        assert result["median"] == pytest.approx(20.0)
        assert result["best"] == pytest.approx(30.0)
        assert result["best_company"] == "C"

    def test_with_nones(self):
        values = [None, 15.0, None, 25.0]
        companies = ["A", "B", "C", "D"]
        result = compute_market_stats(values, companies)
        assert result["avg"] == pytest.approx(20.0)
        assert result["best"] == pytest.approx(25.0)
        assert result["best_company"] == "D"

    def test_all_nones(self):
        result = compute_market_stats([None, None], ["A", "B"])
        assert result["avg"] is None
        assert result["best"] is None
        assert result["best_company"] is None

    def test_empty(self):
        result = compute_market_stats([], [])
        assert result["avg"] is None

    def test_single_value(self):
        result = compute_market_stats([42.0], ["Solo"])
        assert result["avg"] == pytest.approx(42.0)
        assert result["best_company"] == "Solo"


# ---------------------------------------------------------------------------
# generate_workbook -- workbook structure
# ---------------------------------------------------------------------------

class TestGenerateWorkbook:
    def test_workbook_sheet_count_and_names(self, competitors_list, tmp_path):
        data = {
            "competitors": competitors_list,
            "missing_data": [],
            "metadata": {
                "total_folders": 2,
                "with_financial_data": 2,
                "without_financial_data": 0,
                "source_directory": str(tmp_path),
            },
        }
        output = tmp_path / "test_report.xlsx"
        generate_workbook(data, output)
        assert output.exists()

        from openpyxl import load_workbook
        wb = load_workbook(str(output))

        expected_sheets = {
            "Executive Summary",
            "Eneve vs Market",
            "Summary",
            "Revenue Leaderboard",
            "Funding Leaderboard",
            "Employee Growth",
            "SaaS Maturity",
            "Classification Matrix",
            "Efficiency & Profitability",
            "Market Reach",
            "M&A Landscape",
            "Raw Data",
            "Methodology",
        }
        actual_sheets = set(wb.sheetnames)
        # The default "Sheet" created by Workbook() may or may not be present
        actual_sheets.discard("Sheet")
        assert expected_sheets.issubset(actual_sheets), (
            f"Missing sheets: {expected_sheets - actual_sheets}"
        )

    def test_workbook_not_created_for_empty_data(self, tmp_path):
        data = {"competitors": [], "missing_data": [], "metadata": {}}
        output = tmp_path / "empty_report.xlsx"
        generate_workbook(data, output)
        assert not output.exists()

    def test_mna_landscape_sheet_present(self, competitors_list, tmp_path):
        data = {
            "competitors": competitors_list,
            "missing_data": [],
            "metadata": {
                "total_folders": 2,
                "with_financial_data": 2,
                "without_financial_data": 0,
                "source_directory": str(tmp_path),
            },
        }
        output = tmp_path / "test_mna.xlsx"
        generate_workbook(data, output)

        from openpyxl import load_workbook
        wb = load_workbook(str(output))
        assert "M&A Landscape" in wb.sheetnames
        ws = wb["M&A Landscape"]
        header_values = [cell.value for cell in ws[1]]
        assert "M&A Role" in header_values
        assert "Acquirer/Target Rationale" in header_values

    def test_summary_sheet_has_header_row(self, competitors_list, tmp_path):
        data = {
            "competitors": competitors_list,
            "missing_data": [],
            "metadata": {
                "total_folders": 2,
                "with_financial_data": 2,
                "without_financial_data": 0,
                "source_directory": str(tmp_path),
            },
        }
        output = tmp_path / "test_headers.xlsx"
        generate_workbook(data, output)

        from openpyxl import load_workbook
        wb = load_workbook(str(output))
        ws = wb["Summary"]
        header_values = [cell.value for cell in ws[1]]
        assert header_values[0] is not None
        assert any("Company" in str(h) or "Rank" in str(h) for h in header_values if h)


# ---------------------------------------------------------------------------
# M&A Landscape classification logic
# ---------------------------------------------------------------------------

class TestClassifyMna:
    def test_high_mna_score_is_acquirer(self, sample_competitor):
        """M&A score >= 7 should classify as Likely Acquirer."""
        sample_competitor["scorecard"]["dimensions"]["M&A Activity"]["score"] = 9.0
        assert _classify_mna(sample_competitor) == "Likely Acquirer"

    def test_eneve_always_neutral(self, eneve_competitor):
        """Eneve is the reference company and should always be Neutral."""
        assert _classify_mna(eneve_competitor) == "Neutral"

    def test_small_unfunded_low_scores_is_target(self):
        """Small revenue + no funding + low composite + low SaaS => Likely Target."""
        comp = {
            "company_name": "SmallCo",
            "folder": "smallco",
            "scorecard": {
                "dimensions": {
                    "M&A Activity": {"score": 2.0},
                    "Funding Momentum": {"score": 1.0},
                    "SaaS Maturity": {"score": 3.0},
                },
                "composite_score": 3.5,
                "classification": "Dinosaur",
            },
            "revenue": {"latest_revenue_eur_m": 15.0},
            "funding": {"total_raised_text": None},
        }
        assert _classify_mna(comp) == "Likely Target"

    def test_strong_self_funded_is_self_sustaining(self):
        """High composite, self-funded, significant revenue => Self-Sustaining."""
        comp = {
            "company_name": "StableCo",
            "folder": "stableco",
            "scorecard": {
                "dimensions": {
                    "M&A Activity": {"score": 3.0},
                    "Funding Momentum": {"score": 2.0},
                    "SaaS Maturity": {"score": 6.0},
                },
                "composite_score": 5.5,
                "classification": "Riser",
            },
            "revenue": {"latest_revenue_eur_m": 80.0},
            "funding": {"total_raised_text": None},
        }
        assert _classify_mna(comp) == "Self-Sustaining"

    def test_moderate_metrics_is_neutral(self):
        """Competitor with moderate metrics that don't hit any threshold => Neutral."""
        comp = {
            "company_name": "MidCo",
            "folder": "midco",
            "scorecard": {
                "dimensions": {
                    "M&A Activity": {"score": 4.0},
                    "Funding Momentum": {"score": 5.0},
                    "SaaS Maturity": {"score": 6.0},
                },
                "composite_score": 5.0,
                "classification": "Steady",
            },
            "revenue": {"latest_revenue_eur_m": 60.0},
            "funding": {"total_raised_text": "EUR 20M"},
        }
        assert _classify_mna(comp) == "Neutral"


class TestMnaRationale:
    def test_known_narrative_for_hansen(self):
        comp = {"folder": "hansen-technologies", "scorecard": {}, "revenue": {}, "funding": {}}
        result = _mna_rationale(comp, "Likely Acquirer")
        assert "powercloud" in result.lower()

    def test_acquirer_rationale_includes_mna_score(self):
        comp = {
            "folder": "generic-acquirer",
            "scorecard": {
                "dimensions": {"M&A Activity": {"score": 8.0}},
                "composite_score": 7.0,
            },
            "revenue": {},
            "funding": {},
        }
        result = _mna_rationale(comp, "Likely Acquirer")
        assert "M&A score" in result

    def test_target_rationale_includes_vulnerable(self):
        comp = {
            "folder": "small-target",
            "scorecard": {
                "dimensions": {"M&A Activity": {"score": 1.0}, "SaaS Maturity": {"score": 2.0}},
                "composite_score": 2.5,
            },
            "revenue": {"latest_revenue_eur_m": 10.0},
            "funding": {"total_raised_text": None},
        }
        result = _mna_rationale(comp, "Likely Target")
        assert "Vulnerable" in result


class TestCashFlowIndicator:
    def test_high_funding_is_strong(self):
        comp = {
            "scorecard": {"dimensions": {"Funding Momentum": {"score": 9.0}}},
            "revenue": {"latest_revenue_eur_m": 50.0},
        }
        assert _cash_flow_indicator(comp) == "Strong (funded)"

    def test_large_revenue_is_strong(self):
        comp = {
            "scorecard": {"dimensions": {"Funding Momentum": {"score": 3.0}}},
            "revenue": {"latest_revenue_eur_m": 200.0},
        }
        assert _cash_flow_indicator(comp) == "Strong (large co)"

    def test_moderate_revenue(self):
        comp = {
            "scorecard": {"dimensions": {"Funding Momentum": {"score": 3.0}}},
            "revenue": {"latest_revenue_eur_m": 50.0},
        }
        assert _cash_flow_indicator(comp) == "Moderate"

    def test_small_revenue_is_limited(self):
        comp = {
            "scorecard": {"dimensions": {"Funding Momentum": {"score": 1.0}}},
            "revenue": {"latest_revenue_eur_m": 5.0},
        }
        assert _cash_flow_indicator(comp) == "Limited"
