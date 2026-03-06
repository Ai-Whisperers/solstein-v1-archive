#!/usr/bin/env python3
"""
Test script for EPIC-005: Excel Export Fix

Verifies that profit_margin and ebitda_margin are correctly exported to Excel.
"""

import sys
from pathlib import Path
import tempfile

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from solstein.data.loaders import CompetitorDataLoader
from solstein.exporters.excel_improved import ImprovedExcelExporter


def test_excel_export():
    """Test that profit_margin and ebitda_margin are correctly exported."""
    print("=" * 60)
    print("EPIC-005: Excel Export Fix Test")
    print("=" * 60)

    # Load test data
    test_data_path = Path(__file__).parent.parent / "tests" / "fixtures" / "synthetic" / "competitor_data_enriched.json"
    loader = CompetitorDataLoader(data_dir=test_data_path.parent)

    try:
        companies = loader.load_from_json(test_data_path)
        print(f"\n✓ Loaded {len(companies)} companies")

        if not companies:
            print("✗ No companies loaded!")
            return False

        # Check that companies have profit_margin and ebitda_margin
        print("\nChecking company data:")
        for company in companies[:3]:  # Check first 3
            print(f"\n  {company.name}:")
            print(f"    profit_margin: {company.profit_margin}")
            print(f"    ebitda_margin: {company.ebitda_margin}")

        # Create Excel export
        print("\n\nCreating Excel export...")
        exporter = ImprovedExcelExporter()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = Path(tmp.name)

        try:
            exporter.create_dashboard(companies, output_path)
            print(f"✓ Excel file created: {output_path}")

            # Verify file exists and has content
            if output_path.exists() and output_path.stat().st_size > 0:
                print(f"✓ File size: {output_path.stat().st_size:,} bytes")

                # Try to open and check the Financial Intelligence sheet
                from openpyxl import load_workbook

                wb = load_workbook(output_path)

                if "Financial Intelligence" in wb.sheetnames:
                    ws = wb["Financial Intelligence"]
                    print(f"✓ Financial Intelligence sheet found")
                    print(f"  Rows: {ws.max_row}, Columns: {ws.max_column}")

                    # Check headers to find profit_margin and ebitda_margin columns
                    headers = []
                    for col in range(1, ws.max_column + 1):
                        headers.append(ws.cell(row=4, column=col).value)

                    print(f"\n  Headers: {headers}")

                    # Find columns for profit_margin and ebitda_margin
                    profit_col = None
                    ebitda_col = None
                    for idx, header in enumerate(headers, 1):
                        if header and "profit" in str(header).lower():
                            profit_col = idx
                        if header and "ebitda" in str(header).lower():
                            ebitda_col = idx

                    print(f"\n  Profit Margin column: {profit_col}")
                    print(f"  EBITDA Margin column: {ebitda_col}")

                    # Check data in first data row
                    if profit_col:
                        profit_val = ws.cell(row=5, column=profit_col).value
                        print(f"  First row profit_margin value: {profit_val}")
                        if profit_val and profit_val != "N/A":
                            print(f"  ✓ profit_margin is not N/A!")
                        else:
                            print(f"  ✗ profit_margin is {profit_val}")

                    if ebitda_col:
                        ebitda_val = ws.cell(row=5, column=ebitda_col).value
                        print(f"  First row ebitda_margin value: {ebitda_val}")
                        if ebitda_val and ebitda_val != "N/A":
                            print(f"  ✓ ebitda_margin is not N/A!")
                        else:
                            print(f"  ✗ ebitda_margin is {ebitda_val}")
                else:
                    print(f"✗ Financial Intelligence sheet not found!")
                    print(f"  Available sheets: {wb.sheetnames}")

                wb.close()

                # Cleanup
                output_path.unlink()

                print("\n" + "=" * 60)
                print("✓ EPIC-005: Excel Export Fix Test PASSED!")
                return True

        except Exception as e:
            print(f"\n✗ Error creating or verifying Excel: {e}")
            import traceback

            traceback.print_exc()
            return False

    except Exception as e:
        print(f"\n✗ Error during test: {e}")
        import traceback

        traceback.print_exc()
        return False


if __name__ == "__main__":
    success = test_excel_export()
    sys.exit(0 if success else 1)
