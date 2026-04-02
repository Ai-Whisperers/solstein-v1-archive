"""Streaming Excel exporter for large datasets.

STORY-112: Uses OpenPyXL write_only mode to generate Excel files
row-by-row, keeping memory usage O(1) relative to dataset size.

In write_only mode, cell-level access (ws.cell, ws.merge_cells,
auto_adjust_columns) is not supported. Rows are appended sequentially
and cannot be revisited. This is a deliberate trade-off: constant
memory vs. formatting flexibility.

Limitations documented per STORY-112 acceptance criteria:
- Column width auto-calculation is disabled (fixed widths used)
- Cell merging is not supported (headers use plain rows)
- Individual cell styling is limited (row-level only)

Sheets generated:
1. Summary — company overview with key metrics
2. Companies — full company detail rows
3. Signals — signal data per company
4. Financials — financial metrics per company
"""

from __future__ import annotations

from collections.abc import Sequence
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from loguru import logger
from openpyxl import Workbook

from ..domain.models import Company


def _safe_str(value: Any) -> str | None:
    """Convert value to string, return None for None/empty."""
    if value is None:
        return None
    s = str(value)
    return s if s else None


def _safe_num(value: Any) -> float | int | None:
    """Convert value to number, return None if not numeric."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


class StreamingExcelExporter:
    """Streaming Excel exporter using write_only mode.

    Memory usage is O(1) relative to dataset size — only one row
    is in memory at a time. Suitable for exports of 10,000+ rows.
    """

    def __init__(self) -> None:
        """Initialize the streaming exporter."""
        self._sheets_completed = 0
        self._total_sheets = 4  # Summary, Companies, Signals, Financials

    def create_dashboard(
        self,
        profiles: Sequence[Company],
        output_path: Path,
        progress_callback: Any | None = None,
    ) -> None:
        """Create a streaming Excel dashboard.

        Args:
            profiles: Sequence of company profiles to export.
            output_path: Path to save the Excel file.
            progress_callback: Optional callable(pct: int) called
                after each sheet completes (0-100).
        """
        logger.info(
            "[StreamingExcel] Starting export of %d companies to %s",
            len(profiles),
            output_path,
        )

        wb = Workbook(write_only=True)
        self._sheets_completed = 0

        try:
            self._write_summary_sheet(wb, profiles)
            self._report_progress(progress_callback)

            self._write_companies_sheet(wb, profiles)
            self._report_progress(progress_callback)

            self._write_signals_sheet(wb, profiles)
            self._report_progress(progress_callback)

            self._write_financials_sheet(wb, profiles)
            self._report_progress(progress_callback)

            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(output_path))

            logger.info(
                "[StreamingExcel] Export complete: %s (%d companies)",
                output_path,
                len(profiles),
            )
        except (OSError, ValueError, TypeError) as exc:
            logger.error(
                "[StreamingExcel] Export failed: %s",
                exc,
            )
            raise

    def _report_progress(
        self,
        callback: Any | None,
    ) -> None:
        """Report progress after sheet completion."""
        self._sheets_completed += 1
        pct = int((self._sheets_completed / self._total_sheets) * 100)
        if callback is not None:
            try:
                callback(pct)
            except (TypeError, RuntimeError) as exc:
                logger.warning(
                    "[StreamingExcel] Progress callback failed: %s",
                    exc,
                )

    def _write_summary_sheet(
        self,
        wb: Workbook,
        profiles: Sequence[Company],
    ) -> None:
        """Write the Summary sheet with overview metrics."""
        ws = wb.create_sheet("Summary")

        # Header row
        ws.append(
            [
                "Export Summary",
                None,
                None,
                None,
            ]
        )
        ws.append(
            [
                "Generated",
                datetime.now(timezone.utc).isoformat(),
            ]
        )
        ws.append(
            [
                "Total Companies",
                len(profiles),
            ]
        )
        ws.append([])  # blank row

        # Company summary table
        ws.append(
            [
                "Company",
                "Industry",
                "Classification",
                "Overall Score",
                "Growth Score",
                "AI Maturity",
            ]
        )

        for company in profiles:
            ws.append(
                [
                    _safe_str(company.name),
                    _safe_str(company.industry),
                    _safe_str(company.classification),
                    _safe_num(getattr(company, "overall_score", None)),
                    _safe_num(getattr(company, "growth_score", None)),
                    _safe_str(getattr(company, "ai_maturity", None)),
                ]
            )

    def _write_companies_sheet(
        self,
        wb: Workbook,
        profiles: Sequence[Company],
    ) -> None:
        """Write the Companies sheet with full detail rows."""
        ws = wb.create_sheet("Companies")

        # Header
        ws.append(
            [
                "Name",
                "Industry",
                "Classification",
                "Description",
                "Website",
                "Country",
                "Employee Count",
                "Founded Year",
                "Revenue",
                "Revenue Growth",
                "Total Funding",
                "Last Funding Round",
            ]
        )

        for company in profiles:
            ws.append(
                [
                    _safe_str(company.name),
                    _safe_str(company.industry),
                    _safe_str(company.classification),
                    _safe_str(getattr(company, "description", None)),
                    _safe_str(getattr(company, "website", None)),
                    _safe_str(getattr(company, "country", None)),
                    _safe_num(getattr(company, "employee_count", None)),
                    _safe_num(getattr(company, "founded_year", None)),
                    _safe_num(getattr(company, "revenue", None)),
                    _safe_num(getattr(company, "revenue_growth", None)),
                    _safe_num(getattr(company, "total_funding", None)),
                    _safe_str(getattr(company, "last_funding_round", None)),
                ]
            )

    def _write_signals_sheet(
        self,
        wb: Workbook,
        profiles: Sequence[Company],
    ) -> None:
        """Write the Signals sheet with signal data."""
        ws = wb.create_sheet("Signals")

        ws.append(
            [
                "Company",
                "Signal Type",
                "Signal Value",
                "Confidence",
                "Source",
            ]
        )

        for company in profiles:
            signals = getattr(company, "signals", None) or []
            if not signals:
                # Write a placeholder row for companies with no signals
                ws.append(
                    [
                        _safe_str(company.name),
                        "N/A",
                        None,
                        None,
                        None,
                    ]
                )
                continue

            for signal in signals:
                if isinstance(signal, dict):
                    ws.append(
                        [
                            _safe_str(company.name),
                            _safe_str(signal.get("type")),
                            _safe_str(signal.get("value")),
                            _safe_num(signal.get("confidence")),
                            _safe_str(signal.get("source")),
                        ]
                    )
                else:
                    ws.append(
                        [
                            _safe_str(company.name),
                            _safe_str(getattr(signal, "type", None)),
                            _safe_str(getattr(signal, "value", None)),
                            _safe_num(getattr(signal, "confidence", None)),
                            _safe_str(getattr(signal, "source", None)),
                        ]
                    )

    def _write_financials_sheet(
        self,
        wb: Workbook,
        profiles: Sequence[Company],
    ) -> None:
        """Write the Financials sheet with financial metrics."""
        ws = wb.create_sheet("Financials")

        ws.append(
            [
                "Company",
                "Revenue",
                "Revenue Growth (%)",
                "Gross Margin (%)",
                "EBITDA",
                "Net Income",
                "Total Funding",
                "Burn Rate",
                "Runway (months)",
            ]
        )

        for company in profiles:
            ws.append(
                [
                    _safe_str(company.name),
                    _safe_num(getattr(company, "revenue", None)),
                    _safe_num(getattr(company, "revenue_growth", None)),
                    _safe_num(getattr(company, "gross_margin", None)),
                    _safe_num(getattr(company, "ebitda", None)),
                    _safe_num(getattr(company, "net_income", None)),
                    _safe_num(getattr(company, "total_funding", None)),
                    _safe_num(getattr(company, "burn_rate", None)),
                    _safe_num(getattr(company, "runway_months", None)),
                ]
            )
