"""Excel sheet generators.

EPIC-022: Extracted from ImprovedExcelExporter for modularity.
Each sheet type has its own generator function.
"""

from typing import Any

from loguru import logger

from ...domain.models import Company
from .styles import ExcelStyles, LayoutConstants
from .utils import auto_adjust_columns, format_number, format_percentage, safe_get, safe_get_financial


def add_title_banner(ws: Any, styles: ExcelStyles, title: str, subtitle: str = "") -> None:
    """Add a title banner to the worksheet.

    Args:
        ws: Worksheet to add banner to
        styles: Excel styles
        title: Main title
        subtitle: Optional subtitle
    """
    from openpyxl.utils import get_column_letter

    # Merge cells for title
    ws.merge_cells(
        start_row=LayoutConstants.TITLE_ROW,
        start_column=LayoutConstants.TITLE_COLUMN,
        end_row=LayoutConstants.TITLE_ROW,
        end_column=10,
    )

    # Add title
    title_cell = ws.cell(row=LayoutConstants.TITLE_ROW, column=LayoutConstants.TITLE_COLUMN)
    title_cell.value = title
    title_cell.font = styles.title_font
    title_cell.fill = styles.header_fill
    title_cell.alignment = styles.header_alignment

    # Add subtitle if provided
    if subtitle:
        ws.merge_cells(
            start_row=LayoutConstants.SUBTITLE_ROW,
            start_column=LayoutConstants.TITLE_COLUMN,
            end_row=LayoutConstants.SUBTITLE_ROW,
            end_column=10,
        )
        subtitle_cell = ws.cell(row=LayoutConstants.SUBTITLE_ROW, column=LayoutConstants.TITLE_COLUMN)
        subtitle_cell.value = subtitle
        subtitle_cell.font = styles.subheader_font
        subtitle_cell.fill = styles.subheader_fill
        subtitle_cell.alignment = styles.subheader_alignment

    # Set row heights
    ws.row_dimensions[LayoutConstants.TITLE_ROW].height = 30
    if subtitle:
        ws.row_dimensions[LayoutConstants.SUBTITLE_ROW].height = 25


def write_headers(ws: Any, styles: ExcelStyles, headers: list[str], row: int | None = None) -> None:
    """Write column headers to the worksheet.

    Args:
        ws: Worksheet to write to
        styles: Excel styles
        headers: List of header strings
        row: Row number (defaults to HEADER_ROW)
    """
    if row is None:
        row = LayoutConstants.HEADER_ROW

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles.header_font
        cell.fill = styles.header_fill
        cell.alignment = styles.header_alignment
        cell.border = styles.standard_border


def add_executive_summary(ws: Any, styles: ExcelStyles, profiles: list[Company]) -> None:
    """Add executive summary sheet.

    Args:
        ws: Worksheet to add to
        styles: Excel styles
        profiles: List of company profiles
    """
    logger.info(f"Adding executive summary for {len(profiles)} companies")

    # Title
    add_title_banner(ws, styles, "Executive Summary", "Market Intelligence Dashboard")

    # Headers
    headers = ["Company", "Industry", "Revenue (€M)", "Growth", "AI Score", "Tier", "Threat Level"]
    write_headers(ws, styles, headers)

    # Data rows
    for idx, company in enumerate(profiles):
        row = LayoutConstants.DATA_START_ROW + idx

        # Company name
        cell = ws.cell(row=row, column=1)
        cell.value = safe_get(company, "name", "Unknown")
        cell.font = styles.data_font
        cell.alignment = styles.data_alignment
        cell.fill = styles.get_row_fill(idx, safe_get(company, "classification"))

        # Industry
        cell = ws.cell(row=row, column=2)
        cell.value = safe_get(company, "industry", "N/A")
        cell.font = styles.data_font
        cell.alignment = styles.data_alignment
        cell.fill = styles.get_row_fill(idx)

        # Revenue
        cell = ws.cell(row=row, column=3)
        revenue = safe_get_financial(company, "revenue_eur_m")
        cell.value = format_number(revenue, 1, "M")
        cell.font = styles.data_font
        cell.alignment = styles.number_alignment
        cell.fill = styles.get_row_fill(idx)

        # Growth
        cell = ws.cell(row=row, column=4)
        growth = safe_get_financial(company, "growth_rate_pct")
        cell.value = format_percentage(growth, 1)
        cell.font = styles.data_font
        cell.alignment = styles.number_alignment
        cell.fill = styles.get_row_fill(idx)

        # AI Score
        cell = ws.cell(row=row, column=5)
        ai_score = safe_get(company, "ai_score")
        cell.value = format_number(ai_score, 2)
        cell.font = styles.data_font
        cell.alignment = styles.number_alignment
        cell.fill = styles.get_row_fill(idx)

        # Tier
        cell = ws.cell(row=row, column=6)
        cell.value = safe_get(company, "tier", "Unknown")
        cell.font = styles.data_font
        cell.alignment = styles.data_alignment
        cell.fill = styles.get_row_fill(idx)

        # Threat Level
        cell = ws.cell(row=row, column=7)
        cell.value = safe_get(company, "threat_level", "Unknown")
        cell.font = styles.data_font
        cell.alignment = styles.data_alignment
        cell.fill = styles.get_row_fill(idx)

    # Auto-adjust columns
    auto_adjust_columns(ws)

    logger.info("Executive summary complete")


def add_market_rankings(ws: Any, styles: ExcelStyles, profiles: list[Company]) -> None:
    """Add market rankings sheet.

    Args:
        ws: Worksheet to add to
        styles: Excel styles
        profiles: List of company profiles
    """
    logger.info(f"Adding market rankings for {len(profiles)} companies")

    # Title
    add_title_banner(ws, styles, "Market Rankings", "Competitive Position Analysis")

    # Headers
    headers = ["Rank", "Company", "Market Share", "Competitive Score", "Growth Rate", "Employees"]
    write_headers(ws, styles, headers)

    # Sort by competitive score
    sorted_profiles = sorted(
        profiles,
        key=lambda x: safe_get(x, "competitive_position_score", 0) or 0,
        reverse=True,
    )

    # Data rows
    for idx, company in enumerate(sorted_profiles):
        row = LayoutConstants.DATA_START_ROW + idx

        # Rank
        cell = ws.cell(row=row, column=1)
        cell.value = idx + 1
        cell.font = styles.data_font
        cell.alignment = styles.number_alignment
        cell.fill = styles.get_row_fill(idx, safe_get(company, "classification"))

        # Company name
        cell = ws.cell(row=row, column=2)
        cell.value = safe_get(company, "name", "Unknown")
        cell.font = styles.data_font
        cell.alignment = styles.data_alignment
        cell.fill = styles.get_row_fill(idx)

        # Market Share
        cell = ws.cell(row=row, column=3)
        market_share = safe_get(company, "market_share_pct")
        cell.value = format_percentage(market_share, 1)
        cell.font = styles.data_font
        cell.alignment = styles.number_alignment
        cell.fill = styles.get_row_fill(idx)

        # Competitive Score
        cell = ws.cell(row=row, column=4)
        score = safe_get(company, "competitive_position_score")
        cell.value = format_number(score, 2)
        cell.font = styles.data_font
        cell.alignment = styles.number_alignment
        cell.fill = styles.get_row_fill(idx)

        # Growth Rate
        cell = ws.cell(row=row, column=5)
        growth = safe_get_financial(company, "growth_rate_pct")
        cell.value = format_percentage(growth, 1)
        cell.font = styles.data_font
        cell.alignment = styles.number_alignment
        cell.fill = styles.get_row_fill(idx)

        # Employees
        cell = ws.cell(row=row, column=6)
        employees = safe_get(company, "employee_count")
        cell.value = format_number(employees, 0)
        cell.font = styles.data_font
        cell.alignment = styles.number_alignment
        cell.fill = styles.get_row_fill(idx)

    # Auto-adjust columns
    auto_adjust_columns(ws)

    logger.info("Market rankings complete")


def add_financial_intelligence(ws: Any, styles: ExcelStyles, profiles: list[Company]) -> None:
    """Add financial intelligence sheet.

    Args:
        ws: Worksheet to add to
        styles: Excel styles
        profiles: List of company profiles
    """
    logger.info(f"Adding financial intelligence for {len(profiles)} companies")

    # Title
    add_title_banner(ws, styles, "Financial Intelligence", "Revenue, Funding & Valuation")

    # Headers
    headers = [
        "Company",
        "Revenue (€M)",
        "Growth Rate",
        "Profit Margin",
        "Total Funding",
        "Latest Valuation",
        "Investors",
    ]
    write_headers(ws, styles, headers)

    # Data rows
    for idx, company in enumerate(profiles):
        row = LayoutConstants.DATA_START_ROW + idx

        # Company name
        cell = ws.cell(row=row, column=1)
        cell.value = safe_get(company, "name", "Unknown")
        cell.font = styles.data_font
        cell.alignment = styles.data_alignment
        cell.fill = styles.get_row_fill(idx, safe_get(company, "classification"))

        # Revenue
        cell = ws.cell(row=row, column=2)
        revenue = safe_get_financial(company, "revenue_eur_m")
        cell.value = format_number(revenue, 1, "M")
        cell.font = styles.data_font
        cell.alignment = styles.number_alignment
        cell.fill = styles.get_row_fill(idx)

        # Growth Rate
        cell = ws.cell(row=row, column=3)
        growth = safe_get_financial(company, "growth_rate_pct")
        cell.value = format_percentage(growth, 1)
        cell.font = styles.data_font
        cell.alignment = styles.number_alignment
        cell.fill = styles.get_row_fill(idx)

        # Profit Margin
        cell = ws.cell(row=row, column=4)
        margin = safe_get_financial(company, "profit_margin_pct")
        cell.value = format_percentage(margin, 1)
        cell.font = styles.data_font
        cell.alignment = styles.number_alignment
        cell.fill = styles.get_row_fill(idx)

        # Total Funding
        cell = ws.cell(row=row, column=5)
        funding = safe_get_financial(company, "total_funding_raised_eur")
        cell.value = format_number(funding, 1, "M")
        cell.font = styles.data_font
        cell.alignment = styles.number_alignment
        cell.fill = styles.get_row_fill(idx)

        # Latest Valuation
        cell = ws.cell(row=row, column=6)
        valuation = safe_get_financial(company, "latest_valuation_eur")
        cell.value = format_number(valuation, 1, "M")
        cell.font = styles.data_font
        cell.alignment = styles.number_alignment
        cell.fill = styles.get_row_fill(idx)

        # Investors
        cell = ws.cell(row=row, column=7)
        investors = safe_get(company, "lead_investors", [])
        cell.value = ", ".join(investors) if investors else "N/A"
        cell.font = styles.data_font
        cell.alignment = styles.data_alignment
        cell.fill = styles.get_row_fill(idx)

    # Auto-adjust columns
    auto_adjust_columns(ws)

    logger.info("Financial intelligence complete")
