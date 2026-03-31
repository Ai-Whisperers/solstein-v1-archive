"""Extended Excel sheet generators for STORY-125.

STORY-125: Adds Revenue History and Advanced Data sheets to restore
the 20 dropped fields identified in the EPIC-033 forensic audit.
"""

import json
import re
from typing import Any

from loguru import logger

from ...domain.models import Company
from ...exporters.export_schema import get_headers_for_sheet
from .sheets import add_title_banner, write_headers
from .styles import ExcelStyles, LayoutConstants
from .utils import auto_adjust_columns, format_number, safe_get


def _is_valid_url(value: str) -> bool:
    """Check if a string looks like a valid URL.

    Args:
        value: String to check

    Returns:
        True if the string appears to be a URL
    """
    return bool(re.match(r"https?://", value))


def _serialize_list(items: list[Any], separator: str = ", ") -> str:
    """Serialize a list to a human-readable string.

    Args:
        items: List of items to serialize
        separator: Separator between items

    Returns:
        Serialized string or 'N/A' if empty
    """
    if not items:
        return "N/A"
    return separator.join(str(item) for item in items)


def _serialize_dict_list(items: list[dict[str, Any]]) -> str:
    """Serialize a list of dicts to a human-readable string.

    Args:
        items: List of dictionaries

    Returns:
        Serialized string or 'N/A' if empty
    """
    if not items:
        return "N/A"
    summaries = []
    for item in items:
        parts = [f"{k}: {v}" for k, v in item.items() if v is not None]
        summaries.append("; ".join(parts))
    return " | ".join(summaries)


def add_revenue_history(ws: Any, styles: ExcelStyles, profiles: list[Company]) -> None:
    """Add Revenue History sheet with time-series revenue data.

    Renders revenue_timeline as one row per company-year combination,
    supporting variable-length time series across companies.

    Args:
        ws: Worksheet to add to
        styles: Excel styles
        profiles: List of company profiles
    """
    logger.info(f"Adding revenue history for {len(profiles)} companies")

    # Title
    add_title_banner(ws, styles, "Revenue History", "Historical Revenue Time Series", num_columns=4)

    # STORY-250: Headers derived from single authoritative schema — drift-proof
    headers = get_headers_for_sheet("Revenue History")
    write_headers(ws, styles, headers)

    # Data rows — one row per company-year
    current_row = LayoutConstants.DATA_START_ROW
    for company in profiles:
        timeline = safe_get(company, "revenue_timeline", [])
        if not timeline:
            # Even companies with no timeline get a row so they are visible
            cell = ws.cell(row=current_row, column=1)
            cell.value = safe_get(company, "name", "Unknown")
            cell.font = styles.data_font
            cell.alignment = styles.data_alignment

            cell = ws.cell(row=current_row, column=2)
            cell.value = "N/A"
            cell.font = styles.data_font
            cell.alignment = styles.data_alignment

            cell = ws.cell(row=current_row, column=3)
            cell.value = "N/A"
            cell.font = styles.data_font
            cell.alignment = styles.number_alignment

            cell = ws.cell(row=current_row, column=4)
            cell.value = "N/A"
            cell.font = styles.data_font
            cell.alignment = styles.data_alignment

            current_row += 1
            continue

        for entry in timeline:
            if not isinstance(entry, dict):
                continue

            row_idx = current_row - LayoutConstants.DATA_START_ROW
            fill = styles.get_row_fill(row_idx)

            # Company name
            cell = ws.cell(row=current_row, column=1)
            cell.value = safe_get(company, "name", "Unknown")
            cell.font = styles.data_font
            cell.alignment = styles.data_alignment
            cell.fill = fill

            # Year
            cell = ws.cell(row=current_row, column=2)
            cell.value = entry.get("year", "N/A")
            cell.font = styles.data_font
            cell.alignment = styles.data_alignment
            cell.fill = fill

            # Revenue
            cell = ws.cell(row=current_row, column=3)
            rev = entry.get("eur_millions", entry.get("revenue", entry.get("amount")))
            cell.value = format_number(rev, 1) if rev is not None else "N/A"
            cell.font = styles.data_font
            cell.alignment = styles.number_alignment
            cell.fill = fill

            # Source
            cell = ws.cell(row=current_row, column=4)
            cell.value = entry.get("source", entry.get("confidence", "N/A"))
            cell.font = styles.data_font
            cell.alignment = styles.data_alignment
            cell.fill = fill

            current_row += 1

    # Freeze header row
    ws.freeze_panes = ws.cell(row=LayoutConstants.DATA_START_ROW, column=1)

    auto_adjust_columns(ws)
    logger.info(f"Revenue history complete: {current_row - LayoutConstants.DATA_START_ROW} rows")


def add_advanced_data(ws: Any, styles: ExcelStyles, profiles: list[Company]) -> None:
    """Add Advanced Data sheet with corporate structure, notes, and provenance.

    Contains: parent_company, subsidiaries, acquisitions, notes, source_links,
    data_source_per_field (mapped from metric_sources), merge_conflicts
    (mapped from enrichment_quality_metrics).

    Args:
        ws: Worksheet to add to
        styles: Excel styles
        profiles: List of company profiles
    """
    from openpyxl.worksheet.hyperlink import Hyperlink

    logger.info(f"Adding advanced data for {len(profiles)} companies")

    # Title
    add_title_banner(
        ws, styles, "Advanced Data", "Corporate Structure, Provenance & Notes", num_columns=8
    )

    # STORY-250: Headers derived from single authoritative schema — drift-proof
    headers = get_headers_for_sheet("Advanced Data")
    write_headers(ws, styles, headers)

    # Data rows
    for idx, company in enumerate(profiles):
        row = LayoutConstants.DATA_START_ROW + idx
        fill = styles.get_row_fill(idx)

        # Company name
        cell = ws.cell(row=row, column=1)
        cell.value = safe_get(company, "name", "Unknown")
        cell.font = styles.data_font
        cell.alignment = styles.data_alignment
        cell.fill = fill

        # Parent Company
        cell = ws.cell(row=row, column=2)
        cell.value = safe_get(company, "parent_company", "N/A")
        cell.font = styles.data_font
        cell.alignment = styles.data_alignment
        cell.fill = fill

        # Subsidiaries
        cell = ws.cell(row=row, column=3)
        subs = safe_get(company, "subsidiaries", [])
        cell.value = _serialize_list(subs)
        cell.font = styles.data_font
        cell.alignment = styles.data_alignment
        cell.fill = fill

        # Acquisitions
        cell = ws.cell(row=row, column=4)
        acqs = safe_get(company, "acquisitions", [])
        cell.value = _serialize_dict_list(acqs)
        cell.font = styles.data_font
        cell.alignment = styles.data_alignment
        cell.fill = fill

        # Notes
        cell = ws.cell(row=row, column=5)
        cell.value = safe_get(company, "notes", "N/A")
        cell.font = styles.data_font
        cell.alignment = styles.data_alignment
        cell.fill = fill

        # Source Links — render as hyperlinks where possible
        cell = ws.cell(row=row, column=6)
        links = safe_get(company, "source_links", [])
        if links:
            # Put first valid URL as hyperlink, list all in cell value
            cell.value = _serialize_list(links, separator="\n")
            for link in links:
                if isinstance(link, str) and _is_valid_url(link):
                    try:
                        cell.hyperlink = Hyperlink(ref=cell.coordinate, target=link)
                    except Exception as e:
                        logger.debug(f"Could not set hyperlink for {link}: {e}")
                    break  # One hyperlink per cell in openpyxl
        else:
            cell.value = "N/A"
        cell.font = styles.data_font
        cell.alignment = styles.data_alignment
        cell.fill = fill

        # Data Source Per Field — mapped from metric_sources
        cell = ws.cell(row=row, column=7)
        sources = safe_get(company, "metric_sources", {})
        if sources:
            parts = [f"{field}: {', '.join(srcs)}" for field, srcs in sources.items()]
            cell.value = "; ".join(parts)
        else:
            cell.value = "N/A"
        cell.font = styles.data_font
        cell.alignment = styles.data_alignment
        cell.fill = fill

        # Merge Conflicts — mapped from enrichment_quality_metrics
        cell = ws.cell(row=row, column=8)
        quality = safe_get(company, "enrichment_quality_metrics", {})
        conflicts = quality.get("merge_conflicts", quality.get("conflicts", None)) if isinstance(quality, dict) else None
        if conflicts:
            cell.value = json.dumps(conflicts, default=str) if isinstance(conflicts, (dict, list)) else str(conflicts)
        else:
            cell.value = "None"
        cell.font = styles.data_font
        cell.alignment = styles.data_alignment
        cell.fill = fill

    # Freeze header row
    ws.freeze_panes = ws.cell(row=LayoutConstants.DATA_START_ROW, column=1)

    auto_adjust_columns(ws)
    logger.info("Advanced data complete")
