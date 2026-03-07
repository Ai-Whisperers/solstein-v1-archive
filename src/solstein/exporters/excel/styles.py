"""Excel styling constants and utilities.

EPIC-022: Extracted from ImprovedExcelExporter for modularity.
"""

from dataclasses import dataclass

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@dataclass(frozen=True)
class LayoutConstants:
    """Layout constants for Excel sheets."""

    TITLE_ROW = 1
    SUBTITLE_ROW = 2
    BLANK_ROW = 3
    HEADER_ROW = 4
    DATA_START_ROW = 5

    TITLE_COLUMN = 1

    # Column widths
    MIN_COLUMN_WIDTH = 8
    MAX_COLUMN_WIDTH = 60
    DEFAULT_COLUMN_WIDTH = 15
    DESCRIPTION_COLUMN_WIDTH = 40
    NAME_COLUMN_WIDTH = 25


@dataclass(frozen=True)
class ColorPalette:
    """Color palette for consistent styling."""

    OBSIDIAN = "0A0A0F"
    ONYX = "12121A"
    GOLD = "D4A843"
    GOLD_LIGHT = "F0D78C"
    EMERALD = "2ECC71"
    RUBY = "E74C3C"
    SAPPHIRE = "3498DB"
    SLATE = "1A1A2E"
    TEXT = "E8E8F0"
    TEXT_MUTED = "8888A0"

    # Row fills
    PHOENIX_FILL = "D5F5E3"  # Light Emerald
    SALT_FILL = "FEF9E7"  # Light Yellow
    LEAD_FILL = "FADBD8"  # Light Ruby
    EVEN_ROW = "F9F9FB"
    ODD_ROW = "FFFFFF"


class ExcelStyles:
    """Centralized style definitions for Excel exports."""

    def __init__(self):
        self.colors = ColorPalette()
        self._init_borders()
        self._init_fonts()
        self._init_fills()
        self._init_alignments()

    def _init_borders(self):
        """Initialize border styles."""
        self.standard_border = Border(
            left=Side(style="thin", color=self.colors.SLATE),
            right=Side(style="thin", color=self.colors.SLATE),
            top=Side(style="thin", color=self.colors.SLATE),
            bottom=Side(style="thin", color=self.colors.SLATE),
        )
        self.thick_bottom = Border(bottom=Side(style="medium", color=self.colors.GOLD))

    def _init_fonts(self):
        """Initialize font styles."""
        self.title_font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
        self.header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        self.subheader_font = Font(name="Calibri", size=11, bold=True, color=self.colors.GOLD)
        self.data_font = Font(name="Calibri", size=10)

    def _init_fills(self):
        """Initialize fill styles."""
        self.header_fill = PatternFill(
            start_color=self.colors.OBSIDIAN,
            end_color=self.colors.OBSIDIAN,
            fill_type="solid",
        )
        self.subheader_fill = PatternFill(
            start_color=self.colors.SLATE,
            end_color=self.colors.SLATE,
            fill_type="solid",
        )
        self.even_row_fill = PatternFill(
            start_color=self.colors.EVEN_ROW,
            end_color=self.colors.EVEN_ROW,
            fill_type="solid",
        )
        self.odd_row_fill = PatternFill(
            start_color=self.colors.ODD_ROW,
            end_color=self.colors.ODD_ROW,
            fill_type="solid",
        )

    def _init_alignments(self):
        """Initialize alignment styles."""
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        self.subheader_alignment = Alignment(horizontal="left", vertical="center")
        self.data_alignment = Alignment(horizontal="left", vertical="center", indent=1)
        self.number_alignment = Alignment(horizontal="right", vertical="center", indent=1)

    def get_row_fill(self, row_idx: int, classification: str | None = None) -> PatternFill:
        """Get appropriate fill for a row.

        Args:
            row_idx: Row index (0-based)
            classification: Company classification (phoenix, salt, lead)

        Returns:
            PatternFill for the row
        """
        if classification:
            if classification.lower() == "phoenix":
                return PatternFill(
                    start_color=self.colors.PHOENIX_FILL,
                    end_color=self.colors.PHOENIX_FILL,
                    fill_type="solid",
                )
            elif classification.lower() == "salt":
                return PatternFill(
                    start_color=self.colors.SALT_FILL,
                    end_color=self.colors.SALT_FILL,
                    fill_type="solid",
                )
            elif classification.lower() == "lead":
                return PatternFill(
                    start_color=self.colors.LEAD_FILL,
                    end_color=self.colors.LEAD_FILL,
                    fill_type="solid",
                )

        # Default alternating rows
        if row_idx % 2 == 0:
            return self.even_row_fill
        return self.odd_row_fill
