"""Export schema definition and validation for STORY-126.

This module is the single authoritative source of truth for the Excel export
structure. Any structural change to the export (adding, removing, moving, or
renaming a field) requires an explicit version bump here.

Usage:
    from solstein.exporters.export_schema import validate_export, EXPORT_SCHEMA_VERSION
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from loguru import logger
from openpyxl import load_workbook

# ─────────────────────────────────────────────
# Schema version — bump on any structural change
# ─────────────────────────────────────────────
EXPORT_SCHEMA_VERSION = "1.0"


# ─────────────────────────────────────────────
# Field specification
# ─────────────────────────────────────────────
@dataclass(frozen=True)
class FieldSpec:
    """Specification of a single exported field."""

    name: str
    header: str
    sheet: str
    data_type: str  # "string", "number", "percentage", "integer", "list", "structured"
    required: bool = True  # Whether the column header must always be present


# ─────────────────────────────────────────────
# Export schema definition
# ─────────────────────────────────────────────
EXPORT_SCHEMA: list[FieldSpec] = [
    # ── Executive Summary ──
    FieldSpec("name", "Company", "Executive Summary", "string"),
    FieldSpec("industry", "Industry", "Executive Summary", "string"),
    FieldSpec("revenue_eur_m", "Revenue (€M)", "Executive Summary", "number"),
    FieldSpec("growth_rate_pct", "Growth", "Executive Summary", "percentage"),
    FieldSpec("ai_score", "AI Score", "Executive Summary", "number"),
    FieldSpec("ai_readiness_score", "AI Readiness", "Executive Summary", "number"),
    FieldSpec("ai_readiness_tier", "AI Readiness Tier", "Executive Summary", "string"),
    FieldSpec("tier", "Tier", "Executive Summary", "string"),
    FieldSpec("threat_level", "Threat Level", "Executive Summary", "string"),
    FieldSpec("tech_stack", "Tech Stack", "Executive Summary", "list"),
    FieldSpec("key_customers", "Key Customers", "Executive Summary", "list"),
    FieldSpec("open_positions", "Open Positions", "Executive Summary", "integer"),
    FieldSpec("data_availability", "Data Availability", "Executive Summary", "string"),
    # ── Market Rankings ──
    FieldSpec("rank", "Rank", "Market Rankings", "integer"),
    FieldSpec("name_rankings", "Company", "Market Rankings", "string"),
    FieldSpec("market_share_pct", "Market Share", "Market Rankings", "percentage"),
    FieldSpec("competitive_position_score", "Competitive Score", "Market Rankings", "number"),
    FieldSpec("growth_rate_rankings", "Growth Rate", "Market Rankings", "percentage"),
    FieldSpec("employee_count", "Employees", "Market Rankings", "integer"),
    # ── Financial Intelligence ──
    FieldSpec("name_financial", "Company", "Financial Intelligence", "string"),
    FieldSpec("revenue_financial", "Revenue (€M)", "Financial Intelligence", "number"),
    FieldSpec("growth_rate_financial", "Growth Rate", "Financial Intelligence", "percentage"),
    FieldSpec("profit_margin", "Profit Margin", "Financial Intelligence", "percentage"),
    FieldSpec("total_funding", "Total Funding", "Financial Intelligence", "number"),
    FieldSpec("latest_valuation", "Latest Valuation", "Financial Intelligence", "number"),
    FieldSpec("lead_investors", "Investors", "Financial Intelligence", "list"),
    FieldSpec("funding_rounds", "Funding Rounds", "Financial Intelligence", "structured"),
    FieldSpec("funding_war_chest", "Funding War Chest", "Financial Intelligence", "string"),
    FieldSpec("revenue_cagr_5yr", "Revenue CAGR 5yr", "Financial Intelligence", "percentage"),
    FieldSpec("revenue_per_employee", "Revenue/Employee (€K)", "Financial Intelligence", "number"),
    FieldSpec("employee_cagr_3yr", "Employee CAGR 3yr", "Financial Intelligence", "percentage"),
    # ── Revenue History ──
    FieldSpec("name_revenue", "Company", "Revenue History", "string"),
    FieldSpec("year", "Year", "Revenue History", "integer"),
    FieldSpec("revenue_eur_m_history", "Revenue (EUR M)", "Revenue History", "number"),
    FieldSpec("source", "Source", "Revenue History", "string"),
    # ── Advanced Data ──
    FieldSpec("name_advanced", "Company", "Advanced Data", "string"),
    FieldSpec("parent_company", "Parent Company", "Advanced Data", "string"),
    FieldSpec("subsidiaries", "Subsidiaries", "Advanced Data", "list"),
    FieldSpec("acquisitions", "Acquisitions", "Advanced Data", "structured"),
    FieldSpec("notes", "Notes", "Advanced Data", "string"),
    FieldSpec("source_links", "Source Links", "Advanced Data", "list"),
    FieldSpec("data_source_per_field", "Data Sources Per Field", "Advanced Data", "structured"),
    FieldSpec("merge_conflicts", "Merge Conflicts", "Advanced Data", "structured"),
]


class ExportValidationError(Exception):
    """Raised when exported content does not match the export schema."""

    def __init__(self, missing_fields: list[str], wrong_sheet_fields: list[str] | None = None) -> None:
        self.missing_fields = missing_fields
        self.wrong_sheet_fields = wrong_sheet_fields or []
        parts = []
        if missing_fields:
            parts.append(f"Missing fields: {', '.join(missing_fields)}")
        if wrong_sheet_fields:
            parts.append(f"Wrong sheet: {', '.join(wrong_sheet_fields)}")
        super().__init__(f"Export schema validation failed. {'; '.join(parts)}")


def _get_sheet_headers(ws: Any, header_row: int = 4) -> set[str]:
    """Extract header values from a worksheet row.

    Args:
        ws: openpyxl worksheet
        header_row: Row number containing headers

    Returns:
        Set of header strings found in the row
    """
    headers: set[str] = set()
    for col in range(1, 50):
        val = ws.cell(row=header_row, column=col).value
        if val is not None:
            headers.add(str(val))
    return headers


def validate_export(file_path: str | Any, schema: list[FieldSpec] | None = None) -> None:
    """Validate an exported Excel file against the export schema.

    Checks that every required field's header is present on the correct sheet.
    Raises ExportValidationError if validation fails.

    Args:
        file_path: Path to the exported .xlsx file
        schema: Optional override schema (defaults to EXPORT_SCHEMA)

    Raises:
        ExportValidationError: If fields are missing or misplaced
    """
    if schema is None:
        schema = EXPORT_SCHEMA

    wb = load_workbook(str(file_path), read_only=True, data_only=True)

    # Build lookup: sheet_name -> set of headers
    sheet_headers: dict[str, set[str]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_headers[sheet_name] = _get_sheet_headers(ws)

    missing: list[str] = []
    wrong_sheet: list[str] = []

    for spec in schema:
        if not spec.required:
            continue

        target_headers = sheet_headers.get(spec.sheet, set())
        if spec.header not in target_headers:
            # Check if it's on a different sheet (wrong placement)
            found_elsewhere = any(
                spec.header in hdrs
                for name, hdrs in sheet_headers.items()
                if name != spec.sheet
            )
            if found_elsewhere:
                wrong_sheet.append(f"{spec.name} (expected on '{spec.sheet}', found elsewhere)")
            else:
                missing.append(f"{spec.name} ('{spec.header}' on '{spec.sheet}')")

    wb.close()

    if missing or wrong_sheet:
        raise ExportValidationError(missing, wrong_sheet)

    logger.info(f"Export schema validation passed (v{EXPORT_SCHEMA_VERSION}): {len(schema)} fields verified")


def get_schema_by_sheet() -> dict[str, list[FieldSpec]]:
    """Group schema fields by sheet name.

    Returns:
        Dictionary mapping sheet names to their field specs
    """
    result: dict[str, list[FieldSpec]] = {}
    for spec in EXPORT_SCHEMA:
        result.setdefault(spec.sheet, []).append(spec)
    return result


def generate_schema_docs() -> str:
    """Auto-generate markdown documentation from the export schema.

    Returns:
        Markdown string describing the full export schema
    """
    lines = [
        "# Excel Export Schema",
        "",
        f"**Schema Version**: {EXPORT_SCHEMA_VERSION}",
        "**Auto-generated from**: `src/solstein/exporters/export_schema.py`",
        "",
        "---",
        "",
    ]

    by_sheet = get_schema_by_sheet()
    for sheet_name, fields in by_sheet.items():
        lines.append(f"## {sheet_name}")
        lines.append("")
        lines.append("| Column | Header | Domain Field | Type | Required |")
        lines.append("|--------|--------|-------------|------|----------|")
        for i, spec in enumerate(fields):
            col_letter = chr(65 + i)  # A, B, C, ...
            req = "Yes" if spec.required else "No"
            lines.append(f"| {col_letter} | {spec.header} | `{spec.name}` | {spec.data_type} | {req} |")
        lines.append("")

    lines.append("---")
    lines.append("")
    lines.append("## Schema Changelog")
    lines.append("")
    lines.append("| Version | Date | Changes |")
    lines.append("|---------|------|---------|")
    lines.append("| 1.0 | 2026-03-27 | Initial schema: 5 sheets, all 20 STORY-125 fields + original fields |")
    lines.append("")

    return "\n".join(lines)
