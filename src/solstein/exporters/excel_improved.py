"""
Improved Excel exporter for SolStein dashboards.

EPIC-005: Excel Export Improvements
- Story 5.2: Fix headers to match data
- Story 5.3: Add null checks
- Story 5.4: Replace magic numbers with constants
- Story 5.5: Standardize sheet structure

This module provides a refactored, cleaner implementation of the Excel export
functionality with proper error handling and consistent formatting.
"""

from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from dataclasses import dataclass

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..config import Settings
from ..domain.models import Company, FinancialMetric


# Constants for sheet layout (Story 5.4: Replace magic numbers)
@dataclass(frozen=True)
class LayoutConstants:
    """Layout constants for Excel sheets."""
    TITLE_ROW = 1
    SUBTITLE_ROW = 2
    BLANK_ROW = 3
    HEADER_ROW = 4
    DATA_START_ROW = 5
    
    TITLE_COLUMN = 1
    
    # Column widths
    MIN_COLUMN_WIDTH = 8
    MAX_COLUMN_WIDTH = 60
    DEFAULT_COLUMN_WIDTH = 15
    DESCRIPTION_COLUMN_WIDTH = 40
    NAME_COLUMN_WIDTH = 25


@dataclass(frozen=True)
class ColorPalette:
    """Color palette for consistent styling."""
    OBSIDIAN = "0A0A0F"
    ONYX = "12121A"
    GOLD = "D4A843"
    GOLD_LIGHT = "F0D78C"
    EMERALD = "2ECC71"
    RUBY = "E74C3C"
    SAPPHIRE = "3498DB"
    SLATE = "1A1A2E"
    TEXT = "E8E8F0"
    TEXT_MUTED = "8888A0"
    
    # Row fills
    PHOENIX_FILL = "D5F5E3"  # Light Emerald
    SALT_FILL = "FEF9E7"     # Light Yellow
    LEAD_FILL = "FADBD8"     # Light Ruby
    EVEN_ROW = "F9F9FB"
    ODD_ROW = "FFFFFF"


class ExcelStyles:
    """Centralized style definitions for Excel exports."""
    
    def __init__(self):
        self.colors = ColorPalette()
        self._init_borders()
        self._init_fonts()
        self._init_fills()
        self._init_alignments()
    
    def _init_borders(self):
        """Initialize border styles."""
        self.standard_border = Border(
            left=Side(style="thin", color=self.colors.SLATE),
            right=Side(style="thin", color=self.colors.SLATE),
            top=Side(style="thin", color=self.colors.SLATE),
            bottom=Side(style="thin", color=self.colors.SLATE),
        )
        self.thick_bottom = Border(bottom=Side(style="medium", color=self.colors.GOLD))
    
    def _init_fonts(self):
        """Initialize font styles."""
        self.title_font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
        self.header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        self.subheader_font = Font(name="Calibri", size=11, bold=True, color=self.colors.GOLD)
        self.data_font = Font(name="Calibri", size=10)
    
    def _init_fills(self):
        """Initialize fill styles."""
        self.header_fill = PatternFill(
            start_color=self.colors.OBSIDIAN,
            end_color=self.colors.OBSIDIAN,
            fill_type="solid",
        )
        self.subheader_fill = PatternFill(
            start_color=self.colors.SLATE,
            end_color=self.colors.SLATE,
            fill_type="solid",
        )
        self.even_row_fill = PatternFill(
            start_color=self.colors.EVEN_ROW,
            end_color=self.colors.EVEN_ROW,
            fill_type="solid",
        )
        self.odd_row_fill = PatternFill(
            start_color=self.colors.ODD_ROW,
            end_color=self.colors.ODD_ROW,
            fill_type="solid",
        )
    
    def _init_alignments(self):
        """Initialize alignment styles."""
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        self.subheader_alignment = Alignment(horizontal="left", vertical="center")
        self.data_alignment = Alignment(horizontal="left", vertical="center", indent=1)
        self.number_alignment = Alignment(horizontal="right", vertical="center", indent=1)
    
    def get_row_fill(self, row_idx: int, classification: Optional[str] = None) -> PatternFill:
        """Get appropriate fill for a row based on index and classification."""
        if classification:
            if classification == "Phoenix":
                return PatternFill(start_color=self.colors.PHOENIX_FILL, end_color=self.colors.PHOENIX_FILL, fill_type="solid")
            elif classification == "Lead":
                return PatternFill(start_color=self.colors.LEAD_FILL, end_color=self.colors.LEAD_FILL, fill_type="solid")
            elif classification == "Salt":
                return PatternFill(start_color=self.colors.SALT_FILL, end_color=self.colors.SALT_FILL, fill_type="solid")
        
        # Default alternating rows
        return self.even_row_fill if row_idx % 2 == 0 else self.odd_row_fill


class ImprovedExcelExporter:
    """Improved Excel exporter with better error handling and consistency."""
    
    def __init__(self, template_path: Optional[Path] = None):
        self.settings = Settings()
        self.template_path = template_path
        self.styles = ExcelStyles()
        self.layout = LayoutConstants()
        self.colors = ColorPalette()
    
    def create_dashboard(self, profiles: list[Company], output_path: Path) -> None:
        """Create a multi-sheet Professional Intelligence Report."""
        logger.info(f"Generating Excel dashboard for {len(profiles)} companies")
        
        if not profiles:
            logger.warning("No profiles provided, creating empty dashboard")
            profiles = []
        
        wb = Workbook()
        
        # Create sheets in order
        sheets_config = [
            ("Executive Summary", self._add_executive_summary),
            ("Market Rankings", self._add_market_rankings),
            ("Financial Intelligence", self._add_financial_intelligence),
            ("Tech & AI Maturity", self._add_tech_ai_maturity),
            ("Company Details", self._add_company_details),
        ]
        
        for idx, (title, add_func) in enumerate(sheets_config):
            if idx == 0:
                ws = wb.active
                ws.title = title
            else:
                ws = wb.create_sheet(title)
            
            try:
                add_func(ws, profiles)
                self._auto_adjust_columns(ws)
            except Exception as e:
                logger.error(f"Error creating sheet '{title}': {e}")
                # Add error message to sheet
                ws.cell(row=1, column=1, value=f"Error creating sheet: {e}")
        
        # Save workbook
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info(f"Dashboard saved to {output_path}")
    
    def _add_title_banner(self, ws: Any, title: str, subtitle: str = "") -> None:
        """Add consistent title banner to sheet."""
        # Title
        title_cell = ws.cell(row=self.layout.TITLE_ROW, column=self.layout.TITLE_COLUMN, value=title)
        title_cell.font = self.styles.title_font
        title_cell.fill = self.styles.header_fill
        title_cell.alignment = self.styles.header_alignment
        
        # Subtitle
        if subtitle:
            subtitle_cell = ws.cell(row=self.layout.SUBTITLE_ROW, column=self.layout.TITLE_COLUMN, value=subtitle)
            subtitle_cell.font = self.styles.subheader_font
            subtitle_cell.fill = self.styles.subheader_fill
            subtitle_cell.alignment = self.styles.subheader_alignment
    
    def _write_headers(self, ws: Any, headers: list[str], row: Optional[int] = None) -> None:
        """Write headers with consistent styling."""
        row = row or self.layout.HEADER_ROW
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.header_font
            cell.fill = self.styles.header_fill
            cell.alignment = self.styles.header_alignment
            cell.border = self.styles.thick_bottom
    
    def _safe_get(self, obj: Any, attr: str, default: Any = "") -> Any:
        """Safely get attribute with null check (Story 5.3)."""
        if obj is None:
            return default
        
        try:
            value = getattr(obj, attr, default)
            return value if value is not None else default
        except (AttributeError, TypeError):
            return default
    
    def _safe_get_financial(self, company: Company, attr: str, default: Any = "") -> Any:
        """Safely get financial metric with null check."""
        if company is None or company.financials is None:
            return default
        return self._safe_get(company.financials, attr, default)
    
    def _format_number(self, value: Optional[float], decimals: int = 1, suffix: str = "") -> str:
        """Format number with null check."""
        if value is None:
            return "N/A"
        try:
            return f"{value:.{decimals}f}{suffix}"
        except (TypeError, ValueError):
            return "N/A"
    
    def _format_percentage(self, value: Optional[float], decimals: int = 1) -> str:
        """Format percentage with null check."""
        if value is None:
            return "N/A"
        try:
            return f"{value:.{decimals}f}%"
        except (TypeError, ValueError):
            return "N/A"
    
    def _add_executive_summary(self, ws: Any, profiles: list[Company]) -> None:
        """Executive Summary sheet."""
        self._add_title_banner(ws, "Executive Summary", "Solstein Competitive Intelligence Dashboard")
        
        # Calculate summary statistics with null checks
        total = len(profiles)
        
        # Classification counts
        phoenix_count = sum(1 for p in profiles if getattr(p, 'classification', None) == "Phoenix")
        salt_count = sum(1 for p in profiles if getattr(p, 'classification', None) == "Salt")
        lead_count = sum(1 for p in profiles if getattr(p, 'classification', None) == "Lead")
        
        # Average scores with null handling
        scores = [getattr(p, 'composite_score', None) for p in profiles]
        valid_scores = [s for s in scores if s is not None]
        avg_score = sum(valid_scores) / len(valid_scores) if valid_scores else 0
        min_score = min(valid_scores) if valid_scores else 0
        max_score = max(valid_scores) if valid_scores else 0
        synthetic_cnt = sum(1 for p in profiles if getattr(p, 'data_source_type', '') == 'synthetic')
        real_cnt = total - synthetic_cnt

        # Create summary data
        summary_data = [
            ["Metric", "Value"],
            ["Total Companies", total],
            ["Phoenix Companies", f"{phoenix_count} ({self._format_percentage(phoenix_count/total*100 if total else 0, 0)})"],
            ["Salt Companies", f"{salt_count} ({self._format_percentage(salt_count/total*100 if total else 0, 0)})"],
            ["Lead Companies", f"{lead_count} ({self._format_percentage(lead_count/total*100 if total else 0, 0)})"],
            ["Average Composite Score", self._format_number(avg_score, 2)],
            ["Score Range", f"{min_score:.2f} – {max_score:.2f}" if valid_scores else "N/A"],
            ["Synthetic Companies", f"{synthetic_cnt} ({self._format_percentage(synthetic_cnt/total*100 if total else 0, 0)})"],
            ["Real Companies", f"{real_cnt}"],
            ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ]
        
        # Write summary
        start_row = self.layout.DATA_START_ROW
        for row_idx, row_data in enumerate(summary_data, start=start_row):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.styles.data_font
                
                if row_idx == start_row:  # Header row
                    cell.font = self.styles.header_font
                    cell.fill = self.styles.header_fill
                    cell.alignment = self.styles.header_alignment
                else:
                    cell.fill = self.styles.get_row_fill(row_idx)
                    cell.alignment = self.styles.data_alignment if col_idx == 1 else self.styles.number_alignment
        
        # Set column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
    
    def _add_market_rankings(self, ws: Any, profiles: list[Company]) -> None:
        """Market Rankings sheet with consistent headers."""
        self._add_title_banner(ws, "Market Rankings", "Companies ranked by composite score")
        
        # Headers that match the data structure
        headers = [
            "Rank", "Company", "Classification", "Composite Score",
            "Growth Score", "Financial Health", "Competitive Position",
            "Revenue (M)", "Growth Rate (%)", "Employees", "Data Source"
        ]
        self._write_headers(ws, headers)
        
        # Sort profiles by composite score (descending) with null handling
        sorted_profiles = sorted(
            profiles,
            key=lambda p: getattr(p, 'composite_score', 0) or 0,
            reverse=True
        )
        
        # Write data rows
        for rank, profile in enumerate(sorted_profiles, 1):
            row_idx = self.layout.DATA_START_ROW + rank - 1
            classification = getattr(profile, 'classification', 'Unknown')
            
            row_data = [
                rank,
                self._safe_get(profile, 'name', 'Unknown'),
                classification,
                self._format_number(getattr(profile, 'composite_score', None), 2),
                self._format_number(getattr(profile, 'growth_score', None), 2),
                self._format_number(getattr(profile, 'financial_health_score', None), 2),
                self._format_number(getattr(profile, 'competitive_position_score', None), 2),
                self._format_number(self._safe_get_financial(profile, 'revenue'), 1),
                self._format_percentage(self._safe_get_financial(profile, 'growth_rate')),
                self._safe_get_financial(profile, 'employees', 0),
                getattr(profile, 'data_source_type', 'unknown').capitalize(),
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.styles.data_font
                cell.fill = self.styles.get_row_fill(row_idx, classification)
                cell.border = self.styles.standard_border
                cell.alignment = self.styles.number_alignment if col_idx > 2 else self.styles.data_alignment
    
    def _add_financial_intelligence(self, ws: Any, profiles: list[Company]) -> None:
        """Financial Intelligence sheet."""
        self._add_title_banner(ws, "Financial Intelligence", "Detailed financial metrics")
        
        headers = [
            "Company", "Revenue (M)", "Revenue Confidence", "Growth Rate (%)",
            "Growth Confidence", "Employees", "Employee Confidence",
            "Funding (M)", "Funding Confidence", "Valuation (M)",
            "Valuation Confidence", "Profit Margin (%)", "EBITDA Margin (%)"
        ]
        self._write_headers(ws, headers)
        
        for idx, profile in enumerate(profiles):
            row_idx = self.layout.DATA_START_ROW + idx
            
            # Get confidence levels as strings
            rev_conf = self._safe_get_financial(profile, 'revenue_confidence')
            rev_conf_str = rev_conf.value if hasattr(rev_conf, 'value') else str(rev_conf) if rev_conf else "Unknown"
            
            growth_conf = self._safe_get_financial(profile, 'growth_confidence')
            growth_conf_str = growth_conf.value if hasattr(growth_conf, 'value') else str(growth_conf) if growth_conf else "Unknown"
            
            emp_conf = self._safe_get_financial(profile, 'employees_confidence')
            emp_conf_str = emp_conf.value if hasattr(emp_conf, 'value') else str(emp_conf) if emp_conf else "Unknown"
            
            fund_conf = self._safe_get_financial(profile, 'funding_confidence')
            fund_conf_str = fund_conf.value if hasattr(fund_conf, 'value') else str(fund_conf) if fund_conf else "Unknown"
            
            val_conf = self._safe_get_financial(profile, 'valuation_confidence')
            val_conf_str = val_conf.value if hasattr(val_conf, 'value') else str(val_conf) if val_conf else "Unknown"
            
            row_data = [
                self._safe_get(profile, 'name', 'Unknown'),
                self._format_number(self._safe_get_financial(profile, 'revenue'), 1),
                rev_conf_str,
                self._format_percentage(self._safe_get_financial(profile, 'growth_rate')),
                growth_conf_str,
                self._safe_get_financial(profile, 'employees', 0),
                emp_conf_str,
                self._format_number(self._safe_get_financial(profile, 'funding_raised'), 1),
                fund_conf_str,
                self._format_number(self._safe_get_financial(profile, 'valuation'), 1),
                val_conf_str,
                self._format_percentage(self._safe_get_financial(profile, 'profit_margin')),
                self._format_percentage(self._safe_get_financial(profile, 'ebitda_margin')),
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.styles.data_font
                cell.fill = self.styles.get_row_fill(row_idx)
                cell.border = self.styles.standard_border
                cell.alignment = self.styles.number_alignment if col_idx > 1 else self.styles.data_alignment
    
    def _add_tech_ai_maturity(self, ws: Any, profiles: list[Company]) -> None:
        """Tech & AI Maturity sheet."""
        self._add_title_banner(ws, "Tech & AI Maturity", "Technology and AI capabilities assessment")
        
        headers = [
            "Company", "AI Maturity", "SaaS Maturity", "Tech Stack",
            "AI in Production", "Key Capabilities"
        ]
        self._write_headers(ws, headers)
        
        for idx, profile in enumerate(profiles):
            row_idx = self.layout.DATA_START_ROW + idx
            
            ai_maturity = self._safe_get(profile, 'ai_maturity', 'Unknown')
            ai_maturity_str = ai_maturity.value if hasattr(ai_maturity, 'value') else str(ai_maturity)
            
            tech_stack = self._safe_get(profile, 'tech_stack', [])
            tech_stack_str = ", ".join(tech_stack) if tech_stack else "N/A"
            
            row_data = [
                self._safe_get(profile, 'name', 'Unknown'),
                ai_maturity_str,
                self._safe_get(profile, 'saas_maturity', 0),
                tech_stack_str,
                "Yes" if self._safe_get(profile, 'ai_in_production', False) else "No",
                self._safe_get(profile, 'ai_key_capabilities', 'N/A'),
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.styles.data_font
                cell.fill = self.styles.get_row_fill(row_idx)
                cell.border = self.styles.standard_border
                cell.alignment = self.styles.data_alignment
                
                # Wrap text for capabilities column
                if col_idx == 6:
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
        
        # Set wider column for capabilities
        ws.column_dimensions["F"].width = self.layout.DESCRIPTION_COLUMN_WIDTH
    
    def _add_company_details(self, ws: Any, profiles: list[Company]) -> None:
        """Company Details sheet with complete information."""
        self._add_title_banner(ws, "Company Details", "Complete company profiles and metadata")
        
        headers = [
            "Company", "Website", "Headquarters", "Founded", "Industry",
            "Description", "Enrichment Sources", "Data Quality Tier", "Last Updated"
        ]
        self._write_headers(ws, headers)
        
        for idx, profile in enumerate(profiles):
            row_idx = self.layout.DATA_START_ROW + idx
            
            # Format enrichment sources
            enrichment_count = self._safe_get(profile, 'enrichment_source_count', 0)
            
            # Format last updated
            last_updated = self._safe_get(profile, 'last_updated', None)
            last_updated_str = last_updated.strftime("%Y-%m-%d") if hasattr(last_updated, 'strftime') else str(last_updated) if last_updated else "N/A"
            
            row_data = [
                self._safe_get(profile, 'name', 'Unknown'),
                self._safe_get(profile, 'website', ''),
                self._safe_get(profile, 'headquarters', ''),
                self._safe_get(profile, 'founded_year', ''),
                self._safe_get(profile, 'industry', ''),
                (self._safe_get(profile, 'description', '') or '')[:200],  # Truncate long descriptions
                enrichment_count,
                self._safe_get(profile, 'data_quality_tier', 'unknown'),
                last_updated_str,
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.styles.data_font
                cell.fill = self.styles.get_row_fill(row_idx)
                cell.border = self.styles.standard_border
                cell.alignment = self.styles.data_alignment
                
                # Wrap text for description
                if col_idx == 6:
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
    
    def _auto_adjust_columns(self, ws: Any) -> None:
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            if not column:
                continue
                
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except (ValueError, TypeError):
                    continue
            
            # Apply width with constraints
            adjusted_width = min(
                max(max_length + 2, self.layout.MIN_COLUMN_WIDTH),
                self.layout.MAX_COLUMN_WIDTH
            )
            ws.column_dimensions[column_letter].width = adjusted_width


# Backwards compatibility: alias for the old class name
ExcelExporter = ImprovedExcelExporter
