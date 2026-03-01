"""
Excel exporter for SolStein dashboards.

Replaces the monolithic generate_excel_report.py script.
"""

from datetime import datetime
from pathlib import Path
from typing import Any

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..config import Settings
from ..domain.models import Company


class ExcelExporter:
    """Export functionality for SolStein data."""

    def __init__(self, template_path: Path | None = None):
        self.settings = Settings()
        self.template_path = template_path or (self.settings.data.data_dir / "templates" / "dashboard_template.xlsx")
        self.styles: dict[str, Any] = self._create_styles()

    def _create_styles(self) -> dict[str, Any]:
        """Create consistent master-grade styles for the intelligence report."""
        # Solstein Alchemical Palette
        colors = {
            "obsidian": "0A0A0F",  # Deep dark
            "onyx": "12121A",  # Background
            "gold": "D4A843",  # Accents
            "gold_light": "F0D78C",  # Highlight
            "emerald": "2ECC71",  # Phoenix/Positive
            "ruby": "E74C3C",  # Lead/Negative
            "sapphire": "3498DB",  # Info
            "slate": "1A1A2E",  # Borders/Subheaders
            "text": "E8E8F0",  # Light text
            "text_muted": "8888A0",  # Dimmed text
        }

        # Borders
        standard_border = Border(
            left=Side(style="thin", color=colors["slate"]),
            right=Side(style="thin", color=colors["slate"]),
            top=Side(style="thin", color=colors["slate"]),
            bottom=Side(style="thin", color=colors["slate"]),
        )

        thick_bottom = Border(bottom=Side(style="medium", color=colors["gold"]))

        return {
            "title_font": Font(name="Calibri", size=18, bold=True, color="FFFFFF"),
            "header_font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "header_fill": PatternFill(
                start_color=colors["obsidian"],
                end_color=colors["obsidian"],
                fill_type="solid",
            ),
            "header_alignment": Alignment(horizontal="center", vertical="center"),
            "header_border": thick_bottom,
            "subheader_font": Font(name="Calibri", size=11, bold=True, color=colors["gold"]),
            "subheader_fill": PatternFill(
                start_color=colors["slate"],
                end_color=colors["slate"],
                fill_type="solid",
            ),
            "subheader_alignment": Alignment(horizontal="left", vertical="center"),
            "data_font": Font(name="Calibri", size=10),
            "data_fill_even": PatternFill(start_color="F9F9FB", end_color="F9F9FB", fill_type="solid"),
            "data_fill_odd": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
            "data_alignment": Alignment(horizontal="left", vertical="center", indent=1),
            "data_border": standard_border,
            "number_alignment": Alignment(horizontal="right", vertical="center", indent=1),
            "phoenix_fill": PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),  # Light Emerald
            "lead_fill": PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"),  # Light Ruby
            "salt_fill": PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid"),  # Light Yellow
        }

    def create_dashboard(self, profiles: list[Company], output_path: Path) -> None:
        """Create a multi-sheet Professional Intelligence Report."""
        logger.info(f"Aura | Generating Professional Glyph Report for {len(profiles)} companies")

        wb = Workbook()

        # 1. Executive Summary
        ws_summary = wb.active
        if ws_summary is None:
            ws_summary = wb.create_sheet("Executive Summary")
        else:
            ws_summary.title = "Executive Summary"
        self._add_executive_summary(ws_summary, profiles)

        # 2. Market Rankings (Attractiveness Board)
        ws_rankings = wb.create_sheet("Market Rankings")
        self._add_market_rankings(ws_rankings, profiles)

        # 3. Financial Intelligence
        ws_financials = wb.create_sheet("Financial Intelligence")
        self._add_financial_intelligence(ws_financials, profiles)

        # 4. Tech & AI Maturity
        ws_tech = wb.create_sheet("Tech & AI Maturity")
        self._add_tech_maturity(ws_tech, profiles)

        # 5. Company Details — 20 fields missing from other sheets (EPIC-033)
        ws_details = wb.create_sheet("Company Details")
        self._add_company_details(ws_details, profiles)

        # Auto-adjust column widths for all sheets
        for sheet in wb.worksheets:
            self._auto_adjust_columns(sheet)

        # Save workbook
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info(f"Aura | Glyph Report saved to {output_path}")

    def _add_top_banner(self, ws: Any, title: str, subtitle: str) -> None:
        """Add a professional Solstein banner to the top of a worksheet."""
        # Main Title
        ws["A1"] = f"SOLSTEIN | {title.upper()}"
        ws["A1"].font = self.styles["title_font"]
        ws["A1"].fill = self.styles["header_fill"]
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells("A1:G1")
        ws.row_dimensions[1].height = 40

        # Subtitle/Timestamp
        ws["A2"] = f"{subtitle} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font = Font(name="Calibri", size=9, color="8888A0", italic=True)
        ws.merge_cells("A2:G2")

    def _write_headers(self, ws: Any, headers: list[str], row: int) -> None:
        """Helper to write standardized headers."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles["header_font"]
            cell.fill = self.styles["header_fill"]
            cell.alignment = self.styles["header_alignment"]
            cell.border = self.styles["header_border"]

    def _add_executive_summary(self, ws: Any, profiles: list[Company]) -> None:
        """Add high-level market metrics and distribution summary."""
        self._add_top_banner(ws, "Executive Summary", "Market Composition & Growth Analytics")

        # 1. Core KPIs
        kpis = [
            ("Total Portfolio Size", f"{len(profiles)} Companies"),
            (
                "Market Leaders (Phoenix)",
                f"{len([p for p in profiles if p.classification == 'Phoenix'])}",
            ),
            (
                "Stable Assets (Salt)",
                f"{len([p for p in profiles if p.classification == 'Salt'])}",
            ),
            (
                "Underperformers (Lead)",
                f"{len([p for p in profiles if p.classification == 'Lead'])}",
            ),
            (
                "Avg Growth Score",
                f"{sum(p.growth_score or 0 for p in profiles) / len(profiles):.1f}" if profiles else "0.0",
            ),
        ]

        row = 4
        ws.cell(row=row, column=1, value="MARKET CORE KPIS").font = self.styles["subheader_font"]
        row += 1
        for label, value in kpis:
            ws.cell(row=row, column=1, value=label).font = self.styles["data_font"]
            ws.cell(row=row, column=2, value=value).font = Font(bold=True)
            row += 1

        # 2. Top Movers (Placeholder for chart or list)
        row += 2
        ws.cell(row=row, column=1, value="STRATEGIC RECOMMENDATIONS").font = self.styles["subheader_font"]
        row += 1
        recommendations = [
            "• Immediate focus on Phoenix-tier entities for scale expansion.",
            "• Strategic audit required for Lead-tier companies with high tech debt.",
            "• Monitor Salt-tier players for directional growth signals.",
        ]
        for rec in recommendations:
            ws.cell(row=row, column=1, value=rec).font = self.styles["data_font"]
            ws.merge_cells(f"A{row}:D{row}")
            row += 1

    def _add_market_rankings(self, ws: Any, profiles: list[Company]) -> None:
        """Add the primary ranked company list (The Attractiveness Board)."""
        self._add_top_banner(ws, "Market Rankings", "Corporate Attractiveness Board")

        headers = [
            "Rank",
            "Company",
            "Classification",
            "Growth",
            "Financial",
            "Competitive",
            "Composite",
            "Tier",
            "Threat",
        ]
        self._write_headers(ws, headers, row=4)

        sorted_profiles = sorted(profiles, key=lambda p: p.composite_score or 0, reverse=True)

        for i, p in enumerate(sorted_profiles, 5):
            fill = self.styles["data_fill_even"] if i % 2 == 0 else self.styles["data_fill_odd"]

            # Apply classification highlight
            if p.classification == "Phoenix":
                fill = self.styles["phoenix_fill"]
            elif p.classification == "Lead":
                fill = self.styles["lead_fill"]
            elif p.classification == "Salt":
                fill = self.styles["salt_fill"]

            data = [
                i - 4,
                p.name,
                p.classification or "Salt",
                f"{p.growth_score:.1f}" if p.growth_score else "N/A",
                f"{p.financial_health_score:.1f}" if p.financial_health_score else "N/A",
                f"{p.competitive_position_score:.1f}" if p.competitive_position_score else "N/A",
                f"{p.composite_score:.1f}" if p.composite_score else "N/A",
                p.tier.value if p.tier else "N/A",
                p.threat_level.value if p.threat_level else "N/A",
            ]

            for col, val in enumerate(data, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.font = self.styles["data_font"]
                cell.fill = fill
                cell.border = self.styles["data_border"]
                cell.alignment = self.styles["number_alignment"] if col != 2 else self.styles["data_alignment"]

    def _add_financial_intelligence(self, ws: Any, profiles: list[Company]) -> None:
        """Detailed financial metrics and growth timelines."""
        self._add_top_banner(ws, "Financial Intelligence", "Scalability & Capital Health Analysis")

        headers = [
            "Company",
            "Revenue (€M)",
            "Growth (%)",
            "CAGR 3yr",
            "Margin (%)",
            "EBITDA (%)",
            "Recurring (%)",
            "Funding (€M)",
            "Valuation (€M)",
        ]
        self._write_headers(ws, headers, row=4)

        for i, p in enumerate(profiles, 5):
            fill = self.styles["data_fill_even"] if i % 2 == 0 else self.styles["data_fill_odd"]

            data = [
                p.name,
                f"{p.financials.revenue:.1f}" if p.financials.revenue else "N/A",
                f"{p.financials.growth_rate:.1f}%" if p.financials.growth_rate else "N/A",
                f"{p.revenue_cagr_3yr:.1f}%" if p.revenue_cagr_3yr else "N/A",
                f"{p.profit_margin:.1f}%" if p.profit_margin else "N/A",
                f"{p.ebitda_margin:.1f}%" if p.ebitda_margin else "N/A",
                f"{p.recurring_revenue_pct:.0f}%" if p.recurring_revenue_pct is not None else "N/A",
                f"{p.total_funding_raised_eur / 1e6:.1f}M" if p.total_funding_raised_eur else "Bootstrapped",
                f"{p.latest_valuation_eur / 1e6:.1f}M" if p.latest_valuation_eur else "Undisclosed",
            ]

            for col, val in enumerate(data, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.font = self.styles["data_font"]
                cell.fill = fill
                cell.border = self.styles["data_border"]
                cell.alignment = self.styles["number_alignment"] if col != 1 else self.styles["data_alignment"]

    def _add_tech_maturity(self, ws: Any, profiles: list[Company]) -> None:
        """Deep dive into AI maturity and technology stack signal strength."""
        self._add_top_banner(ws, "Tech & AI Maturity", "Intelligence Signaling & Architectural Moats")

        headers = [
            "Company",
            "AI Score",
            "AI Signal",
            "In Production",
            "SaaS Maturity",
            "Key AI Capabilities",
            "Moat Strength",
        ]
        self._write_headers(ws, headers, row=4)

        for i, p in enumerate(profiles, 5):
            fill = self.styles["data_fill_even"] if i % 2 == 0 else self.styles["data_fill_odd"]

            moat = "Low"
            if p.ai_score and p.ai_score > 7:
                moat = "Critical" if p.ai_in_production else "High"
            elif p.saas_maturity > 8:
                moat = "Medium"

            data = [
                p.name,
                p.ai_score if p.ai_score is not None else "N/A",
                p.ai_signal_level or "Unknown",
                "YES" if p.ai_in_production else "NO",
                p.saas_maturity,
                p.ai_key_capabilities or "Generic SaaS",
                moat,
            ]

            for col, val in enumerate(data, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.font = self.styles["data_font"]
                cell.fill = fill
                cell.border = self.styles["data_border"]
                cell.alignment = Alignment(wrap_text=True) if col == 6 else self.styles["data_alignment"]
                if col != 1 and col != 6:
                    cell.alignment = self.styles["number_alignment"]

        # Set column width for AI capabilities
        ws.column_dimensions["F"].width = 40

    def _add_company_details(self, ws: Any, profiles: list[Company]) -> None:
        """Company Details sheet — 20 fields not covered in other sheets (EPIC-033)."""
        self._add_top_banner(ws, "Company Details", "Full profile including identifiers and metadata")
        headers = [
            "Company", "Website", "Headquarters", "Founded Year", "Employees",
            "Ticker", "Company Number", "ISIN", "Geography Code", "Industry",
            "Sub Sector", "AI Maturity", "Confidence Level", "Signal Count",
            "Source Count", "Data Quality", "Competitive Positioning",
            "Comp. Overlap Count", "Last Updated", "Description",
        ]
        self._write_headers(ws, headers, start_row=5)

        for row_idx, p in enumerate(profiles, start=6):
            # Source / signal counts
            source_count = len(p.sources) if hasattr(p, "sources") and p.sources else 0
            signal_count = getattr(p, "signal_count", 0)

            # Data quality — compute from confidence if not directly available
            dq_raw = getattr(p, "data_quality", None)
            data_quality = f"{dq_raw:.0%}" if isinstance(dq_raw, float) else "N/A"

            # Confidence level
            conf = getattr(p, "confidence_level", None)
            confidence = conf.value if hasattr(conf, "value") else str(conf) if conf else "N/A"

            # Competitive positioning / overlap
            comp_pos = getattr(p, "competitive_positioning", "")
            overlaps = getattr(p, "competitive_overlaps", [])
            overlap_count = len(overlaps) if overlaps else 0

            # Sub-sector
            sub_sector = getattr(p, "sub_sector", "") or ""

            row = [
                p.name,
                getattr(p, "website", "") or "",
                getattr(p, "headquarters", "") or "",
                getattr(p, "founded_year", "") or "",
                getattr(p, "employees", "") or "",
                getattr(p, "ticker", "") or "",
                getattr(p, "company_number", "") or "",
                getattr(p, "isin", "") or "",
                getattr(p, "geography_code", "") or "",
                getattr(p, "industry", "") or "",
                sub_sector,
                str(getattr(p, "ai_maturity", "") or ""),
                confidence,
                signal_count,
                source_count,
                data_quality,
                comp_pos or "",
                overlap_count,
                getattr(p, "updated_at", "") or "",
                (getattr(p, "description", "") or "")[:200],  # truncate long descriptions
            ]
            for col, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col, value=value)

    def _auto_adjust_columns(self, ws: Any) -> None:
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            # Skip F on Tech sheet as we set it manually
            if ws.title == "Tech & AI Maturity" and column_letter == "F":
                continue

            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (ValueError, TypeError):
                    continue

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width


class TemplateExporter(ExcelExporter):
    """Export using a template file."""

    def create_dashboard(self, profiles: list[Company], output_path: Path) -> None:
        """Create dashboard using a template (Not Implemented)."""
        logger.error("Template-based export is not yet implemented.")
        raise NotImplementedError(
            "Template-based export requires mapping profiles to specific template cells. "  # noqa: E501
            "Use standard ExcelExporter for generated dashboards."
        )
