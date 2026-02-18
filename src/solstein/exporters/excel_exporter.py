"""
Excel exporter for SolStein dashboards.

Replaces the monolithic generate_excel_report.py script.
"""

from pathlib import Path
from typing import List, Optional, Dict, Any
from datetime import datetime
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, Series
from loguru import logger

from ..data.models import CompanyProfile


class ExcelExporter:
    """Export company profiles to Excel dashboard."""
    
    def __init__(self, template_path: Optional[Path] = None):
        self.template_path = template_path
        self.styles = self._create_styles()
    
    def _create_styles(self) -> Dict[str, Any]:
        """Create consistent styles for the dashboard."""
        # Colors
        colors = {
            "header": "366092",  # Dark blue
            "subheader": "4F81BD",  # Medium blue
            "data_even": "DCE6F1",  # Light blue
            "data_odd": "FFFFFF",  # White
            "positive": "C6EFCE",  # Green
            "negative": "FFC7CE",  # Red
            "warning": "FFEB9C",  # Yellow
        }
        
        # Borders
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        return {
            "header_font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "header_fill": PatternFill(start_color=colors["header"], end_color=colors["header"], fill_type="solid"),
            "header_alignment": Alignment(horizontal="center", vertical="center"),
            
            "subheader_font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "subheader_fill": PatternFill(start_color=colors["subheader"], end_color=colors["subheader"], fill_type="solid"),
            "subheader_alignment": Alignment(horizontal="center", vertical="center"),
            
            "data_font": Font(name="Calibri", size=11),
            "data_fill_even": PatternFill(start_color=colors["data_even"], end_color=colors["data_even"], fill_type="solid"),
            "data_fill_odd": PatternFill(start_color=colors["data_odd"], end_color=colors["data_odd"], fill_type="solid"),
            "data_alignment": Alignment(horizontal="left", vertical="center"),
            "data_border": thin_border,
            
            "number_font": Font(name="Calibri", size=11),
            "number_alignment": Alignment(horizontal="right", vertical="center"),
            
            "positive_fill": PatternFill(start_color=colors["positive"], end_color=colors["positive"], fill_type="solid"),
            "negative_fill": PatternFill(start_color=colors["negative"], end_color=colors["negative"], fill_type="solid"),
            "warning_fill": PatternFill(start_color=colors["warning"], end_color=colors["warning"], fill_type="solid"),
        }
    
    def create_dashboard(self, profiles: List[CompanyProfile], output_path: Path) -> None:
        """Create Excel dashboard from company profiles."""
        logger.info(f"Creating dashboard with {len(profiles)} companies")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Competitive Dashboard"
        
        # Add metadata
        self._add_metadata(ws, len(profiles))
        
        # Add summary table
        self._add_summary_table(ws, profiles, start_row=5)
        
        # Add detailed analysis
        detail_start = len(profiles) + 8
        self._add_detailed_analysis(ws, profiles, start_row=detail_start)
        
        # Add charts
        chart_start = detail_start + len(profiles) + 5
        self._add_charts(ws, profiles, start_row=chart_start)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        # Save workbook
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info(f"Dashboard saved to {output_path}")
    
    def _add_metadata(self, ws, company_count: int) -> None:
        """Add metadata section to worksheet."""
        ws["A1"] = "SolStein Competitive Intelligence Dashboard"
        ws["A1"].font = self.styles["header_font"]
        ws["A1"].fill = self.styles["header_fill"]
        ws["A1"].alignment = self.styles["header_alignment"]
        ws.merge_cells("A1:H1")
        
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(name="Calibri", size=10, italic=True)
        
        ws["A3"] = f"Companies Analyzed: {company_count}"
        ws["A3"].font = Font(name="Calibri", size=10, bold=True)
    
    def _add_summary_table(self, ws, profiles: List[CompanyProfile], start_row: int) -> None:
        """Add summary table to worksheet."""
        # Table headers
        headers = [
            "Company",
            "Tier",
            "Revenue (€M)",
            "Growth Rate (%)",
            "Employees",
            "AI Maturity",
            "Threat Level",
            "Growth Score",
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = self.styles["subheader_font"]
            cell.fill = self.styles["subheader_fill"]
            cell.alignment = self.styles["subheader_alignment"]
            cell.border = self.styles["data_border"]
        
        # Write data rows
        for row_idx, profile in enumerate(profiles, start_row + 1):
            # Alternate row colors
            fill = self.styles["data_fill_even"] if row_idx % 2 == 0 else self.styles["data_fill_odd"]
            
            # Format revenue in millions
            revenue = profile.financials.revenue
            revenue_str = f"{revenue/1_000_000:.1f}" if revenue else "N/A"
            
            # Format growth rate
            growth = profile.financials.growth_rate
            growth_str = f"{growth:.1f}%" if growth else "N/A"
            
            # Format employees
            employees = profile.financials.employees
            employees_str = f"{employees:,}" if employees else "N/A"
            
            # Data cells
            data = [
                profile.name,
                profile.tier,
                revenue_str,
                growth_str,
                employees_str,
                profile.ai_maturity,
                profile.threat_level,
                profile.growth_score or "N/A",
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.styles["data_font"]
                cell.fill = fill
                cell.alignment = self.styles["data_alignment"] if col_idx == 1 else self.styles["number_alignment"]
                cell.border = self.styles["data_border"]
                
                # Conditional formatting for growth rate
                if col_idx == 4 and growth:
                    if growth > 20:
                        cell.fill = self.styles["positive_fill"]
                    elif growth < 0:
                        cell.fill = self.styles["negative_fill"]
                
                # Conditional formatting for threat level
                if col_idx == 7:
                    if profile.threat_level == "High":
                        cell.fill = self.styles["negative_fill"]
                    elif profile.threat_level == "Critical":
                        cell.font = Font(name="Calibri", size=11, bold=True, color="FF0000")
    
    def _add_detailed_analysis(self, ws, profiles: List[CompanyProfile], start_row: int) -> None:
        """Add detailed analysis section."""
        # Section header
        ws.cell(row=start_row, column=1, value="Detailed Financial Analysis")
        ws.cell(row=start_row, column=1).font = self.styles["subheader_font"]
        ws.cell(row=start_row, column=1).fill = self.styles["subheader_fill"]
        ws.merge_cells(f"A{start_row}:H{start_row}")
        
        # Detailed headers
        detail_headers = [
            "Company",
            "Revenue (€)",
            "Profit Margin (%)",
            "Funding Raised (€)",
            "Valuation (€)",
            "SaaS Maturity",
            "Geographic Presence",
            "Key Customers",
        ]
        
        detail_start = start_row + 1
        for col_idx, header in enumerate(detail_headers, 1):
            cell = ws.cell(row=detail_start, column=col_idx, value=header)
            cell.font = self.styles["subheader_font"]
            cell.fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
            cell.alignment = self.styles["subheader_alignment"]
            cell.border = self.styles["data_border"]
        
        # Detailed data
        for row_idx, profile in enumerate(profiles, detail_start + 1):
            fill = self.styles["data_fill_even"] if row_idx % 2 == 0 else self.styles["data_fill_odd"]
            
            # Format large numbers with commas
            revenue = profile.financials.revenue
            revenue_str = f"€{revenue:,.0f}" if revenue else "N/A"
            
            margin = profile.financials.profit_margin
            margin_str = f"{margin:.1f}%" if margin else "N/A"
            
            funding = profile.financials.funding_raised
            funding_str = f"€{funding:,.0f}" if funding else "N/A"
            
            valuation = profile.financials.valuation
            valuation_str = f"€{valuation:,.0f}" if valuation else "N/A"
            
            # Truncate geographic presence and customers
            geo_presence = ", ".join(profile.geographic_presence[:3])
            if len(profile.geographic_presence) > 3:
                geo_presence += "..."
            
            customers = ", ".join(profile.key_customers[:2])
            if len(profile.key_customers) > 2:
                customers += "..."
            
            data = [
                profile.name,
                revenue_str,
                margin_str,
                funding_str,
                valuation_str,
                profile.saas_maturity,
                geo_presence,
                customers or "N/A",
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.styles["data_font"]
                cell.fill = fill
                cell.alignment = self.styles["data_alignment"] if col_idx in [1, 7, 8] else self.styles["number_alignment"]
                cell.border = self.styles["data_border"]
    
    def _add_charts(self, ws, profiles: List[CompanyProfile], start_row: int) -> None:
        """Add charts to the dashboard."""
        # Revenue comparison chart
        ws.cell(row=start_row, column=1, value="Revenue Comparison (Top 10)")
        ws.cell(row=start_row, column=1).font = Font(name="Calibri", size=12, bold=True)
        
        # Sort by revenue and take top 10
        sorted_profiles = sorted(
            [p for p in profiles if p.financials.revenue],
            key=lambda p: p.financials.revenue or 0,
            reverse=True
        )[:10]
        
        if sorted_profiles:
            # Prepare data for chart
            chart_data_start = start_row + 2
            for i, profile in enumerate(sorted_profiles):
                ws.cell(row=chart_data_start + i, column=1, value=profile.name)
                ws.cell(row=chart_data_start + i, column=2, value=profile.financials.revenue or 0)
            
            # Create bar chart
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Revenue Comparison"
            chart.y_axis.title = "Revenue (€)"
            chart.x_axis.title = "Company"
            
            data = Reference(ws, min_col=2, min_row=chart_data_start, max_row=chart_data_start + len(sorted_profiles) - 1)
            categories = Reference(ws, min_col=1, min_row=chart_data_start, max_row=chart_data_start + len(sorted_profiles) - 1)
            
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(categories)
            
            ws.add_chart(chart, f"D{start_row}")
    
    def _auto_adjust_columns(self, ws) -> None:
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width


class TemplateExporter(ExcelExporter):
    """Export using a template file."""
    
    def create_dashboard(self, profiles: List[CompanyProfile], output_path: Path) -> None:
        """Create dashboard using a template."""
        if not self.template_path:
            raise ValueError("Template path required for TemplateExporter")
        
        logger.info(f"Using template: {self.template_path}")
        
        # Load template
        wb = Workbook()
        wb = wb.load_workbook(self.template_path)
        
        # TODO: Populate template with data
        # This would involve mapping profiles to specific cells in the template
        
        # Save output
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info(f"Dashboard saved to {output_path}")